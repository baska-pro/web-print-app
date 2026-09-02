#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PrintBot - WebApp Local Printer Edition
=======================================
File manager + print console berbasis web, responsif (mobile & desktop).
File tunggal dan mandiri: HTML/CSS/JS dirender inline, dependency Python
di-install otomatis, dan seluruh proses print memakai printer yang TERPASANG
DI PC/HOST ini - tidak memakai Google Apps Script, Print Bridge, atau relay.

Antrean print disimpan secara persisten di SQLite (`data/printbot.db`).
Worker lokal memproses job BERURUTAN supaya PDF/Office tidak saling tumpang
tindih. Pada Windows, engine memakai pywin32 + GDI + Microsoft Office COM
(Word/Excel/PowerPoint) serta memantau Windows Spooler/Job ID secara nyata.

Fitur print utama:
  - deteksi printer lokal + default printer
  - A4/F4/Legal/Letter, portrait/landscape, Fit/Actual
  - duplex dan color/mono bila driver mendukung
  - PDF/gambar menjadi satu Windows spool document per job
  - Word/Excel/PowerPoint melalui Office COM dengan COM threading aman
  - status Offline/Paper Out/Paper Jam/Error + Windows Job ID
  - cancel job yang sudah masuk Windows Spooler
  - recovery queue setelah restart tanpa duplicate print
  - Task Scheduler autostart Windows (pythonw.exe) + single instance

Konfigurasi `.env` di folder yang sama:
    FILE_MANAGER_ROOTS=Label:F:\\Data;Label2:D:\\Arsip
    DATA_DIR=./data
    DEFAULT_PRINTER=
    PRINT_JOB_TIMEOUT_SEC=900
    PRINT_RENDER_DPI=144
    PRINT_MAX_RETRIES=1
    AUTO_START_TASK=1
    TASK_NAME=PrintBot-WebApp
    MAX_UPLOAD_MB=50
    MAX_DOWNLOAD_MB=50
    WEB_ITEMS_PER_PAGE=20
    SEARCH_MAX_RESULTS=40
    WEBAPP_HOST=0.0.0.0
    WEBAPP_PORT=8000
    WEBAPP_USERNAME=admin
    WEBAPP_PASSWORD=...
    WEBAPP_SECRET_KEY=...
    LOG_LEVEL=INFO

Jalankan langsung: python webapp-PrintBot.py
"""
from __future__ import annotations

import importlib.util
import os
import subprocess
import sys

# ============================================================================
# BOOTSTRAP - cek & install dependency Python yang belum ada (stdlib only,
# dijalankan SEBELUM import library pihak ketiga di bawah).
# ============================================================================
REQUIRED_PACKAGES = {
    "dotenv": "python-dotenv",
    "fastapi": "fastapi",
    "uvicorn": "uvicorn[standard]",
    "starlette": "starlette",
    "itsdangerous": "itsdangerous",
    "multipart": "python-multipart",
    "pydantic": "pydantic",
    "openpyxl": "openpyxl",
    "PIL": "Pillow",
    "fitz": "PyMuPDF",
    "docx": "python-docx",
}

# Python 3.8 masih umum dipakai pada Windows 7. Gunakan versi dependency
# yang masih menyediakan wheel Python 3.8 agar bootstrap tidak menarik versi
# modern yang sudah menghentikan dukungan.
if sys.version_info < (3, 9):
    REQUIRED_PACKAGES.update({
        "dotenv": "python-dotenv<1.1", "fastapi": "fastapi<0.116",
        "uvicorn": "uvicorn[standard]<0.31", "starlette": "starlette<0.42",
        "pydantic": "pydantic<2.11", "openpyxl": "openpyxl<3.2",
        "PIL": "Pillow<11", "fitz": "PyMuPDF<1.25", "docx": "python-docx<1.2",
    })

# pywin32 hanya valid di Windows. Modul ini memberi akses printer, GDI,
# Task Scheduler COM, dan Windows Spooler.
if sys.platform.startswith("win"):
    REQUIRED_PACKAGES["win32print"] = "pywin32==306" if sys.version_info < (3, 9) else "pywin32"


def ensure_dependencies(quiet: bool = False) -> None:
    missing = []
    for mod_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            found = importlib.util.find_spec(mod_name) is not None
        except (ImportError, ValueError):
            found = False
        if not found:
            missing.append(pip_name)
    if not missing:
        return
    if not quiet:
        print(f"[bootstrap] Dependency belum lengkap, menginstall: {', '.join(missing)}")
    cmd = [sys.executable, "-m", "pip", "install", "--disable-pip-version-check", "-q", *missing]
    try:
        subprocess.check_call(cmd)
    except subprocess.CalledProcessError as exc:
        print(f"[bootstrap] ERROR: gagal install dependency otomatis ({exc}).")
        print(f"[bootstrap] Jalankan manual: pip install {' '.join(missing)}")
        sys.exit(1)
    if not quiet:
        print("[bootstrap] Dependency siap.")


ensure_dependencies()

# ============================================================================
# IMPORTS (pihak ketiga)
# ============================================================================
import asyncio                                                    # noqa: E402
import ast                                                        # noqa: E402
import hashlib                                                    # noqa: E402
import functools                                                   # noqa: E402
import hmac                                                        # noqa: E402
import json                                                        # noqa: E402
import logging                                                     # noqa: E402
import platform                                                    # noqa: E402
import csv                                                         # noqa: E402
import io                                                          # noqa: E402
import urllib.request                                              # noqa: E402
import urllib.parse                                                # noqa: E402
import urllib.error                                                # noqa: E402
import tempfile                                                    # noqa: E402
import zipfile                                                     # noqa: E402
import re                                                          # noqa: E402
import secrets                                                     # noqa: E402
import shutil                                                      # noqa: E402
import sqlite3                                                     # noqa: E402
import threading                                                   # noqa: E402
import time                                                        # noqa: E402
import uuid                                                        # noqa: E402
from dataclasses import asdict, dataclass, field                   # noqa: E402
from logging.handlers import RotatingFileHandler                   # noqa: E402
from pathlib import Path                                           # noqa: E402
from typing import Dict, List, Optional, Set, Any, AsyncIterator     # noqa: E402

from dotenv import load_dotenv                                     # noqa: E402
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect  # noqa: E402
from fastapi.responses import (                                    # noqa: E402
    FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse, Response,
)
from pydantic import BaseModel, Field                               # noqa: E402
from starlette.background import BackgroundTask                    # noqa: E402
from starlette.middleware.sessions import SessionMiddleware         # noqa: E402


async def _to_thread(func, *args, **kwargs):
    """asyncio.to_thread kompatibel Python 3.8+ tanpa memblokir event loop."""
    native = getattr(asyncio, "to_thread", None)
    if native is not None:
        return await native(func, *args, **kwargs)
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, functools.partial(func, *args, **kwargs))


APP_VERSION = "4.0.3-advanced"
CONFIG_VERSION = 4
DB_SCHEMA_VERSION = 5

# ============================================================================
# CONFIG - baca .env di folder ini
# ============================================================================
BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

def _get(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def _get_int(name: str, default: int) -> int:
    try:
        return int(_get(name, str(default)))
    except ValueError:
        return default


def _get_bool(name: str, default: bool = False) -> bool:
    raw = _get(name, "1" if default else "0").lower()
    return raw in {"1", "true", "yes", "on", "y"}


def _resolve_data_dir(raw: str) -> Path:
    # Task Scheduler sering memulai proses dari C:\Windows\System32.
    # DATA_DIR relatif harus tetap menunjuk ke folder script, bukan CWD.
    candidate = Path(raw or "./data").expanduser()
    if not candidate.is_absolute():
        candidate = BASE_DIR / candidate
    return candidate.resolve()


def _get_list(name: str) -> List[str]:
    raw = _get(name, "")
    return [x.strip() for x in raw.split(",") if x.strip()]


def _parse_roots(raw: str) -> Dict[str, str]:
    """FILE_MANAGER_ROOTS format: 'Label:/abs/path;Label2:/abs/path2'"""
    roots: Dict[str, str] = {}
    for part in raw.split(";"):
        part = part.strip()
        if not part or ":" not in part:
            continue
        label, path = part.split(":", 1)
        label, path = label.strip(), path.strip()
        if label and path:
            roots[label] = str(Path(path).expanduser())
    return roots


@dataclass
class Settings:
    roots: Dict[str, str] = field(default_factory=lambda: _parse_roots(_get("FILE_MANAGER_ROOTS")))
    data_dir: Path = field(default_factory=lambda: _resolve_data_dir(_get("DATA_DIR", "./data")))
    max_upload_mb: int = field(default_factory=lambda: _get_int("MAX_UPLOAD_MB", 50))
    max_download_mb: int = field(default_factory=lambda: _get_int("MAX_DOWNLOAD_MB", 50))
    items_per_page: int = field(default_factory=lambda: _get_int("WEB_ITEMS_PER_PAGE", 20))
    search_max_results: int = field(default_factory=lambda: _get_int("SEARCH_MAX_RESULTS", 40))

    # Print lokal. Tidak ada dependency GAS/Bridge dalam alur print aktif.
    default_printer: str = field(default_factory=lambda: _get("DEFAULT_PRINTER"))
    print_job_timeout_sec: int = field(default_factory=lambda: _get_int("PRINT_JOB_TIMEOUT_SEC", 900))
    print_render_dpi: int = field(default_factory=lambda: _get_int("PRINT_RENDER_DPI", 144))
    print_max_retries: int = field(default_factory=lambda: _get_int("PRINT_MAX_RETRIES", 1))
    backup_retention_days: int = field(default_factory=lambda: _get_int("BACKUP_RETENTION_DAYS", 14))
    auto_start_task: bool = field(default_factory=lambda: _get_bool("AUTO_START_TASK", True))
    task_name: str = field(default_factory=lambda: _get("TASK_NAME", "PrintBot-WebApp"))

    # Security / session / quota / maintenance.
    session_max_age_sec: int = field(default_factory=lambda: _get_int("SESSION_MAX_AGE_SEC", 7 * 86400))
    session_idle_sec: int = field(default_factory=lambda: _get_int("SESSION_IDLE_SEC", 8 * 3600))
    login_max_attempts: int = field(default_factory=lambda: _get_int("LOGIN_MAX_ATTEMPTS", 5))
    login_lockout_sec: int = field(default_factory=lambda: _get_int("LOGIN_LOCKOUT_SEC", 300))
    secure_cookie: bool = field(default_factory=lambda: _get_bool("WEBAPP_SECURE_COOKIE", False))
    max_jobs_per_hour: int = field(default_factory=lambda: _get_int("MAX_JOBS_PER_HOUR", 30))
    max_copies_per_day: int = field(default_factory=lambda: _get_int("MAX_COPIES_PER_DAY", 200))
    max_copies_per_job: int = field(default_factory=lambda: _get_int("MAX_COPIES_PER_JOB", 50))
    index_refresh_sec: int = field(default_factory=lambda: _get_int("INDEX_REFRESH_SEC", 300))
    printer_monitor_sec: int = field(default_factory=lambda: _get_int("PRINTER_MONITOR_SEC", 15))
    db_maintenance_hours: int = field(default_factory=lambda: _get_int("DB_MAINTENANCE_HOURS", 24))
    update_url: str = field(default_factory=lambda: _get("UPDATE_URL", ""))
    update_token: str = field(default_factory=lambda: _get("UPDATE_TOKEN", ""))
    auto_update_apply: bool = field(default_factory=lambda: _get_bool("AUTO_UPDATE_APPLY", False))
    update_check_hours: int = field(default_factory=lambda: _get_int("UPDATE_CHECK_HOURS", 6))
    api_token: str = field(default_factory=lambda: _get("WEBAPP_API_TOKEN", ""))
    api_scopes: List[str] = field(default_factory=lambda: _get_list("WEBAPP_API_SCOPES") or ["read", "print"])
    telegram_bot_token: str = field(default_factory=lambda: _get("TELEGRAM_BOT_TOKEN", ""))
    telegram_admin_chat_id: str = field(default_factory=lambda: _get("TELEGRAM_ADMIN_CHAT_ID", ""))

    webapp_host: str = field(default_factory=lambda: _get("WEBAPP_HOST", "0.0.0.0"))
    webapp_port: int = field(default_factory=lambda: _get_int("WEBAPP_PORT", 8000))
    webapp_username: str = field(default_factory=lambda: _get("WEBAPP_USERNAME", "admin"))
    webapp_password: str = field(default_factory=lambda: _get("WEBAPP_PASSWORD"))
    webapp_secret_key: str = field(default_factory=lambda: _get("WEBAPP_SECRET_KEY"))
    log_level: str = field(default_factory=lambda: _get("LOG_LEVEL", "INFO"))

    def __post_init__(self) -> None:
        self.data_dir.mkdir(parents=True, exist_ok=True)
        for name in ("uploads", "previews", "logs", "roots", "print_archive", "backup", "updates", "exports"):
            (self.data_dir / name).mkdir(exist_ok=True)
        if not self.webapp_secret_key:
            # Persist agar sesi tidak invalid setiap restart.
            secret_file = self.data_dir / "webapp_secret.key"
            try:
                if secret_file.is_file():
                    self.webapp_secret_key = secret_file.read_text("utf-8").strip()
                if not self.webapp_secret_key:
                    self.webapp_secret_key = secrets.token_urlsafe(48)
                    secret_file.write_text(self.webapp_secret_key, encoding="utf-8")
                    try:
                        os.chmod(str(secret_file), 0o600)
                    except OSError:
                        pass
            except Exception:
                self.webapp_secret_key = secrets.token_urlsafe(48)
        if not self.api_token:
            token_file=self.data_dir/"api_token.key"
            try:
                if token_file.is_file(): self.api_token=token_file.read_text("utf-8").strip()
                if not self.api_token:
                    self.api_token=secrets.token_urlsafe(36); token_file.write_text(self.api_token,encoding="utf-8")
                    try: os.chmod(str(token_file),0o600)
                    except OSError: pass
            except Exception:
                self.api_token=secrets.token_urlsafe(36)

    @property
    def db_path(self) -> Path:
        return self.data_dir / "printbot.db"

    @property
    def log_path(self) -> Path:
        return self.data_dir / "logs" / "printbot.log"

    @property
    def roots_base_dir(self) -> Path:
        raw = _get("ROOTS_BASE_DIR", "")
        if raw:
            p = Path(raw).expanduser()
            return p if p.is_absolute() else (BASE_DIR / p).resolve()
        return self.data_dir / "roots"

    @property
    def print_archive_dir(self) -> Path:
        return self.data_dir / "print_archive"

    @property
    def backup_dir(self) -> Path:
        return self.data_dir / "backup"

    def validate(self) -> List[str]:
        errs: List[str] = []
        if not self.roots and not (self.data_dir / "roots.json").exists():
            print("[INFO] Belum ada folder root - tambahkan lewat tombol '+ Folder' di sidebar WebApp.")
        if not self.webapp_password:
            errs.append("WEBAPP_PASSWORD belum diisi di .env")
        elif len(self.webapp_password) < 8:
            print("[SECURITY WARNING] WEBAPP_PASSWORD sebaiknya minimal 8 karakter.")
        if self.webapp_port < 1 or self.webapp_port > 65535:
            errs.append("WEBAPP_PORT tidak valid")
        if self.print_render_dpi < 72 or self.print_render_dpi > 600:
            errs.append("PRINT_RENDER_DPI harus 72-600")
        return errs


settings = Settings()


# ============================================================================
# LOGGING
# ============================================================================
def setup_logging(name: str) -> logging.Logger:
    logger = logging.getLogger(name)
    if logger.handlers:
        return logger
    settings.log_path.parent.mkdir(parents=True, exist_ok=True)
    logger.setLevel(settings.log_level.upper())
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s", "%Y-%m-%d %H:%M:%S")
    fh = RotatingFileHandler(settings.log_path, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    logger.propagate = False
    return logger


# ============================================================================
# SECURITY - anti path-traversal, blocklist ekstensi, perbandingan token aman
# ============================================================================
BLOCKED_UPLOAD_EXT = {
    ".exe", ".bat", ".cmd", ".vbs", ".js", ".ps1", ".msi", ".com", ".scr",
    ".jar", ".sh", ".dll", ".vbe", ".wsf", ".jse", ".ps1xml", ".psc1",
}


class PathSecurityError(Exception):
    pass


def safe_join(root: str, *parts: str) -> Path:
    """Resolve `parts` di bawah `root`, raise PathSecurityError kalau ada
    upaya traversal / symlink-escape. Selalu mengembalikan Path yang
    dijamin berada di dalam `root`."""
    root_resolved = Path(root).resolve()
    candidate = root_resolved
    for part in parts:
        if not part:
            continue
        cleaned = str(part).replace("\\", "/")
        for piece in cleaned.split("/"):
            if piece in ("", "."):
                continue
            if piece == "..":
                raise PathSecurityError("Path traversal terdeteksi.")
            if "\x00" in piece or (len(piece) == 2 and piece[1] == ":"):
                raise PathSecurityError("Path tidak valid.")
            candidate = candidate / piece
    resolved = candidate.resolve()
    try:
        resolved.relative_to(root_resolved)
    except ValueError:
        raise PathSecurityError("Akses di luar root direktori ditolak.")
    return resolved


def is_upload_blocked(filename: str) -> bool:
    return Path(filename).suffix.lower() in BLOCKED_UPLOAD_EXT


def constant_time_eq(a: str, b: str) -> bool:
    return hmac.compare_digest(a or "", b or "")


WINDOWS_RESERVED_NAMES = {"CON", "PRN", "AUX", "NUL", *("COM%d" % i for i in range(1, 10)), *("LPT%d" % i for i in range(1, 10))}
WINDOWS_INVALID_CHARS = set('<>:"/\\|?*')


def validate_windows_filename(name: str, allow_extension: bool = True) -> str:
    """Validasi nama yang aman di Windows 7-11, juga aman untuk host lain."""
    value = (name or "").strip()
    if not value or value in {".", ".."}:
        raise FileManagerError("Nama tidak boleh kosong.")
    if len(value) > 240:
        raise FileManagerError("Nama terlalu panjang (maksimal 240 karakter).")
    if any(ord(ch) < 32 or ch in WINDOWS_INVALID_CHARS for ch in value):
        raise FileManagerError('Nama mengandung karakter yang tidak valid di Windows: < > : " / \\ | ? *')
    if value.endswith((" ", ".")):
        raise FileManagerError("Nama Windows tidak boleh diakhiri spasi atau titik.")
    stem = value.split(".", 1)[0].upper()
    if stem in WINDOWS_RESERVED_NAMES:
        raise FileManagerError("Nama '%s' merupakan nama perangkat khusus Windows." % stem)
    return value


# ============================================================================
# FILE MANAGER - browsing, search, info (CORE - identik dengan webapp.py)
# ============================================================================
DOC_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".xlsm", ".ppt", ".pptx", ".txt", ".rtf"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff"}
PRINTABLE_EXTS = DOC_EXTS | IMAGE_EXTS


class FileManagerError(Exception):
    pass


@dataclass
class Entry:
    name: str
    rel_path: str
    is_dir: bool
    size: int = 0
    modified: float = 0.0
    ext: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


def fm_root_path(root_key: str) -> str:
    root = get_roots().get(root_key)
    if not root:
        raise FileManagerError(f"Root '{root_key}' tidak dikenal.")
    if not Path(root).is_dir():
        raise FileManagerError(f"Folder root '{root_key}' tidak ditemukan di server: {root}")
    return root


def fm_resolve(root_key: str, rel_path: str = "") -> Path:
    root = fm_root_path(root_key)
    try:
        return safe_join(root, rel_path or "")
    except PathSecurityError as e:
        raise FileManagerError(str(e))


def fm_to_rel(root_key: str, abs_path: Path) -> str:
    root = Path(fm_root_path(root_key))
    return str(abs_path.relative_to(root)).replace("\\", "/")


def fm_breadcrumb(root_key: str, rel_path: str) -> List[dict]:
    crumbs = [{"label": root_key, "rel_path": ""}]
    if not rel_path:
        return crumbs
    parts = [p for p in rel_path.replace("\\", "/").split("/") if p]
    acc: List[str] = []
    for p in parts:
        acc.append(p)
        crumbs.append({"label": p, "rel_path": "/".join(acc)})
    return crumbs


def fm_list_dir(root_key: str, rel_path: str = "", sort_by: str = "name", page: int = 0,
                 page_size: Optional[int] = None) -> dict:
    target = fm_resolve(root_key, rel_path)
    if not target.is_dir():
        raise FileManagerError("Folder tidak ditemukan.")
    page_size = page_size or settings.items_per_page

    dirs: List[Entry] = []
    files: List[Entry] = []
    try:
        for child in target.iterdir():
            try:
                st = child.stat()
            except OSError:
                continue
            rel = fm_to_rel(root_key, child)
            if child.is_dir():
                dirs.append(Entry(child.name, rel, True, 0, st.st_mtime))
            else:
                files.append(Entry(child.name, rel, False, st.st_size, st.st_mtime, child.suffix.lower()))
    except PermissionError:
        raise FileManagerError("Akses ditolak oleh sistem operasi.")

    keyfn = {
        "name": lambda e: e.name.lower(),
        "size": lambda e: e.size,
        "type": lambda e: (e.ext, e.name.lower()),
        "modified": lambda e: e.modified,
    }.get(sort_by, lambda e: e.name.lower())
    reverse = sort_by in ("size", "modified")
    # Folder tidak punya "jenis" - saat urut per jenis, folder tetap diurut per nama.
    dirs.sort(key=(lambda e: e.name.lower()) if sort_by == "type" else keyfn,
              reverse=False if sort_by == "type" else reverse)
    files.sort(key=keyfn, reverse=reverse)

    items = dirs + files
    total = len(items)
    start = page * page_size
    end = min(start + page_size, total)
    page_items = items[start:end]

    return {
        "root": root_key,
        "rel_path": rel_path.strip("/"),
        "breadcrumb": fm_breadcrumb(root_key, rel_path),
        "items": [e.to_dict() for e in page_items],
        "total": total,
        "page": page,
        "page_size": page_size,
        "has_next": end < total,
        "has_prev": page > 0,
        "dir_count": len(dirs),
        "file_count": len(files),
    }


def fm_build_search_pattern(keyword: str) -> re.Pattern:
    """Cari fleksibel per-kata: 'Anggota Keluar' cocok ke
    Form_Anggota_Keluar.pdf, Anggota-Keluar.docx, ANGGOTA KELUAR (1).xlsx, dst."""
    words = [re.escape(w) for w in keyword.strip().split() if w]
    if not words:
        words = [re.escape(keyword.strip())]
    pattern = r"[\s_\-.()]+".join(words)
    return re.compile(pattern, re.IGNORECASE)


def fm_search(root_key: Optional[str], keyword: str, max_results: Optional[int] = None) -> List[dict]:
    return indexed_search(root_key, keyword, max_results)


def fm_get_excel_sheet_names(path: Path) -> List[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def fm_human_size(num: float) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if num < 1024:
            return f"{int(num)} {unit}" if unit == "B" else f"{num:.1f} {unit}"
        num /= 1024
    return f"{num:.1f} TB"


def fm_rename(root_key: str, rel_path: str, new_name: str) -> Path:
    """Ganti nama file/folder (tetap di folder induk yang sama)."""
    new_name = validate_windows_filename(new_name)
    old_path = fm_resolve(root_key, rel_path)
    if not old_path.exists():
        raise FileManagerError("Item tidak ditemukan.")
    new_path = old_path.parent / new_name
    try:
        root_base = Path(fm_root_path(root_key))
        new_path.resolve().relative_to(root_base)
    except Exception:
        raise FileManagerError("Nama baru tidak valid.")
    if new_path.exists():
        raise FileManagerError("Sudah ada file/folder dengan nama itu di sini.")
    try:
        old_path.rename(new_path)
    except OSError as e:
        raise FileManagerError(f"Gagal mengganti nama: {e}")
    return new_path


def fm_delete(root_key: str, rel_path: str) -> None:
    """Hapus file, atau folder beserta isinya."""
    p = fm_resolve(root_key, rel_path)
    if not p.exists():
        raise FileManagerError("Item tidak ditemukan.")
    try:
        if p.is_dir():
            shutil.rmtree(p)
        else:
            p.unlink()
    except OSError as e:
        raise FileManagerError(f"Gagal menghapus: {e}")


def fm_mkdir(root_key: str, rel_path: str, name: str) -> Path:
    """Buat subfolder baru di dalam rel_path (folder yang sedang dibuka)."""
    name = validate_windows_filename(name)
    parent = fm_resolve(root_key, rel_path)
    if not parent.is_dir():
        raise FileManagerError("Folder induk tidak ditemukan.")
    new_dir = parent / name
    try:
        root_base = Path(fm_root_path(root_key))
        new_dir.resolve().relative_to(root_base)
    except Exception:
        raise FileManagerError("Path tidak valid.")
    if new_dir.exists():
        raise FileManagerError("Folder dengan nama itu sudah ada di sini.")
    try:
        new_dir.mkdir(parents=False)
    except OSError as e:
        raise FileManagerError(f"Gagal membuat folder: {e}")
    return new_dir


def fm_save_upload(root_key: str, rel_path: str, filename: str, data: bytes) -> Path:
    """Simpan bytes upload sebagai file baru di rel_path (folder tujuan).
    Kalau nama sudah ada, otomatis diberi suffix (1), (2), dst agar tidak
    menimpa file lain."""
    filename = validate_windows_filename(Path(filename).name)  # buang komponen path apa pun dari nama asli
    if is_upload_blocked(filename):
        raise FileManagerError("Ekstensi file ditolak demi keamanan.")
    target_dir = fm_resolve(root_key, rel_path)
    if not target_dir.is_dir():
        raise FileManagerError("Folder tujuan tidak ditemukan.")
    dest = target_dir / filename
    if dest.exists():
        stem, suf = Path(filename).stem, Path(filename).suffix
        i = 1
        while dest.exists():
            dest = target_dir / f"{stem} ({i}){suf}"
            i += 1
    try:
        dest.write_bytes(data)
    except OSError as e:
        raise FileManagerError(f"Gagal menyimpan file: {e}")
    return dest



# ============================================================================
# PREVIEW - render halaman 1 dokumen jadi JPEG. Beberapa metode dicoba
# berurutan, dari yang paling tidak butuh dependency sistem (PyMuPDF/
# openpyxl+Pillow, murni pip - selalu tersedia lewat auto-install) sampai
# yang butuh system tool (LibreOffice/poppler/ghostscript, best-effort).
# Ini memastikan preview PDF & Excel tetap jalan walau LibreOffice/
# poppler-utils gagal/tidak terinstall di server.
# ============================================================================
def _tool_find(*names: str) -> Optional[str]:
    for n in names:
        p = shutil.which(n)
        if p:
            return p
    return None


def preview_available_tools() -> dict:
    return {
        "PyMuPDF (PDF)": importlib.util.find_spec("fitz") is not None,
        "openpyxl+Pillow (Excel)": importlib.util.find_spec("openpyxl") is not None
        and importlib.util.find_spec("PIL") is not None,
        "python-docx (Word)": importlib.util.find_spec("docx") is not None,
        "libreoffice": _tool_find("libreoffice", "soffice") is not None,
        "pdftoppm": _tool_find("pdftoppm") is not None,
        "ghostscript": _tool_find("gs") is not None,
    }


def _run_tool(cmd: list, timeout: int = 90) -> bool:
    try:
        r = subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=timeout)
        return r.returncode == 0
    except Exception:
        return False


def preview_pdf_to_jpg(pdf_path: Path, out_jpg: Path, dpi: int = 130) -> bool:
    """Coba beberapa metode berurutan sampai salah satu berhasil:
    1) PyMuPDF (fitz) - murni pip, tanpa dependency sistem, PALING andal.
    2) pdftoppm (poppler-utils) - kalau terinstall di sistem.
    3) ghostscript - fallback terakhir."""
    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        try:
            if doc.page_count > 0:
                page = doc.load_page(0)
                zoom = dpi / 72.0
                pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
                pix.save(str(out_jpg))
                if out_jpg.exists() and out_jpg.stat().st_size > 100:
                    return True
        finally:
            doc.close()
    except Exception as e:
        log.debug("preview_pdf_to_jpg: PyMuPDF gagal (%s), coba metode lain.", e)

    pdftoppm = _tool_find("pdftoppm")
    if pdftoppm:
        out_prefix = str(out_jpg.with_suffix(""))
        _run_tool([pdftoppm, "-jpeg", "-r", str(dpi), "-f", "1", "-l", "1", str(pdf_path), out_prefix])
        produced = out_jpg.with_name(out_jpg.stem + "-1.jpg")
        if produced.exists():
            produced.rename(out_jpg)
        if out_jpg.exists() and out_jpg.stat().st_size > 100:
            return True

    gs = _tool_find("gs")
    if gs:
        ok = _run_tool([gs, "-dNOPAUSE", "-dBATCH", "-dSAFER", "-sDEVICE=jpeg", f"-r{dpi}",
                         "-dJPEGQ=85", "-dFirstPage=1", "-dLastPage=1",
                         f"-sOutputFile={out_jpg}", str(pdf_path)])
        if ok and out_jpg.exists() and out_jpg.stat().st_size > 100:
            return True

    return False


def preview_office_to_pdf(src: Path, out_dir: Path) -> Optional[Path]:
    """Preview Office: prioritaskan Microsoft Office COM di Windows agar hasil
    sama dengan layout print asli; LibreOffice hanya fallback lintas platform."""
    out_dir.mkdir(parents=True, exist_ok=True)
    out_pdf = out_dir / (src.stem + "_preview_%s.pdf" % uuid.uuid4().hex[:8])
    ext = src.suffix.lower()
    if sys.platform.startswith("win") and ext in {".xls", ".xlsx", ".xlsm", ".doc", ".docx", ".rtf", ".ppt", ".pptx"}:
        try:
            import pythoncom
            import win32com.client
            pythoncom.CoInitialize()
            try:
                if ext in {".xls", ".xlsx", ".xlsm"}:
                    app = wb = None
                    try:
                        app = win32com.client.DispatchEx("Excel.Application")
                        app.Visible = False; app.DisplayAlerts = False
                        wb = app.Workbooks.Open(str(src), ReadOnly=True, UpdateLinks=0)
                        wb.ExportAsFixedFormat(0, str(out_pdf))
                    finally:
                        try:
                            if wb is not None: wb.Close(False)
                        except Exception: pass
                        try:
                            if app is not None: app.Quit()
                        except Exception: pass
                elif ext in {".doc", ".docx", ".rtf"}:
                    app = doc = None
                    try:
                        app = win32com.client.DispatchEx("Word.Application")
                        app.Visible = False; app.DisplayAlerts = 0
                        doc = app.Documents.Open(str(src), ReadOnly=True, AddToRecentFiles=False)
                        doc.ExportAsFixedFormat(str(out_pdf), 17)
                    finally:
                        try:
                            if doc is not None: doc.Close(False)
                        except Exception: pass
                        try:
                            if app is not None: app.Quit()
                        except Exception: pass
                else:
                    app = pres = None
                    try:
                        app = win32com.client.DispatchEx("PowerPoint.Application")
                        pres = app.Presentations.Open(str(src), WithWindow=False, ReadOnly=True)
                        pres.ExportAsFixedFormat(str(out_pdf), 2)
                    finally:
                        try:
                            if pres is not None: pres.Close()
                        except Exception: pass
                        try:
                            if app is not None: app.Quit()
                        except Exception: pass
            finally:
                pythoncom.CoUninitialize()
            if out_pdf.is_file() and out_pdf.stat().st_size > 100:
                return out_pdf
        except Exception as exc:
            log.debug("Preview Office COM gagal (%s), mencoba LibreOffice.", exc)
            try: out_pdf.unlink(missing_ok=True)
            except Exception: pass

    soffice = _tool_find("libreoffice", "soffice")
    if not soffice:
        return None
    # LibreOffice menentukan nama sendiri, jadi pakai direktori unik agar tidak
    # bentrok dengan preview paralel.
    temp_out = out_dir / ("office_%s" % uuid.uuid4().hex[:8])
    temp_out.mkdir(exist_ok=True)
    try:
        _run_tool([soffice, "--headless", "--norestart", "--convert-to", "pdf",
                   "--outdir", str(temp_out), str(src)], timeout=90)
        candidate = temp_out / (src.stem + ".pdf")
        if candidate.exists():
            shutil.move(str(candidate), str(out_pdf))
            return out_pdf
    finally:
        shutil.rmtree(temp_out, ignore_errors=True)
    return None


def _draw_text_image(out_jpg: Path, title: str, lines: list, mono: bool = False) -> bool:
    """Helper umum: render daftar baris teks jadi satu gambar JPEG rapi -
    dipakai fallback Excel (tabel) & Word (paragraf) tanpa LibreOffice."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return False
    try:
        font_name = "DejaVuSansMono.ttf" if mono else "DejaVuSans.ttf"
        font = ImageFont.truetype(font_name, 14)
        title_font = ImageFont.truetype("DejaVuSans-Bold.ttf", 16)
    except Exception:
        font = ImageFont.load_default()
        title_font = font

    line_h = 20
    img_w = 900
    img_h = line_h * (len(lines) + 2) + 24
    img = Image.new("RGB", (img_w, max(img_h, 120)), "white")
    draw = ImageDraw.Draw(img)
    y = 12
    draw.text((14, y), title[:110], fill="#111111", font=title_font)
    y += line_h + 6
    for ln in lines:
        draw.text((14, y), ln[:130], fill="#222222", font=font)
        y += line_h
    img.save(out_jpg, "JPEG", quality=85)
    return out_jpg.exists() and out_jpg.stat().st_size > 100


def preview_xlsx_to_jpg(path: Path, out_jpg: Path, max_rows: int = 35, max_cols: int = 10) -> bool:
    """Fallback preview Excel TANPA LibreOffice: baca isi sheet aktif lewat
    openpyxl, render sebagai tabel teks sederhana lewat Pillow."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
        wb.close()
    except Exception as e:
        log.debug("preview_xlsx_to_jpg gagal baca workbook: %s", e)
        return False
    if not rows:
        return False

    col_count = max(len(r) for r in rows)
    rows = [r + [""] * (col_count - len(r)) for r in rows]
    col_w = [min(max(max(len(r[c]) for r in rows), 3), 18) for c in range(col_count)]

    lines = []
    header = " | ".join(v[:w].ljust(w) for v, w in zip(rows[0], col_w))
    lines.append(header)
    lines.append("-" * len(header))
    for r in rows[1:]:
        lines.append(" | ".join(v[:w].ljust(w) for v, w in zip(r, col_w)))

    sheet_name = getattr(ws, "title", "Sheet1")
    return _draw_text_image(out_jpg, f"{path.name}  \u2022  Sheet: {sheet_name}", lines, mono=True)


def preview_docx_to_jpg(path: Path, out_jpg: Path, max_paragraphs: int = 45) -> bool:
    """Fallback preview Word TANPA LibreOffice: ekstrak teks lewat
    python-docx, render sebagai gambar teks lewat Pillow (hanya .docx -
    format .doc lama tidak didukung tanpa LibreOffice)."""
    try:
        import docx
        import textwrap
        d = docx.Document(str(path))
        paras = [p.text for p in d.paragraphs if p.text.strip()][:max_paragraphs]
    except Exception as e:
        log.debug("preview_docx_to_jpg gagal baca dokumen: %s", e)
        return False
    if not paras:
        return False
    lines = []
    for p in paras:
        wrapped = textwrap.wrap(p, width=95) or [""]
        lines.extend(wrapped)
    return _draw_text_image(out_jpg, path.name, lines[:200])


def generate_preview(src_path: Path) -> Optional[Path]:
    ext = src_path.suffix.lower()
    work_dir = settings.data_dir / "previews"
    work_dir.mkdir(exist_ok=True)
    out_jpg = work_dir / f"pvw_{uuid.uuid4().hex}.jpg"

    if ext in (".jpg", ".jpeg"):
        shutil.copy(src_path, out_jpg)
        return out_jpg

    if ext in IMAGE_EXTS:
        try:
            from PIL import Image
            with Image.open(src_path) as im:
                im.convert("RGB").save(out_jpg, "JPEG", quality=85)
            return out_jpg
        except Exception:
            return None

    if ext == ".pdf":
        return out_jpg if preview_pdf_to_jpg(src_path, out_jpg) else None

    if ext in (".xls", ".xlsx", ".xlsm"):
        # Coba LibreOffice dulu (hasil paling presisi, sesuai format asli).
        pdf = preview_office_to_pdf(src_path, work_dir)
        if pdf:
            ok = preview_pdf_to_jpg(pdf, out_jpg)
            pdf.unlink(missing_ok=True)
            if ok:
                return out_jpg
        # Fallback: render tabel langsung dari openpyxl - tanpa LibreOffice.
        if preview_xlsx_to_jpg(src_path, out_jpg):
            return out_jpg
        return None

    if ext in (".doc", ".docx", ".ppt", ".pptx", ".rtf", ".odt"):
        pdf = preview_office_to_pdf(src_path, work_dir)
        if pdf:
            ok = preview_pdf_to_jpg(pdf, out_jpg)
            pdf.unlink(missing_ok=True)
            if ok:
                return out_jpg
        # Fallback tanpa LibreOffice (hanya format .docx yang didukung).
        if ext == ".docx" and preview_docx_to_jpg(src_path, out_jpg):
            return out_jpg
        return None

    return None


def preview_document_meta(src_path: Path) -> dict:
    ext=src_path.suffix.lower(); count=1
    if ext==".pdf":
        try:
            import fitz
            d=fitz.open(str(src_path)); count=d.page_count; d.close()
        except Exception: count=1
    elif ext in {".xls",".xlsx",".xlsm",".doc",".docx",".rtf",".ppt",".pptx",".odt"}:
        pdf=preview_office_to_pdf(src_path,settings.data_dir/"previews")
        if pdf:
            try:
                import fitz
                d=fitz.open(str(pdf)); count=d.page_count; d.close()
            except Exception: count=1
            try: pdf.unlink()
            except OSError: pass
    return {"pages":max(1,int(count)),"ext":ext,"name":src_path.name}


def generate_preview_page(src_path: Path, page_num: int = 1, dpi: int = 130) -> Optional[Path]:
    ext=src_path.suffix.lower(); out=settings.data_dir/"previews"/("pvwp_%s.jpg"%uuid.uuid4().hex); source_pdf=None; cleanup_pdf=False
    if ext in IMAGE_EXTS:
        if page_num!=1: return None
        return generate_preview(src_path)
    if ext==".pdf": source_pdf=src_path
    elif ext in {".xls",".xlsx",".xlsm",".doc",".docx",".rtf",".ppt",".pptx",".odt"}:
        source_pdf=preview_office_to_pdf(src_path,settings.data_dir/"previews"); cleanup_pdf=bool(source_pdf)
    if source_pdf:
        try:
            import fitz
            doc=fitz.open(str(source_pdf))
            try:
                idx=max(0,min(int(page_num)-1,doc.page_count-1)); page=doc.load_page(idx); pix=page.get_pixmap(matrix=fitz.Matrix(dpi/72.0,dpi/72.0)); pix.save(str(out))
            finally: doc.close()
            return out if out.is_file() else None
        except Exception as exc:
            log.debug("generate_preview_page gagal: %s",exc); return None
        finally:
            if cleanup_pdf and source_pdf:
                try: source_pdf.unlink()
                except OSError: pass
    if page_num==1: return generate_preview(src_path)
    return None


def ensure_system_tools_async() -> None:
    """Best-effort, TIDAK memblokir startup: coba install otomatis
    LibreOffice/poppler-utils/ghostscript (untuk hasil Preview Office/PDF
    paling presisi) lewat apt-get kalau belum ada dan proses berjalan
    sebagai root atau lewat sudo non-interaktif. Preview PDF & Excel SUDAH
    tetap berfungsi tanpa ini (fallback PyMuPDF/openpyxl+Pillow murni pip),
    jadi ini hanya peningkatan kualitas, bukan syarat - kalau gagal/tidak
    memenuhi syarat cukup dicatat di log, fitur lain tetap berjalan normal."""
    tools = preview_available_tools()
    system_tools = {k: v for k, v in tools.items() if k in ("libreoffice", "pdftoppm", "ghostscript")}
    if all(system_tools.values()):
        return
    missing_names = ", ".join(k for k, v in system_tools.items() if not v)
    if not sys.platform.startswith("linux"):
        # Di Windows tool ini opsional. Preview utama menggunakan Microsoft Office COM
        # (jika Office tersedia), PyMuPDF, openpyxl/Pillow, dan python-docx.
        log.debug("Tool preview opsional tidak terpasang: %s. Engine native/fallback tetap aktif.", missing_names)
        return
    apt = _tool_find("apt-get")
    if not apt:
        log.info("apt-get tidak ada - install manual jika perlu: libreoffice poppler-utils ghostscript "
                  "(preview PDF/Excel/Word tetap jalan lewat fallback pip).")
        return
    is_root = hasattr(os, "geteuid") and os.geteuid() == 0
    sudo = _tool_find("sudo")
    if not is_root and not sudo:
        log.info("Bukan root & sudo tidak tersedia - install manual jika perlu: sudo apt-get install -y "
                  "libreoffice poppler-utils ghostscript (preview tetap jalan lewat fallback pip).")
        return

    def _worker():
        try:
            base = [] if is_root else ["sudo", "-n"]
            log.info("Mencoba install otomatis tool Preview yang belum ada: %s ...", missing_names)
            subprocess.run(base + ["apt-get", "update", "-qq"], timeout=300,
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            r = subprocess.run(base + ["apt-get", "install", "-y", "-qq",
                                        "libreoffice", "poppler-utils", "ghostscript"],
                                timeout=1800, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            if r.returncode == 0:
                log.info("Tool Preview berhasil diinstall otomatis.")
            else:
                log.warning("Install otomatis tool Preview gagal (kode %s) - install manual jika perlu:\n"
                             "  sudo apt-get install -y libreoffice poppler-utils ghostscript", r.returncode)
        except Exception as e:
            log.warning("Install otomatis tool Preview gagal: %s", e)

    threading.Thread(target=_worker, daemon=True, name="printbot-tools-install").start()


# ============================================================================
# PRINT QUEUE - SQLite (WAL), dapat dibagi dengan bot.py (data/printbot.db).
# Semua job baru WebApp memakai method='direct' dan diproses worker printer lokal.
# Kolom legacy GAS/Bridge hanya dipertahankan agar database lama tetap terbaca.
# ============================================================================
_db_lock = threading.Lock()


class JobStatus:
    QUEUED = "QUEUED"          # menunggu worker printer lokal
    PROCESSING = "PROCESSING"  # sudah diklaim worker
    PRINTING = "PRINTING"      # sedang render/spooling/printing
    SUCCESS = "SUCCESS"
    FAILED = "FAILED"
    TERMINAL = {SUCCESS, FAILED}
    LABELS = {
        QUEUED: "Queued", PROCESSING: "Processing", PRINTING: "Printing",
        SUCCESS: "Success", FAILED: "Failed",
    }


class PrintMethod:
    # BRIDGE/GAS dipertahankan hanya agar riwayat DB lama tetap dapat dibaca.
    # Semua job BARU dan reprint selalu DIRECT.
    BRIDGE = "bridge"
    GAS = "gas"
    DIRECT = "direct"
    LABELS = {BRIDGE: "Legacy Bridge", GAS: "Legacy GAS", DIRECT: "Printer Lokal"}


def _db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(str(settings.db_path), check_same_thread=False, timeout=15)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA busy_timeout=8000;")
    conn.row_factory = sqlite3.Row
    return conn


def db_init() -> None:
    with _db_lock, _db_connect() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS jobs (
                id TEXT PRIMARY KEY,
                source TEXT NOT NULL,
                sender TEXT NOT NULL,
                sender_ref TEXT,
                file_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                label TEXT,
                copies INTEGER NOT NULL DEFAULT 1,
                pages TEXT,
                sheet TEXT,
                method TEXT NOT NULL DEFAULT 'direct',
                printer TEXT,
                gas_job_id TEXT,
                reprint_of TEXT,
                archive_path TEXT,
                status TEXT NOT NULL DEFAULT 'QUEUED',
                error TEXT,
                retries INTEGER NOT NULL DEFAULT 0,
                created_at REAL NOT NULL,
                dispatched_at REAL,
                finished_at REAL
            )
        """)
        # Migrasi lembut untuk DB lama (sebelum kolom method/printer ada) - aman
        # dijalankan berulang kali, kolom yang sudah ada akan diabaikan (error ditelan).
        existing_cols = {row["name"] for row in conn.execute("PRAGMA table_info(jobs)").fetchall()}
        required_v4={"method","printer","archive_path","paper","orientation","scale_mode","duplex","color_mode","spool_job_id","priority","scheduled_at"}
        if settings.db_path.is_file() and not required_v4.issubset(existing_cols):
            try:
                pre=settings.backup_dir/("PreMigration_v4_%s.db"%time.strftime("%Y%m%d_%H%M%S")); shutil.copy2(str(settings.db_path),str(pre))
            except Exception: pass
        if "method" not in existing_cols:
            conn.execute("ALTER TABLE jobs ADD COLUMN method TEXT NOT NULL DEFAULT 'direct'")
        if "printer" not in existing_cols:
            conn.execute("ALTER TABLE jobs ADD COLUMN printer TEXT")
        if "gas_job_id" not in existing_cols:
            conn.execute("ALTER TABLE jobs ADD COLUMN gas_job_id TEXT")
        if "reprint_of" not in existing_cols:
            conn.execute("ALTER TABLE jobs ADD COLUMN reprint_of TEXT")
        if "archive_path" not in existing_cols:
            conn.execute("ALTER TABLE jobs ADD COLUMN archive_path TEXT")
        # Kolom print lokal/spooler. Migrasi aman untuk DB dari versi Bridge lama.
        migrations = [
            ("paper", "TEXT"), ("orientation", "TEXT"), ("scale_mode", "TEXT"),
            ("duplex", "TEXT"), ("color_mode", "TEXT"), ("spool_job_id", "INTEGER"),
            ("spool_status", "TEXT"), ("started_at", "REAL"), ("cancelled_at", "REAL"),
            ("spool_cancelled", "INTEGER NOT NULL DEFAULT 0"), ("priority", "INTEGER NOT NULL DEFAULT 0"),
            ("scheduled_at", "REAL"),
        ]
        for col, ddl in migrations:
            if col not in existing_cols:
                conn.execute("ALTER TABLE jobs ADD COLUMN %s %s" % (col, ddl))
        conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_status ON jobs(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_direct_queue ON jobs(method,status,priority,created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_gas_job_id ON jobs(gas_job_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_file_path ON jobs(file_path)")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS bridge_state (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                last_poll_at REAL
            )
        """)
        conn.execute("INSERT OR IGNORE INTO bridge_state (id, last_poll_at) VALUES (1, NULL)")
        conn.commit()


def _row_to_dict(r: sqlite3.Row) -> dict:
    d = dict(r)
    if d.get("status") == JobStatus.FAILED and str(d.get("error") or "").lower().startswith("dibatalkan"):
        d["status_label"] = "Dibatalkan"
    else:
        d["status_label"] = JobStatus.LABELS.get(d["status"], d["status"])
    return d


def job_create(*, source: str, sender: str, sender_ref: str, file_path: str, file_name: str,
               label: str = "", copies: int = 1, pages: str = "", sheet: str = "",
               method: str = PrintMethod.DIRECT, printer: str = "", reprint_of: str = "",
               archive_path: str = "", paper: str = "AUTO", orientation: str = "AUTO",
               scale_mode: str = "FIT", duplex: str = "DEFAULT", color_mode: str = "DEFAULT",
               priority: int = 0, scheduled_at: Optional[float] = None) -> dict:
    job_id = uuid.uuid4().hex[:16]
    now = time.time()
    with _db_lock, _db_connect() as conn:
        conn.execute(
            """INSERT INTO jobs (id, source, sender, sender_ref, file_path, file_name, label,
                                  copies, pages, sheet, method, printer, reprint_of, archive_path,
                                  paper, orientation, scale_mode, duplex, color_mode, priority, scheduled_at,
                                  status, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (job_id, source, sender, sender_ref, str(file_path), file_name, label or file_name,
             max(1, min(int(copies), 99)), pages or None, sheet or None,
             PrintMethod.DIRECT, printer or None, reprint_of or None, archive_path or None,
             (paper or "AUTO").upper(), (orientation or "AUTO").upper(),
             (scale_mode or "FIT").upper(), (duplex or "DEFAULT").upper(),
             (color_mode or "DEFAULT").upper(), max(0, min(int(priority), 100)),
             float(scheduled_at) if scheduled_at else None, JobStatus.QUEUED, now),
        )
        conn.commit()
    return job_get(job_id)


def job_get(job_id: str) -> Optional[dict]:
    with _db_lock, _db_connect() as conn:
        row = conn.execute(
            """SELECT j.*,
                      (SELECT COUNT(*) FROM jobs x
                       WHERE x.file_path=j.file_path AND x.status=?) AS print_count
               FROM jobs j WHERE j.id=?""",
            (JobStatus.SUCCESS, job_id),
        ).fetchone()
        return _row_to_dict(row) if row else None


def job_list(status: Optional[str] = None, limit: int = 50, offset: int = 0) -> List[dict]:
    q = """SELECT j.*,
                  (SELECT COUNT(*) FROM jobs x
                   WHERE x.file_path=j.file_path AND x.status=?) AS print_count
           FROM jobs j WHERE 1=1"""
    args: list = [JobStatus.SUCCESS]
    if status:
        q += " AND j.status=?"
        args.append(status)
    q += " ORDER BY j.created_at DESC LIMIT ? OFFSET ?"
    args += [limit, offset]
    with _db_lock, _db_connect() as conn:
        rows = conn.execute(q, args).fetchall()
        return [_row_to_dict(r) for r in rows]


def job_list_for_sender(sender: str, status: Optional[str] = None, limit: int = 50, offset: int = 0) -> List[dict]:
    q="""SELECT j.*, (SELECT COUNT(*) FROM jobs x WHERE x.file_path=j.file_path AND x.status=?) AS print_count
           FROM jobs j WHERE j.sender=?"""; args=[JobStatus.SUCCESS,sender]
    if status: q+=" AND j.status=?"; args.append(status)
    q+=" ORDER BY j.created_at DESC LIMIT ? OFFSET ?"; args += [limit,offset]
    with _db_lock,_db_connect() as conn: rows=conn.execute(q,args).fetchall()
    return [_row_to_dict(r) for r in rows]


def job_mark_cancelled(job_id: str) -> bool:
    with _db_lock, _db_connect() as conn:
        cur = conn.execute(
            """UPDATE jobs SET status=?, error=?, finished_at=?
               WHERE id=? AND status IN (?,?,?)""",
            (JobStatus.FAILED, "Dibatalkan pengguna.", time.time(), job_id,
             JobStatus.QUEUED, JobStatus.PROCESSING, JobStatus.PRINTING),
        )
        conn.commit()
        return cur.rowcount == 1


def job_counts() -> dict:
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("SELECT status, COUNT(*) c FROM jobs GROUP BY status").fetchall()
        out = {s: 0 for s in (JobStatus.QUEUED, JobStatus.PROCESSING, JobStatus.PRINTING,
                               JobStatus.SUCCESS, JobStatus.FAILED)}
        for r in rows:
            out[r["status"]] = r["c"]
        return out


def job_mark_failed(job_id: str, error: str) -> None:
    with _db_lock, _db_connect() as conn:
        conn.execute("UPDATE jobs SET status=?, error=?, finished_at=? WHERE id=?",
                     (JobStatus.FAILED, error, time.time(), job_id))
        conn.commit()


def job_sweep_stuck(timeout_sec: Optional[int] = None, max_retries: Optional[int] = None) -> int:
    """Recovery runtime untuk job DIRECT yang terlalu lama. Job yang sudah
    memperoleh Windows spool ID tidak pernah di-retry otomatis karena berisiko
    mencetak ganda; job yang belum masuk spooler boleh satu kali diulang."""
    timeout_sec = timeout_sec or settings.print_job_timeout_sec
    max_retries = settings.print_max_retries if max_retries is None else max_retries
    cutoff = time.time() - max(60, timeout_sec)
    with _db_lock, _db_connect() as conn:
        rows = conn.execute(
            "SELECT id, retries, spool_job_id FROM jobs WHERE method=? AND status IN (?,?) "
            "AND dispatched_at IS NOT NULL AND dispatched_at < ?",
            (PrintMethod.DIRECT, JobStatus.PROCESSING, JobStatus.PRINTING, cutoff),
        ).fetchall()
        n = 0
        for r in rows:
            if r["spool_job_id"]:
                conn.execute("UPDATE jobs SET status=?, error=?, finished_at=?, spool_status=? WHERE id=?",
                             (JobStatus.FAILED, "Timeout setelah job masuk Windows spooler; tidak diulang untuk mencegah duplicate print.",
                              time.time(), "Timeout/unknown", r["id"]))
            elif int(r["retries"] or 0) < max_retries:
                conn.execute("UPDATE jobs SET status=?, dispatched_at=NULL, started_at=NULL, retries=retries+1, error=? WHERE id=?",
                             (JobStatus.QUEUED, "Job timeout sebelum masuk spooler; dijadwalkan ulang.", r["id"]))
            else:
                conn.execute("UPDATE jobs SET status=?, error=?, finished_at=? WHERE id=?",
                             (JobStatus.FAILED, "Print lokal timeout.", time.time(), r["id"]))
            n += 1
        conn.commit(); return n


def job_cancel(job_id: str) -> bool:
    with _db_lock, _db_connect() as conn:
        cur = conn.execute(
            "UPDATE jobs SET status=?, error=?, finished_at=? WHERE id=? AND status=?",
            (JobStatus.FAILED, "Dibatalkan pengguna.", time.time(), job_id, JobStatus.QUEUED),
        )
        conn.commit()
        return cur.rowcount == 1


def job_clear_queued() -> int:
    with _db_lock, _db_connect() as conn:
        cur = conn.execute(
            "UPDATE jobs SET status=?, error=?, finished_at=? WHERE status=?",
            (JobStatus.FAILED, "Dibersihkan admin.", time.time(), JobStatus.QUEUED),
        )
        conn.commit()
        return cur.rowcount


def job_stats(top_n: int = 10) -> dict:
    """Agregat untuk panel Statistik (menu hamburger web / /stats bot):
    total sukses/gagal, total salinan tercetak, file paling sering
    diprint, pengirim paling aktif, dan metode print."""
    with _db_lock, _db_connect() as conn:
        totals = conn.execute(
            "SELECT status, COUNT(*) c FROM jobs GROUP BY status"
        ).fetchall()
        total_by_status = {s: 0 for s in (JobStatus.QUEUED, JobStatus.PROCESSING, JobStatus.PRINTING,
                                           JobStatus.SUCCESS, JobStatus.FAILED)}
        for r in totals:
            total_by_status[r["status"]] = r["c"]

        total_copies = conn.execute(
            "SELECT COALESCE(SUM(copies),0) c FROM jobs WHERE status=?", (JobStatus.SUCCESS,)
        ).fetchone()["c"]

        top_files = conn.execute(
            """SELECT file_name, COUNT(*) cnt, COALESCE(SUM(copies),0) copies_sum
               FROM jobs WHERE status=? GROUP BY file_name ORDER BY cnt DESC LIMIT ?""",
            (JobStatus.SUCCESS, top_n),
        ).fetchall()

        top_senders = conn.execute(
            """SELECT sender, COUNT(*) cnt FROM jobs WHERE status=?
               GROUP BY sender ORDER BY cnt DESC LIMIT ?""",
            (JobStatus.SUCCESS, top_n),
        ).fetchall()

        by_method = conn.execute(
            """SELECT method, COUNT(*) cnt FROM jobs WHERE status=? GROUP BY method""",
            (JobStatus.SUCCESS,),
        ).fetchall()

        first_row = conn.execute("SELECT MIN(created_at) t FROM jobs").fetchone()

    total_jobs = sum(total_by_status.values())
    return {
        "total_jobs": total_jobs,
        "total_success": total_by_status[JobStatus.SUCCESS],
        "total_failed": total_by_status[JobStatus.FAILED],
        "total_in_progress": total_by_status[JobStatus.QUEUED] + total_by_status[JobStatus.PROCESSING]
        + total_by_status[JobStatus.PRINTING],
        "total_copies_printed": total_copies,
        "top_files": [{"name": r["file_name"], "count": r["cnt"], "copies": r["copies_sum"]} for r in top_files],
        "top_senders": [{"sender": r["sender"], "count": r["cnt"]} for r in top_senders],
        "by_method": {r["method"]: r["cnt"] for r in by_method},
        "since": first_row["t"] if first_row else None,
    }




# GAS/Print Bridge telah dihapus dari alur aktif. Hanya snapshot lokal yang
# dipertahankan untuk riwayat dan reprint.
def archive_print_source(path: Path) -> Path:
    """Copy snapshot untuk reprint tanpa pernah memindahkan/mengubah file sumber."""
    settings.print_archive_dir.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()[:16]
    archived = settings.print_archive_dir / f"{uuid.uuid4().hex}{suffix}"
    shutil.copy2(path, archived)
    return archived


def resolve_print_archive(raw_path: str) -> Optional[Path]:
    if not raw_path:
        return None
    candidate = Path(raw_path).resolve(); archive_root = settings.print_archive_dir.resolve()
    try:
        candidate.relative_to(archive_root)
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


# ---------------------------------------------------------------------------
# PRINT DARI PERANGKAT - staging sementara untuk file drag/drop atau file picker
# ---------------------------------------------------------------------------
DEVICE_PRINT_UPLOAD_TTL_SEC = 2 * 3600

def _device_print_dir() -> Path:
    d = settings.data_dir / "uploads" / "device_print"
    d.mkdir(parents=True, exist_ok=True)
    return d

def _valid_device_token(token: str) -> str:
    token = (token or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{32}", token):
        raise FileManagerError("Token file perangkat tidak valid.")
    return token

def device_print_meta_path(token: str) -> Path:
    return _device_print_dir() / (_valid_device_token(token) + ".json")

def device_print_resolve(token: str) -> "tuple[Path, dict]":
    token = _valid_device_token(token); meta_path = device_print_meta_path(token)
    if not meta_path.is_file():
        raise FileManagerError("File perangkat sudah tidak tersedia/kedaluwarsa.")
    try:
        meta = json.loads(meta_path.read_text("utf-8"))
    except Exception:
        raise FileManagerError("Metadata file perangkat rusak.")
    ext = str(meta.get("ext") or "").lower()
    if ext not in PRINTABLE_EXTS:
        raise FileManagerError("Format file perangkat tidak didukung untuk print.")
    data_path = (_device_print_dir() / (token + ext)).resolve()
    try:
        data_path.relative_to(_device_print_dir().resolve())
    except ValueError:
        raise FileManagerError("Path staging file perangkat tidak valid.")
    if not data_path.is_file():
        raise FileManagerError("File perangkat sudah tidak tersedia/kedaluwarsa.")
    return data_path, meta

def device_print_remove(token: str) -> None:
    try:
        data_path, _meta = device_print_resolve(token)
        try: data_path.unlink()
        except OSError: pass
    except Exception:
        pass
    try: device_print_meta_path(token).unlink()
    except OSError: pass

def cleanup_device_print_uploads(max_age_sec: int = DEVICE_PRINT_UPLOAD_TTL_SEC) -> int:
    root = _device_print_dir(); cutoff = time.time() - max(300, int(max_age_sec)); removed = 0
    for meta_path in root.glob("*.json"):
        try:
            if meta_path.stat().st_mtime >= cutoff:
                continue
            token = meta_path.stem
            try:
                data_path, _ = device_print_resolve(token)
                try: data_path.unlink(); removed += 1
                except OSError: pass
            except Exception:
                pass
            try: meta_path.unlink()
            except OSError: pass
        except OSError:
            pass
    # Bersihkan orphan file yang tidak punya metadata.
    for f in root.iterdir() if root.exists() else []:
        try:
            if f.is_file() and f.suffix.lower() != ".json" and f.stat().st_mtime < cutoff:
                if not (root / (f.stem + ".json")).exists():
                    f.unlink(); removed += 1
        except OSError:
            pass
    return removed


db_init()


def _ensure_table_columns(conn: sqlite3.Connection, table: str, columns: Dict[str, str]) -> set:
    """Tambah kolom yang belum ada tanpa menghapus/mengganti data lama."""
    existing = {str(r["name"]) for r in conn.execute("PRAGMA table_info(%s)" % table).fetchall()}
    for name, ddl in columns.items():
        if name not in existing:
            conn.execute("ALTER TABLE %s ADD COLUMN %s %s" % (table, name, ddl))
            existing.add(name)
    return existing


def advanced_db_init() -> None:
    """Schema v5 yang kompatibel dengan DB WebApp lama dan PrintBot Telegram.

    PrintBot Telegram v7 memakai audit_log(created_at, actor_id, actor_name,
    details), sedangkan WebApp memakai audit_log(ts, username, ip, detail).
    Tabel audit dibuat sebagai union kedua schema agar satu printbot.db aman
    digunakan bersama tanpa error ``no such column``.
    """
    with _db_lock, _db_connect() as conn:
        conn.execute("CREATE TABLE IF NOT EXISTS app_meta (key TEXT PRIMARY KEY, value TEXT)")
        conn.execute("CREATE TABLE IF NOT EXISTS schema_migrations (version INTEGER PRIMARY KEY, applied_at REAL NOT NULL, app_version TEXT NOT NULL)")
        conn.execute("INSERT OR IGNORE INTO schema_migrations(version,applied_at,app_version) VALUES(?,?,?)",(DB_SCHEMA_VERSION,time.time(),APP_VERSION))

        # Audit union-schema: kompatibel dua arah dengan PrintBot Telegram v7.
        conn.execute("""CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ts REAL, created_at REAL,
            username TEXT, actor_id TEXT, actor_name TEXT, ip TEXT,
            action TEXT, target TEXT, detail TEXT, details TEXT,
            success INTEGER NOT NULL DEFAULT 1)""")
        audit_cols = _ensure_table_columns(conn, "audit_log", {
            "ts":"REAL", "created_at":"REAL", "username":"TEXT", "actor_id":"TEXT",
            "actor_name":"TEXT", "ip":"TEXT", "action":"TEXT", "target":"TEXT",
            "detail":"TEXT", "details":"TEXT", "success":"INTEGER NOT NULL DEFAULT 1",
        })
        # WebApp v4.0.0 pernah membuat `ts REAL NOT NULL`. Itu aman untuk WebApp
        # tetapi membuat Bot Telegram gagal INSERT karena bot hanya mengisi created_at.
        # Rebuild sekali dengan union-schema nullable sambil mempertahankan seluruh data.
        audit_info = conn.execute("PRAGMA table_info(audit_log)").fetchall()
        if any(str(r["name"]) == "ts" and int(r["notnull"] or 0) for r in audit_info):
            conn.execute("DROP TABLE IF EXISTS audit_log_legacy_v5")
            conn.execute("ALTER TABLE audit_log RENAME TO audit_log_legacy_v5")
            conn.execute("""CREATE TABLE audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT, ts REAL, created_at REAL,
                username TEXT, actor_id TEXT, actor_name TEXT, ip TEXT, action TEXT,
                target TEXT, detail TEXT, details TEXT, success INTEGER NOT NULL DEFAULT 1)""")
            conn.execute("""INSERT INTO audit_log(id,ts,created_at,username,actor_id,actor_name,ip,action,target,detail,details,success)
                SELECT id,ts,created_at,username,actor_id,actor_name,ip,action,target,detail,details,COALESCE(success,1)
                FROM audit_log_legacy_v5""")
            conn.execute("DROP TABLE audit_log_legacy_v5")
            audit_cols = {str(r["name"]) for r in conn.execute("PRAGMA table_info(audit_log)").fetchall()}
        # Normalisasi data lama tanpa merusak kolom asal.
        if "created_at" in audit_cols:
            conn.execute("UPDATE audit_log SET ts=COALESCE(ts, created_at, ?) WHERE ts IS NULL", (time.time(),))
        else:
            conn.execute("UPDATE audit_log SET ts=COALESCE(ts, ?) WHERE ts IS NULL", (time.time(),))
        conn.execute("UPDATE audit_log SET created_at=COALESCE(created_at, ts, ?) WHERE created_at IS NULL", (time.time(),))
        conn.execute("UPDATE audit_log SET username=COALESCE(NULLIF(username,''), NULLIF(actor_name,''), NULLIF(actor_id,''), '')")
        conn.execute("UPDATE audit_log SET actor_name=COALESCE(NULLIF(actor_name,''), NULLIF(username,''), '')")
        conn.execute("UPDATE audit_log SET actor_id=COALESCE(NULLIF(actor_id,''), NULLIF(username,''), '')")
        conn.execute("UPDATE audit_log SET detail=COALESCE(detail, details, '')")
        conn.execute("UPDATE audit_log SET details=COALESCE(details, detail, '')")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_ts ON audit_log(ts DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_log(created_at DESC)")

        conn.execute("CREATE TABLE IF NOT EXISTS error_log (id INTEGER PRIMARY KEY AUTOINCREMENT, ts REAL, category TEXT, message TEXT, detail TEXT)")
        _ensure_table_columns(conn, "error_log", {"ts":"REAL", "category":"TEXT", "message":"TEXT", "detail":"TEXT"})
        conn.execute("UPDATE error_log SET ts=COALESCE(ts, ?) WHERE ts IS NULL", (time.time(),))
        conn.execute("CREATE INDEX IF NOT EXISTS idx_error_ts ON error_log(ts DESC)")
        conn.execute("CREATE TABLE IF NOT EXISTS file_index (root TEXT NOT NULL, rel_path TEXT NOT NULL, name TEXT NOT NULL, ext TEXT, size INTEGER, modified REAL, indexed_at REAL NOT NULL, PRIMARY KEY(root, rel_path))")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_file_index_name ON file_index(name COLLATE NOCASE)")
        conn.execute("CREATE TABLE IF NOT EXISTS favorites (username TEXT NOT NULL, root TEXT NOT NULL, rel_path TEXT NOT NULL, created_at REAL NOT NULL, PRIMARY KEY(username, root, rel_path))")
        conn.execute("CREATE TABLE IF NOT EXISTS recent_files (username TEXT NOT NULL, root TEXT NOT NULL, rel_path TEXT NOT NULL, action TEXT NOT NULL, ts REAL NOT NULL, PRIMARY KEY(username, root, rel_path, action))")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_recent_user_ts ON recent_files(username, ts DESC)")
        conn.execute("CREATE TABLE IF NOT EXISTS printer_profiles (name TEXT PRIMARY KEY, printer TEXT, paper TEXT, orientation TEXT, scale_mode TEXT, duplex TEXT, color_mode TEXT, copies INTEGER NOT NULL DEFAULT 1, updated_at REAL NOT NULL)")
        conn.execute("INSERT OR REPLACE INTO app_meta(key,value) VALUES('db_schema_version',?)", (str(DB_SCHEMA_VERSION),))
        conn.execute("INSERT OR REPLACE INTO app_meta(key,value) VALUES('config_version',?)", (str(CONFIG_VERSION),))
        conn.commit()


def app_setting_get(key: str, default: str = "") -> str:
    with _db_lock, _db_connect() as conn:
        row = conn.execute("SELECT value FROM app_meta WHERE key=?", (key,)).fetchone()
    return str(row["value"]) if row and row["value"] is not None else default


def app_setting_set(key: str, value: Any) -> None:
    with _db_lock, _db_connect() as conn:
        conn.execute("INSERT OR REPLACE INTO app_meta(key,value) VALUES(?,?)", (key, str(value)))
        conn.commit()


def audit_event(username: str, action: str, target: str = "", detail: str = "", success: bool = True, ip: str = "") -> None:
    try:
        now = time.time(); actor = username or ""
        with _db_lock, _db_connect() as conn:
            # Isi kedua keluarga kolom agar audit tetap dapat dibaca WebApp dan Bot Telegram.
            conn.execute("""INSERT INTO audit_log(
                ts,created_at,username,actor_id,actor_name,ip,action,target,detail,details,success
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                         (now, now, actor, actor, actor, ip or "", action, target or "",
                          detail or "", detail or "", 1 if success else 0))
            conn.commit()
    except Exception:
        logging.getLogger("webapp").exception("audit_event gagal")


def audit_list(limit: int = 200) -> List[dict]:
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT ?", (max(1, min(limit, 1000)),)).fetchall()
    return [dict(r) for r in rows]


def record_error(category: str, message: str, detail: str = "") -> None:
    try:
        with _db_lock, _db_connect() as conn:
            conn.execute("INSERT INTO error_log(ts,category,message,detail) VALUES(?,?,?,?)",
                         (time.time(), category[:80], str(message)[:2000], str(detail)[:8000]))
            conn.commit()
    except Exception:
        pass


def error_list(limit: int = 200) -> List[dict]:
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("SELECT * FROM error_log ORDER BY id DESC LIMIT ?", (max(1, min(limit, 1000)),)).fetchall()
    return [dict(r) for r in rows]


def is_maintenance_mode() -> bool:
    return app_setting_get("maintenance_mode", "0") == "1"


def set_maintenance_mode(enabled: bool) -> None:
    app_setting_set("maintenance_mode", "1" if enabled else "0")


def favorite_toggle(username: str, root: str, rel_path: str) -> bool:
    now = time.time()
    with _db_lock, _db_connect() as conn:
        row = conn.execute("SELECT 1 FROM favorites WHERE username=? AND root=? AND rel_path=?", (username, root, rel_path)).fetchone()
        if row:
            conn.execute("DELETE FROM favorites WHERE username=? AND root=? AND rel_path=?", (username, root, rel_path)); state = False
        else:
            conn.execute("INSERT INTO favorites(username,root,rel_path,created_at) VALUES(?,?,?,?)", (username, root, rel_path, now)); state = True
        conn.commit()
    return state


def favorite_list(username: str, limit: int = 100) -> List[dict]:
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("""SELECT f.username,f.root,f.rel_path,f.created_at,i.name,i.ext,i.size,i.modified
                               FROM favorites f LEFT JOIN file_index i ON i.root=f.root AND i.rel_path=f.rel_path
                               WHERE f.username=? ORDER BY f.created_at DESC LIMIT ?""", (username, limit)).fetchall()
    out=[]
    for r in rows:
        d=dict(r); d["name"]=d.get("name") or Path(d["rel_path"]).name; d["ext"]=d.get("ext") or Path(d["rel_path"]).suffix.lower(); d["size"]=d.get("size") or 0; d["modified"]=d.get("modified") or 0; out.append(d)
    return out


def recent_touch(username: str, root: str, rel_path: str, action: str) -> None:
    with _db_lock, _db_connect() as conn:
        conn.execute("INSERT OR REPLACE INTO recent_files(username,root,rel_path,action,ts) VALUES(?,?,?,?,?)", (username, root, rel_path, action, time.time()))
        conn.commit()


def recent_list(username: str, limit: int = 50) -> List[dict]:
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("""SELECT r.username,r.root,r.rel_path,r.action,r.ts,i.name,i.ext,i.size,i.modified
                               FROM recent_files r LEFT JOIN file_index i ON i.root=r.root AND i.rel_path=r.rel_path
                               WHERE r.username=? ORDER BY r.ts DESC LIMIT ?""", (username, limit)).fetchall()
    out=[]
    for r in rows:
        d=dict(r); d["name"]=d.get("name") or Path(d["rel_path"]).name; d["ext"]=d.get("ext") or Path(d["rel_path"]).suffix.lower(); d["size"]=d.get("size") or 0; d["modified"]=d.get("modified") or d.get("ts") or 0; out.append(d)
    return out


def printer_profile_save(name: str, profile: dict) -> None:
    name = (name or "").strip()
    if not name: raise ValueError("Nama profil wajib diisi.")
    with _db_lock, _db_connect() as conn:
        conn.execute("""INSERT OR REPLACE INTO printer_profiles(name,printer,paper,orientation,scale_mode,duplex,color_mode,copies,updated_at)
                        VALUES(?,?,?,?,?,?,?,?,?)""",
                     (name, profile.get("printer") or "", profile.get("paper") or "AUTO", profile.get("orientation") or "AUTO",
                      profile.get("scale_mode") or "FIT", profile.get("duplex") or "DEFAULT", profile.get("color_mode") or "DEFAULT",
                      max(1, min(int(profile.get("copies") or 1), 99)), time.time()))
        conn.commit()


def printer_profile_list() -> List[dict]:
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("SELECT * FROM printer_profiles ORDER BY name COLLATE NOCASE").fetchall()
    return [dict(r) for r in rows]


def printer_profile_remove(name: str) -> bool:
    with _db_lock, _db_connect() as conn:
        cur = conn.execute("DELETE FROM printer_profiles WHERE name=?", (name,)); conn.commit(); return cur.rowcount == 1


def print_quota_status(username: str) -> dict:
    now=time.time(); hour=now-3600; day=now-86400
    with _db_lock, _db_connect() as conn:
        jobs_hour=conn.execute("SELECT COUNT(*) c FROM jobs WHERE sender=? AND created_at>=?",(username,hour)).fetchone()["c"]
        copies_day=conn.execute("SELECT COALESCE(SUM(copies),0) c FROM jobs WHERE sender=? AND created_at>=? AND status<>?",(username,day,JobStatus.FAILED)).fetchone()["c"]
    return {"jobs_last_hour":int(jobs_hour or 0),"copies_last_day":int(copies_day or 0),
            "max_jobs_per_hour":settings.max_jobs_per_hour,"max_copies_per_day":settings.max_copies_per_day,
            "max_copies_per_job":settings.max_copies_per_job}


def enforce_print_quota(username: str, copies: int, is_admin: bool = False) -> None:
    if is_admin: return
    q=print_quota_status(username)
    if copies > max(1,settings.max_copies_per_job):
        raise FileManagerError("Maksimal %s copy per job." % settings.max_copies_per_job)
    if q["jobs_last_hour"] >= max(1,settings.max_jobs_per_hour):
        raise FileManagerError("Batas %s job per jam tercapai." % settings.max_jobs_per_hour)
    if q["copies_last_day"] + copies > max(1,settings.max_copies_per_day):
        raise FileManagerError("Batas %s copy per hari terlampaui." % settings.max_copies_per_day)


advanced_db_init()


# ============================================================================
# ROOT FOLDER MANAGEMENT - folder statis dari .env (FILE_MANAGER_ROOTS) +
# folder dinamis yang ditambahkan lewat UI/API, tersimpan di
# data/roots.json. File ini DIBAGI dengan bot.py sehingga menambah folder
# lewat WebApp otomatis langsung tampil di Telegram Bot juga (dan
# sebaliknya) - tanpa restart, karena selalu dibaca ulang tiap dipanggil.
# ============================================================================
_roots_lock = threading.Lock()


def _roots_json_path() -> Path:
    return settings.data_dir / "roots.json"


def roots_dynamic_load() -> Dict[str, str]:
    p = _roots_json_path()
    if p.exists():
        try:
            data = json.loads(p.read_text("utf-8"))
            return {str(k): str(v) for k, v in data.items()}
        except Exception:
            return {}
    return {}


def roots_dynamic_save(data: Dict[str, str]) -> None:
    _roots_json_path().write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def get_roots() -> Dict[str, str]:
    """Root statis (.env) digabung root dinamis (roots.json); root dinamis
    bisa menambah ATAU menimpa label yang sama dari .env."""
    with _roots_lock:
        merged = dict(settings.roots)
        merged.update(roots_dynamic_load())
        return merged


def root_is_dynamic(label: str) -> bool:
    with _roots_lock:
        return label in roots_dynamic_load()


def root_add(label: str, path: str = "") -> dict:
    """Tambah folder root. Kalau `path` dikosongkan, folder dibuat OTOMATIS
    (mkdir) di dalam data/roots/<label> - jadi admin tidak perlu tahu/isi
    path absolut sama sekali. Kalau `path` diisi (path absolut ke folder
    yang sudah ada ATAU belum), folder itu juga akan dibuat otomatis kalau
    belum ada (mkdir -p), lalu langsung diizinkan diakses."""
    label = (label or "").strip()
    if not label:
        raise FileManagerError("Nama folder tidak boleh kosong.")
    if any(c in label for c in "/\\:"):
        raise FileManagerError("Nama folder tidak boleh mengandung '/', '\\', atau ':'.")

    path = (path or "").strip()
    auto_created_under_base = False
    if path:
        try:
            p = Path(path).expanduser().resolve()
        except Exception:
            raise FileManagerError("Path tidak valid.")
    else:
        p = (settings.roots_base_dir / label).resolve()
        auto_created_under_base = True

    try:
        p.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        raise FileManagerError(f"Gagal membuat folder di server: {e}")

    with _roots_lock:
        data = roots_dynamic_load()
        data[label] = str(p)
        roots_dynamic_save(data)
    log.info("ROOT ADD label=%s path=%s auto_created=%s", label, p, auto_created_under_base)
    return {"label": label, "path": str(p), "auto_created": auto_created_under_base}


def root_remove(label: str) -> bool:
    with _roots_lock:
        data = roots_dynamic_load()
        if label not in data:
            return False
        del data[label]
        roots_dynamic_save(data)
    log.info("ROOT REMOVE label=%s", label)
    return True


# ============================================================================
# FILE INDEX - pencarian cepat SQLite + refresh background
# ============================================================================
def rebuild_file_index(root_key: Optional[str] = None) -> dict:
    roots = get_roots(); selected = {root_key: roots[root_key]} if root_key and root_key in roots else roots
    indexed = 0; errors = 0; now = time.time()
    with _db_lock, _db_connect() as conn:
        if root_key:
            conn.execute("DELETE FROM file_index WHERE root=?", (root_key,))
        else:
            conn.execute("DELETE FROM file_index")
        batch = []
        for rk, raw in selected.items():
            base = Path(raw)
            if not base.is_dir():
                errors += 1; continue
            try:
                for f in base.rglob("*"):
                    try:
                        if not f.is_file(): continue
                        st = f.stat(); rel = str(f.relative_to(base)).replace("\\", "/")
                        batch.append((rk, rel, f.name, f.suffix.lower(), st.st_size, st.st_mtime, now)); indexed += 1
                        if len(batch) >= 1000:
                            conn.executemany("INSERT OR REPLACE INTO file_index(root,rel_path,name,ext,size,modified,indexed_at) VALUES(?,?,?,?,?,?,?)", batch); batch=[]
                    except OSError:
                        errors += 1
            except (OSError, PermissionError):
                errors += 1
        if batch:
            conn.executemany("INSERT OR REPLACE INTO file_index(root,rel_path,name,ext,size,modified,indexed_at) VALUES(?,?,?,?,?,?,?)", batch)
        conn.commit()
    app_setting_set("file_index_last_build", str(now))
    return {"indexed": indexed, "errors": errors, "at": now}


def indexed_search(root_key: Optional[str], keyword: str, max_results: Optional[int] = None) -> List[dict]:
    if not keyword or len(keyword.strip()) < 2:
        raise FileManagerError("Kata kunci minimal 2 karakter.")
    max_results = max_results or settings.search_max_results
    tokens = [t.lower() for t in re.split(r"[\s_\-.()]+", keyword.strip()) if t]
    if not tokens: return []
    where = ["LOWER(name) LIKE ?" for _ in tokens]; args: list = ["%%%s%%" % t for t in tokens]
    sql = "SELECT root,name,rel_path,size,modified,ext FROM file_index WHERE " + " AND ".join(where)
    if root_key:
        sql += " AND root=?"; args.append(root_key)
    sql += " ORDER BY name COLLATE NOCASE LIMIT ?"; args.append(max_results)
    with _db_lock, _db_connect() as conn:
        rows = conn.execute(sql, args).fetchall()
    if rows:
        return [dict(r) for r in rows]
    # Fallback jika index belum sempat dibangun.
    pattern = fm_build_search_pattern(keyword); roots = [root_key] if root_key else list(get_roots().keys()); out=[]
    for rk in roots:
        try: base=Path(fm_root_path(rk))
        except FileManagerError: continue
        for f in base.rglob("*"):
            if len(out)>=max_results: break
            try:
                if f.is_file() and pattern.search(f.name):
                    st=f.stat(); out.append({"root":rk,"name":f.name,"rel_path":fm_to_rel(rk,f),"size":st.st_size,"modified":st.st_mtime,"ext":f.suffix.lower()})
            except OSError: pass
    return sorted(out,key=lambda x:x["name"].lower())[:max_results]


def file_index_count() -> int:
    with _db_lock, _db_connect() as conn:
        return int(conn.execute("SELECT COUNT(*) c FROM file_index").fetchone()["c"] or 0)


_WATCHDOG_OBSERVERS: List[Any] = []

def _index_event_path(root_key: str, base: Path, raw_path: str) -> None:
    try:
        p=Path(raw_path); rel=str(p.resolve().relative_to(base.resolve())).replace("\\","/")
    except Exception:
        return
    with _db_lock, _db_connect() as conn:
        if p.is_file():
            try:
                st=p.stat(); conn.execute("INSERT OR REPLACE INTO file_index(root,rel_path,name,ext,size,modified,indexed_at) VALUES(?,?,?,?,?,?,?)",
                                          (root_key,rel,p.name,p.suffix.lower(),st.st_size,st.st_mtime,time.time()))
            except OSError: pass
        else:
            conn.execute("DELETE FROM file_index WHERE root=? AND (rel_path=? OR rel_path LIKE ?)",(root_key,rel,rel.rstrip("/")+"/%"))
        conn.commit()

def start_file_watchers() -> int:
    if _WATCHDOG_OBSERVERS: return len(_WATCHDOG_OBSERVERS)
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except Exception:
        try:
            pkg="watchdog<5" if sys.version_info < (3,9) else "watchdog"
            subprocess.run([sys.executable,"-m","pip","install","--disable-pip-version-check","-q",pkg],timeout=180,stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL,check=False)
            import importlib; importlib.invalidate_caches()
            from watchdog.observers import Observer
            from watchdog.events import FileSystemEventHandler
        except Exception:
            log.warning("watchdog tidak tersedia; file index tetap direfresh berkala.")
            return 0
    for rk,raw in get_roots().items():
        base=Path(raw)
        if not base.is_dir(): continue
        class Handler(FileSystemEventHandler):
            def on_created(self,event,_rk=rk,_base=base):
                if not event.is_directory: _index_event_path(_rk,_base,event.src_path)
            def on_modified(self,event,_rk=rk,_base=base):
                if not event.is_directory: _index_event_path(_rk,_base,event.src_path)
            def on_deleted(self,event,_rk=rk,_base=base): _index_event_path(_rk,_base,event.src_path)
            def on_moved(self,event,_rk=rk,_base=base):
                _index_event_path(_rk,_base,event.src_path); _index_event_path(_rk,_base,event.dest_path)
        try:
            obs=Observer(); obs.schedule(Handler(),str(base),recursive=True); obs.daemon=True; obs.start(); _WATCHDOG_OBSERVERS.append(obs)
        except Exception as exc:
            record_error("FILE_WATCHER",str(exc),"root=%s"%rk)
    return len(_WATCHDOG_OBSERVERS)

def stop_file_watchers() -> None:
    for obs in list(_WATCHDOG_OBSERVERS):
        try: obs.stop()
        except Exception: pass
    for obs in list(_WATCHDOG_OBSERVERS):
        try: obs.join(timeout=3)
        except Exception: pass
    _WATCHDOG_OBSERVERS.clear()

def restart_file_watchers() -> int:
    stop_file_watchers(); return start_file_watchers()

async def file_index_loop() -> None:
    first=True
    while True:
        try:
            await _to_thread(rebuild_file_index, None)
            if first:
                watchers=await _to_thread(start_file_watchers); log.info("File index watcher aktif: %s root",watchers); first=False
        except asyncio.CancelledError: raise
        except Exception as exc:
            record_error("FILE_INDEX", str(exc)); log.exception("File index refresh gagal")
        await asyncio.sleep(max(60, settings.index_refresh_sec))


# ============================================================================
# LOCAL PRINT ENGINE - Windows printer lokal / CUPS. Tidak memakai Bridge/GAS.
# ============================================================================
def _win_import_print_modules():
    try:
        import win32api
        import win32con
        import win32print
        import win32ui
        return win32api, win32con, win32print, win32ui
    except Exception:
        return None, None, None, None


def local_print_available() -> bool:
    if sys.platform.startswith("win"):
        _a, _c, wp, _u = _win_import_print_modules()
        return wp is not None
    return _tool_find("lp") is not None or _tool_find("lpr") is not None


def local_list_printers() -> List[str]:
    """Daftar printer lokal/network connection. Tidak memakai Get-Printer agar Windows 7 aman."""
    if sys.platform.startswith("win"):
        try:
            _a, _c, win32print, _u = _win_import_print_modules()
            if win32print is not None:
                flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
                rows = win32print.EnumPrinters(flags, None, 2) or []
                names = [str(r.get("pPrinterName") or "").strip() for r in rows]
                return sorted(set(x for x in names if x), key=lambda x: x.lower())
        except Exception as exc:
            log.warning("EnumPrinters gagal: %s", exc)
        try:
            r = subprocess.run(["wmic", "printer", "get", "name"], stdout=subprocess.PIPE,
                               stderr=subprocess.PIPE, universal_newlines=True, timeout=15)
            lines = [x.strip() for x in r.stdout.splitlines() if x.strip()]
            return [x for x in lines if x.lower() != "name"]
        except Exception:
            return []
    try:
        r = subprocess.run(["lpstat", "-p"], stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                           universal_newlines=True, timeout=10)
        names = []
        for ln in r.stdout.splitlines():
            m = re.match(r"printer\s+(\S+)\s+is", ln)
            if m:
                names.append(m.group(1))
        return names
    except Exception:
        return []


_runtime_lock = threading.Lock()

def _runtime_settings_path() -> Path:
    return settings.data_dir / "runtime_settings.json"

def runtime_settings_load() -> dict:
    with _runtime_lock:
        p = _runtime_settings_path()
        try:
            data = json.loads(p.read_text("utf-8")) if p.is_file() else {}
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

def runtime_settings_update(**values) -> dict:
    with _runtime_lock:
        p = _runtime_settings_path()
        try:
            data = json.loads(p.read_text("utf-8")) if p.is_file() else {}
            if not isinstance(data, dict): data = {}
        except Exception:
            data = {}
        data.update(values)
        tmp = p.with_suffix(".tmp"); tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"); os.replace(str(tmp), str(p))
        return data


def local_default_printer() -> str:
    runtime_default = str(runtime_settings_load().get("default_printer") or "").strip()
    if runtime_default:
        return runtime_default
    if settings.default_printer:
        return settings.default_printer
    if sys.platform.startswith("win"):
        try:
            _a, _c, win32print, _u = _win_import_print_modules()
            if win32print is not None:
                return str(win32print.GetDefaultPrinter() or "")
        except Exception:
            pass
    return ""


def printer_status_details(printer_name: str) -> dict:
    result = {"printer": printer_name, "online": False, "problem": True, "flags": ["UNKNOWN"],
              "jobs": 0, "status_code": 0, "message": "Printer tidak dapat dibaca", "default": False}
    if not printer_name:
        return result
    if not sys.platform.startswith("win"):
        result.update({"online": True, "problem": False, "flags": ["AVAILABLE"], "message": "Available"})
        return result
    try:
        _a, _c, win32print, _u = _win_import_print_modules()
        if win32print is None:
            result["message"] = "pywin32 tidak tersedia"
            return result
        h = win32print.OpenPrinter(printer_name)
        try:
            info = win32print.GetPrinter(h, 2)
            status = int(info.get("Status") or 0)
            jobs = int(info.get("cJobs") or 0)
        finally:
            win32print.ClosePrinter(h)
        defs = [
            (0x00000001, "PAUSED"), (0x00000002, "ERROR"), (0x00000008, "PAPER_JAM"),
            (0x00000010, "PAPER_OUT"), (0x00000040, "PAPER_PROBLEM"), (0x00000080, "OFFLINE"),
            (0x00000400, "PRINTING"), (0x00000800, "OUTPUT_BIN_FULL"), (0x00001000, "NOT_AVAILABLE"),
            (0x00020000, "TONER_LOW"), (0x00040000, "NO_TONER"), (0x00100000, "USER_INTERVENTION"),
            (0x00200000, "OUT_OF_MEMORY"), (0x00400000, "DOOR_OPEN"), (0x01000000, "POWER_SAVE"),
        ]
        flags = [label for bit, label in defs if status & bit] or ["READY"]
        problems = {"PAUSED", "ERROR", "PAPER_JAM", "PAPER_OUT", "PAPER_PROBLEM", "OFFLINE",
                    "OUTPUT_BIN_FULL", "NOT_AVAILABLE", "NO_TONER", "USER_INTERVENTION",
                    "OUT_OF_MEMORY", "DOOR_OPEN"}
        problem = bool(set(flags) & problems)
        result.update({"online": "OFFLINE" not in flags, "problem": problem, "flags": flags, "jobs": jobs,
                       "status_code": status, "message": ", ".join(flags),
                       "default": printer_name == local_default_printer()})
        return result
    except Exception as exc:
        result["message"] = str(exc)
        return result


def local_printer_details() -> List[dict]:
    return [printer_status_details(p) for p in local_list_printers()]


_PRINTER_LAST_STATE: Dict[str, str] = {}

async def printer_monitor_loop() -> None:
    while True:
        try:
            printers=await _to_thread(local_list_printers)
            current=set(printers)
            for old in list(_PRINTER_LAST_STATE):
                if old not in current:
                    prev=_PRINTER_LAST_STATE.pop(old,None)
                    if prev is not None:
                        await event_publish({"type":"printer","printer":old,"state":"REMOVED","problem":True,"message":"Printer tidak lagi terdeteksi"})
            for printer in printers:
                st=await _to_thread(printer_status_details,printer); key="|".join(st.get("flags") or [])
                prev=_PRINTER_LAST_STATE.get(printer); _PRINTER_LAST_STATE[printer]=key
                if (prev is None and st.get("problem")) or (prev is not None and prev != key):
                    action="problem" if st.get("problem") else "recovered"
                    audit_event("system","PRINTER_"+action.upper(),target=printer,detail=st.get("message") or "")
                    await event_publish({"type":"printer","action":action,**st})
                    if settings.telegram_bot_token and settings.telegram_admin_chat_id:
                        icon="⚠️" if st.get("problem") else "✅"
                        await _to_thread(notify_admin_text,"%s PrintBot Printer\n%s\nStatus: %s"%(icon,printer,st.get("message") or action))
        except asyncio.CancelledError: raise
        except Exception as exc:
            record_error("PRINTER_MONITOR",str(exc))
        await asyncio.sleep(max(5,settings.printer_monitor_sec))


def _parse_page_numbers(pages: str, total_pages: int = 0) -> List[int]:
    raw = (pages or "").replace(" ", "")
    if not raw:
        return list(range(1, total_pages + 1)) if total_pages else []
    out: List[int] = []
    for part in raw.split(","):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            if not (a.isdigit() and b.isdigit()):
                raise ValueError("Format halaman tidak valid: %s" % part)
            lo, hi = int(a), int(b)
            if lo > hi:
                lo, hi = hi, lo
            out.extend(range(max(1, lo), hi + 1))
        elif part.isdigit():
            out.append(max(1, int(part)))
        else:
            raise ValueError("Format halaman tidak valid: %s" % part)
    if total_pages:
        out = [n for n in out if n <= total_pages]
    return list(dict.fromkeys(out))


def _printer_devmode(printer_name: str, paper: str = "AUTO", orientation: str = "AUTO",
                     duplex: str = "DEFAULT", color_mode: str = "DEFAULT"):
    _a, _c, win32print, _u = _win_import_print_modules()
    if win32print is None:
        return None, None
    handle = win32print.OpenPrinter(printer_name)
    try:
        info = win32print.GetPrinter(handle, 2)
        dm = info.get("pDevMode")
        if dm is None:
            return info, None
        try:
            if orientation == "PORTRAIT": dm.Orientation = 1
            elif orientation == "LANDSCAPE": dm.Orientation = 2
        except Exception: pass
        paper_map = {"A4": 9, "LETTER": 1, "LEGAL": 5, "F4": 14}
        if paper in paper_map:
            try: dm.PaperSize = paper_map[paper]
            except Exception: pass
        duplex_map = {"OFF": 1, "LONG": 2, "SHORT": 3}
        if duplex in duplex_map:
            try: dm.Duplex = duplex_map[duplex]
            except Exception: pass
        if color_mode in ("COLOR", "MONO"):
            try: dm.Color = 2 if color_mode == "COLOR" else 1
            except Exception: pass
        return info, dm
    finally:
        win32print.ClosePrinter(handle)


def _create_printer_dc(printer_name: str, paper: str = "AUTO", orientation: str = "AUTO",
                       duplex: str = "DEFAULT", color_mode: str = "DEFAULT"):
    _a, _c, win32print, win32ui = _win_import_print_modules()
    if win32print is None or win32ui is None:
        raise RuntimeError("pywin32 belum tersedia.")
    printer_name = printer_name or local_default_printer() or win32print.GetDefaultPrinter()
    try:
        import win32gui
        info, dm = _printer_devmode(printer_name, paper, orientation, duplex, color_mode)
        if info and dm is not None:
            hdc = win32gui.CreateDC(info.get("pDriverName"), printer_name, None, dm)
            return win32ui.CreateDCFromHandle(hdc), printer_name
    except Exception as exc:
        log.debug("DEVMODE printer fallback ke default driver: %s", exc)
    dc = win32ui.CreateDC(); dc.CreatePrinterDC(printer_name)
    return dc, printer_name


def _draw_pil_page(dc, image, scale_mode: str = "FIT", source_dpi: int = 144) -> None:
    from PIL import ImageWin
    printable_w = dc.GetDeviceCaps(8); printable_h = dc.GetDeviceCaps(10)
    dpi_x = max(1, dc.GetDeviceCaps(88)); dpi_y = max(1, dc.GetDeviceCaps(90))
    img = image.convert("RGB")
    if (scale_mode or "FIT").upper() == "ACTUAL":
        draw_w = max(1, int(img.width * float(dpi_x) / max(1, source_dpi)))
        draw_h = max(1, int(img.height * float(dpi_y) / max(1, source_dpi)))
        ratio = min(1.0, float(printable_w) / draw_w, float(printable_h) / draw_h)
    else:
        ratio = min(float(printable_w) / max(1, img.width), float(printable_h) / max(1, img.height))
        draw_w, draw_h = img.width, img.height
    draw_w = max(1, int(draw_w * ratio)); draw_h = max(1, int(draw_h * ratio))
    x1 = max(0, int((printable_w - draw_w) / 2)); y1 = max(0, int((printable_h - draw_h) / 2))
    ImageWin.Dib(img).draw(dc.GetHandleOutput(), (x1, y1, x1 + draw_w, y1 + draw_h))


def _spooler_jobs(printer_name: str) -> dict:
    if not sys.platform.startswith("win") or not printer_name:
        return {}
    try:
        _a, _c, win32print, _u = _win_import_print_modules()
        if win32print is None: return {}
        h = win32print.OpenPrinter(printer_name)
        try:
            rows = win32print.EnumJobs(h, 0, 999, 1) or []
            return {int(r.get("JobId")): r for r in rows if r.get("JobId") is not None}
        finally:
            win32print.ClosePrinter(h)
    except Exception:
        return {}


def _spool_status_text(row: dict) -> str:
    raw = str(row.get("pStatus") or "").strip()
    if raw: return raw
    status = int(row.get("Status") or 0)
    return "Spooling/Printing" if not status else "Windows status 0x%X" % status


def _wait_spool_job(printer_name: str, spool_job_id: int, timeout: int = 180, status_callback=None) -> "tuple[bool, str]":
    if not spool_job_id or not printer_name:
        return True, "Submitted"
    deadline = time.time() + max(10, timeout); seen = False; grace = time.time() + 5.0
    try:
        _a, _c, win32print, _u = _win_import_print_modules()
        error_bits = 0
        for name in ("JOB_STATUS_ERROR", "JOB_STATUS_OFFLINE", "JOB_STATUS_PAPEROUT", "JOB_STATUS_BLOCKED_DEVQ",
                     "JOB_STATUS_USER_INTERVENTION", "JOB_STATUS_DELETED"):
            error_bits |= int(getattr(win32print, name, 0) or 0)
    except Exception:
        error_bits = 0
    while time.time() < deadline:
        pstat = printer_status_details(printer_name)
        if pstat.get("problem"):
            msg = "Printer bermasalah: %s" % pstat.get("message", "Unknown")
            if status_callback: status_callback(msg)
            return False, msg
        row = _spooler_jobs(printer_name).get(int(spool_job_id))
        if row is None:
            if seen or time.time() >= grace:
                if status_callback: status_callback("Completed")
                return True, "Completed"
            time.sleep(0.35); continue
        seen = True
        txt = _spool_status_text(row)
        if status_callback: status_callback(txt)
        st = int(row.get("Status") or 0)
        if error_bits and (st & error_bits):
            return False, "Windows spooler bermasalah: %s" % txt
        time.sleep(0.8)
    return False, "Timeout menunggu Windows spooler menyelesaikan job."


def cancel_windows_spool_job(printer_name: str, spool_job_id: int) -> "tuple[bool, str]":
    if not sys.platform.startswith("win") or not printer_name or not spool_job_id:
        return False, "Windows Job ID/printer tidak tersedia."
    try:
        _a, _c, win32print, _u = _win_import_print_modules()
        h = win32print.OpenPrinter(printer_name)
        try:
            try: win32print.GetJob(h, int(spool_job_id), 1)
            except Exception: return False, "Windows spool job sudah tidak ditemukan (mungkin sudah selesai)."
            try:
                win32print.SetJob(h, int(spool_job_id), 0, None, getattr(win32print, "JOB_CONTROL_CANCEL", 3))
            except Exception:
                win32print.SetJob(h, int(spool_job_id), 0, None, getattr(win32print, "JOB_CONTROL_DELETE", 5))
            return True, "Windows spool job #%s dibatalkan." % spool_job_id
        finally:
            win32print.ClosePrinter(h)
    except Exception as exc:
        return False, "Gagal membatalkan Windows spool job: %s" % exc


def _print_pdf_windows(path: Path, printer: str, copies: int, pages: str,
                       paper: str = "AUTO", orientation: str = "AUTO", scale_mode: str = "FIT",
                       duplex: str = "DEFAULT", color_mode: str = "DEFAULT", spool_callback=None) -> "tuple[bool, str, int]":
    try:
        import fitz
        from PIL import Image
        render_dpi = max(96, min(int(settings.print_render_dpi), 300))
        doc = fitz.open(str(path)); dc = None; spool_id = 0
        try:
            selected = _parse_page_numbers(pages, doc.page_count) or list(range(1, doc.page_count + 1))
            if not selected: return False, "PDF tidak memiliki halaman yang dapat dicetak.", 0
            dc, printer = _create_printer_dc(printer, paper, orientation, duplex, color_mode)
            result = dc.StartDoc("PrintBot Web - %s" % path.name)
            try: spool_id = int(result or 0)
            except Exception: spool_id = 0
            if spool_id and spool_callback: spool_callback(spool_id)
            zoom = float(render_dpi) / 72.0
            for _copy in range(max(1, int(copies))):
                for page_no in selected:
                    page = doc.load_page(page_no - 1)
                    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
                    mode = "RGB" if pix.n < 4 else "RGBA"
                    img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                    dc.StartPage(); _draw_pil_page(dc, img, scale_mode, render_dpi); dc.EndPage()
            dc.EndDoc(); return True, "", spool_id
        finally:
            if dc is not None:
                try: dc.DeleteDC()
                except Exception: pass
            doc.close()
    except Exception as exc:
        return False, "Gagal mencetak PDF langsung: %s" % exc, 0


def _print_image_windows(path: Path, printer: str, copies: int, paper: str = "AUTO",
                         orientation: str = "AUTO", scale_mode: str = "FIT", duplex: str = "DEFAULT",
                         color_mode: str = "DEFAULT", spool_callback=None) -> "tuple[bool, str, int]":
    dc = None
    try:
        from PIL import Image
        with Image.open(str(path)) as source:
            dc, printer = _create_printer_dc(printer, paper, orientation, duplex, color_mode)
            result = dc.StartDoc("PrintBot Web - %s" % path.name)
            try: spool_id = int(result or 0)
            except Exception: spool_id = 0
            if spool_id and spool_callback: spool_callback(spool_id)
            dpi_meta = source.info.get("dpi", (96, 96))
            try: source_dpi = int(dpi_meta[0]) if dpi_meta else 96
            except Exception: source_dpi = 96
            for _copy in range(max(1, int(copies))):
                dc.StartPage(); _draw_pil_page(dc, source.copy(), scale_mode, source_dpi); dc.EndPage()
            dc.EndDoc(); return True, "", spool_id
    except Exception as exc:
        return False, "Gagal mencetak gambar langsung: %s" % exc, 0
    finally:
        if dc is not None:
            try: dc.DeleteDC()
            except Exception: pass


def _office_detect_new_spool(printer: str, before_ids: set, timeout: float = 6.0) -> int:
    deadline = time.time() + timeout
    while time.time() < deadline:
        new_ids = sorted(set(_spooler_jobs(printer).keys()) - set(before_ids))
        if new_ids: return int(new_ids[-1])
        time.sleep(0.25)
    return 0


def _apply_office_page_setup(page_setup, paper: str, orientation: str, kind: str = "excel") -> None:
    try:
        if kind == "word":
            if orientation == "PORTRAIT": page_setup.Orientation = 0
            elif orientation == "LANDSCAPE": page_setup.Orientation = 1
            sizes = {"LETTER": (612, 792), "LEGAL": (612, 1008), "A4": (595, 842), "F4": (612, 936)}
            if paper in sizes:
                w, h = sizes[paper]
                if orientation == "LANDSCAPE": w, h = h, w
                page_setup.PageWidth = w; page_setup.PageHeight = h
        else:
            if orientation == "PORTRAIT": page_setup.Orientation = 1
            elif orientation == "LANDSCAPE": page_setup.Orientation = 2
            paper_map = {"LETTER": 1, "LEGAL": 5, "A4": 9, "F4": 14}
            if paper in paper_map: page_setup.PaperSize = paper_map[paper]
    except Exception:
        pass


def _print_office_windows(path: Path, printer: str, copies: int, pages: str, sheet: str,
                          paper: str = "AUTO", orientation: str = "AUTO", scale_mode: str = "FIT",
                          duplex: str = "DEFAULT", color_mode: str = "DEFAULT", spool_callback=None) -> "tuple[bool, str, int]":
    ext = path.suffix.lower()
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        return False, "Microsoft Office COM/pywin32 tidak tersedia: %s" % exc, 0
    pythoncom.CoInitialize()
    try:
        printer = printer or local_default_printer()
        before = set(_spooler_jobs(printer).keys()) if printer else set()
        if ext in (".xls", ".xlsx", ".xlsm"):
            app = wb = None
            try:
                app = win32com.client.DispatchEx("Excel.Application"); app.Visible = False; app.DisplayAlerts = False
                wb = app.Workbooks.Open(str(path), ReadOnly=True, UpdateLinks=0)
                target = wb.Worksheets(sheet) if sheet else wb
                if sheet: _apply_office_page_setup(target.PageSetup, paper, orientation, "excel")
                else:
                    for ws in wb.Worksheets: _apply_office_page_setup(ws.PageSetup, paper, orientation, "excel")
                if scale_mode == "FIT":
                    try:
                        ps = target.PageSetup if sheet else wb.ActiveSheet.PageSetup
                        ps.Zoom = False; ps.FitToPagesWide = 1
                    except Exception: pass
                kwargs = {"Copies": max(1, int(copies))}
                if printer: kwargs["ActivePrinter"] = printer
                nums = _parse_page_numbers(pages)
                if nums:
                    if nums != list(range(min(nums), max(nums) + 1)):
                        return False, "Untuk Excel, halaman harus berupa rentang berurutan (contoh 2-5).", 0
                    kwargs["From"] = min(nums); kwargs["To"] = max(nums)
                target.PrintOut(**kwargs)
            except Exception as exc:
                return False, "Gagal mencetak Excel melalui Microsoft Office: %s" % exc, 0
            finally:
                try:
                    if wb is not None: wb.Close(False)
                except Exception: pass
                try:
                    if app is not None: app.Quit()
                except Exception: pass
        elif ext in (".doc", ".docx", ".rtf"):
            app = doc = None
            try:
                app = win32com.client.DispatchEx("Word.Application"); app.Visible = False; app.DisplayAlerts = 0
                doc = app.Documents.Open(str(path), ReadOnly=True, AddToRecentFiles=False)
                if printer: app.ActivePrinter = printer
                _apply_office_page_setup(doc.PageSetup, paper, orientation, "word")
                kwargs = {"Background": False, "Copies": max(1, int(copies))}
                if pages: kwargs.update({"Range": 4, "Pages": pages})
                doc.PrintOut(**kwargs)
            except Exception as exc:
                return False, "Gagal mencetak Word melalui Microsoft Office: %s" % exc, 0
            finally:
                try:
                    if doc is not None: doc.Close(False)
                except Exception: pass
                try:
                    if app is not None: app.Quit()
                except Exception: pass
        elif ext in (".ppt", ".pptx"):
            app = pres = None
            try:
                app = win32com.client.DispatchEx("PowerPoint.Application")
                pres = app.Presentations.Open(str(path), WithWindow=False, ReadOnly=True)
                if printer:
                    try: pres.PrintOptions.ActivePrinter = printer
                    except Exception: pass
                nums = _parse_page_numbers(pages)
                if nums and nums != list(range(min(nums), max(nums) + 1)):
                    return False, "Untuk PowerPoint, halaman harus berupa rentang berurutan (contoh 2-5).", 0
                if nums: pres.PrintOut(From=min(nums), To=max(nums), Copies=max(1, int(copies)))
                else: pres.PrintOut(Copies=max(1, int(copies)))
            except Exception as exc:
                return False, "Gagal mencetak PowerPoint melalui Microsoft Office: %s" % exc, 0
            finally:
                try:
                    if pres is not None: pres.Close()
                except Exception: pass
                try:
                    if app is not None: app.Quit()
                except Exception: pass
        else:
            return False, "Format Office belum memiliki handler COM khusus.", 0
        spool_id = _office_detect_new_spool(printer, before) if printer else 0
        if spool_id and spool_callback: spool_callback(spool_id)
        return True, "", spool_id
    finally:
        try: pythoncom.CoUninitialize()
        except Exception: pass


def _windows_shell_print(path: Path, printer: str) -> "tuple[bool, str]":
    try:
        win32api, win32con, _wp, _wu = _win_import_print_modules()
        if win32api is None: return False, "pywin32 belum tersedia."
        if printer:
            rc = win32api.ShellExecute(0, "printto", str(path), '"%s"' % printer.replace('"', ''),
                                       str(path.parent), win32con.SW_HIDE)
        else:
            rc = win32api.ShellExecute(0, "print", str(path), None, str(path.parent), win32con.SW_HIDE)
        if int(rc) <= 32: return False, "Windows tidak menemukan aplikasi/verb Print untuk file ini (kode %s)." % rc
        return True, ""
    except Exception as exc:
        return False, "Fallback Print Windows gagal: %s" % exc


def local_print_file(path: Path, printer: str, copies: int, pages: str = "", sheet: str = "",
                     paper: str = "AUTO", orientation: str = "AUTO", scale_mode: str = "FIT",
                     duplex: str = "DEFAULT", color_mode: str = "DEFAULT",
                     spool_callback=None, spool_status_callback=None) -> "tuple[bool, str, int]":
    try:
        if sys.platform.startswith("win"):
            printers = local_list_printers(); printer = (printer or "").strip() or local_default_printer()
            if not printer: return False, "Tidak ada default printer. Pilih printer terlebih dahulu.", 0
            if printers and printer not in printers:
                return False, "Printer '%s' tidak ditemukan." % printer, 0
            pstat = printer_status_details(printer)
            if pstat.get("problem"):
                return False, "Printer belum siap: %s" % pstat.get("message", "Unknown"), 0
            ext = path.suffix.lower()
            if ext == ".pdf":
                result = _print_pdf_windows(path, printer, copies, pages, paper, orientation, scale_mode, duplex, color_mode, spool_callback)
            elif ext in IMAGE_EXTS:
                result = _print_image_windows(path, printer, copies, paper, orientation, scale_mode, duplex, color_mode, spool_callback)
            elif ext in (".xls", ".xlsx", ".xlsm", ".doc", ".docx", ".rtf", ".ppt", ".pptx"):
                result = _print_office_windows(path, printer, copies, pages, sheet, paper, orientation, scale_mode, duplex, color_mode, spool_callback)
                if not result[0]:
                    log.warning("Office COM print gagal: %s", result[1])
                    return result
            else:
                ok = True; msg = ""
                for _ in range(max(1, int(copies))):
                    ok, msg = _windows_shell_print(path, printer)
                    if not ok: break
                    time.sleep(0.4)
                return ok, msg, 0
            ok, msg, spool_id = result
            if not ok: return result
            if spool_id:
                def cb(st):
                    if spool_status_callback: spool_status_callback(st)
                spool_ok, spool_msg = _wait_spool_job(printer, spool_id,
                    timeout=min(max(settings.print_job_timeout_sec, 60), 3600), status_callback=cb)
                if not spool_ok: return False, spool_msg, spool_id
            return True, "", spool_id

        # Linux/macOS best effort via CUPS.
        cmd = ["lp"]
        if printer: cmd += ["-d", printer]
        cmd += ["-n", str(max(1, int(copies))), str(path)]
        r = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, timeout=60)
        return r.returncode == 0, "" if r.returncode == 0 else (r.stderr.strip() or r.stdout.strip()), 0
    except Exception as exc:
        return False, str(exc), 0


def job_set_status(job_id: str, status: str, error: str = "", mark_dispatched: bool = False) -> Optional[dict]:
    now = time.time()
    with _db_lock, _db_connect() as conn:
        if mark_dispatched:
            conn.execute("UPDATE jobs SET status=?, dispatched_at=?, started_at=? WHERE id=?", (status, now, now, job_id))
        elif status in (JobStatus.SUCCESS, JobStatus.FAILED):
            conn.execute("UPDATE jobs SET status=?, error=?, finished_at=? WHERE id=?", (status, error or None, now, job_id))
        else:
            conn.execute("UPDATE jobs SET status=? WHERE id=?", (status, job_id))
        conn.commit()
    return job_get(job_id)


def job_update_spool(job_id: str, spool_job_id=None, spool_status: str = "") -> None:
    with _db_lock, _db_connect() as conn:
        conn.execute("UPDATE jobs SET spool_job_id=COALESCE(?, spool_job_id), spool_status=? WHERE id=?",
                     (spool_job_id, spool_status or None, job_id)); conn.commit()


def job_claim_next_direct() -> Optional[dict]:
    """Ambil tepat satu job secara atomik. Satu worker = urutan cetak stabil."""
    now = time.time()
    with _db_lock, _db_connect() as conn:
        row = conn.execute("SELECT id FROM jobs WHERE status=? AND method=? AND (scheduled_at IS NULL OR scheduled_at<=?) "
                           "ORDER BY priority DESC, created_at ASC LIMIT 1",
                           (JobStatus.QUEUED, PrintMethod.DIRECT, now)).fetchone()
        if not row: return None
        cur = conn.execute("UPDATE jobs SET status=?, dispatched_at=?, started_at=?, error=NULL WHERE id=? AND status=?",
                           (JobStatus.PROCESSING, now, now, row["id"], JobStatus.QUEUED))
        conn.commit()
        if cur.rowcount != 1: return None
    return job_get(row["id"])


def job_recover_direct_after_restart() -> int:
    """Job yang belum masuk spooler boleh diulang; job yang sudah punya Windows Job ID
    tidak diulang otomatis untuk mencegah cetak ganda setelah crash/restart."""
    n = 0
    with _db_lock, _db_connect() as conn:
        rows = conn.execute("SELECT id, retries, spool_job_id FROM jobs WHERE method=? AND status IN (?,?)",
                            (PrintMethod.DIRECT, JobStatus.PROCESSING, JobStatus.PRINTING)).fetchall()
        for r in rows:
            if r["spool_job_id"]:
                conn.execute("UPDATE jobs SET status=?, error=?, finished_at=?, spool_status=? WHERE id=?",
                             (JobStatus.FAILED, "Aplikasi restart setelah job masuk Windows spooler; tidak diulang otomatis untuk mencegah cetak ganda.",
                              time.time(), "Unknown after restart", r["id"]))
            elif int(r["retries"] or 0) < settings.print_max_retries:
                conn.execute("UPDATE jobs SET status=?, dispatched_at=NULL, started_at=NULL, retries=retries+1, error=? WHERE id=?",
                             (JobStatus.QUEUED, "Dipulihkan setelah restart.", r["id"]))
            else:
                conn.execute("UPDATE jobs SET status=?, error=?, finished_at=? WHERE id=?",
                             (JobStatus.FAILED, "Job gagal dipulihkan setelah restart.", time.time(), r["id"]))
            n += 1
        conn.commit()
    return n


def _job_retry_or_fail(job_id: str, error: str, spool_job_id: int = 0) -> None:
    with _db_lock, _db_connect() as conn:
        row = conn.execute("SELECT retries FROM jobs WHERE id=?", (job_id,)).fetchone()
        retries = int(row["retries"] or 0) if row else settings.print_max_retries
        if not spool_job_id and retries < settings.print_max_retries:
            conn.execute("UPDATE jobs SET status=?, error=?, retries=retries+1, dispatched_at=NULL, started_at=NULL, spool_status=? WHERE id=?",
                         (JobStatus.QUEUED, error, "Retry queued", job_id))
        else:
            conn.execute("UPDATE jobs SET status=?, error=?, finished_at=? WHERE id=?",
                         (JobStatus.FAILED, error, time.time(), job_id))
        conn.commit()


_PRINT_QUEUE_PAUSED = app_setting_get("queue_paused","0") == "1"
_PRINT_WORKER_ACTIVE_JOB: Optional[str] = None
_ACTIVE_JOB_BY_PRINTER: Dict[str, str] = {}
_ACTIVE_PRINTER_TASKS: Dict[str, asyncio.Task] = {}
_ACTIVE_CANCEL_EVENTS: Dict[str, threading.Event] = {}


async def _execute_direct_job(job: dict) -> None:
    global _PRINT_WORKER_ACTIVE_JOB
    job_id = job["id"]
    printer_key = (job.get("printer") or local_default_printer() or "(default)").strip()
    _ACTIVE_JOB_BY_PRINTER[printer_key] = job_id
    _PRINT_WORKER_ACTIVE_JOB = ",".join(_ACTIVE_JOB_BY_PRINTER.values()) or None
    cancel_event = threading.Event(); _ACTIVE_CANCEL_EVENTS[job_id] = cancel_event
    try:
        source = resolve_print_archive(job.get("archive_path") or "")
        if source is None:
            p = Path(job.get("file_path") or "")
            source = p if p.is_file() else None
        if source is None:
            job_set_status(job_id, JobStatus.FAILED, "File sumber tidak ditemukan."); return
        if cancel_event.is_set():
            job_set_status(job_id, JobStatus.FAILED, "Dibatalkan pengguna sebelum diproses."); return
        printer = (job.get("printer") or "").strip() or local_default_printer()
        job_set_status(job_id, JobStatus.PRINTING)
        await event_publish({"type":"job","action":"printing","job_id":job_id,"printer":printer})

        def spool_id_cb(spool_id):
            job_update_spool(job_id, int(spool_id), "Submitted to Windows spooler")
            if cancel_event.is_set():
                ok, msg = cancel_windows_spool_job(printer, int(spool_id))
                with _db_lock, _db_connect() as conn:
                    conn.execute("UPDATE jobs SET spool_cancelled=1, spool_status=? WHERE id=?",
                                 ("Cancelled" if ok else msg, job_id)); conn.commit()
        def spool_status_cb(status_text):
            job_update_spool(job_id, None, str(status_text))

        ok, err, spool_id = await _to_thread(
            local_print_file, source, printer, int(job.get("copies") or 1), job.get("pages") or "", job.get("sheet") or "",
            job.get("paper") or "AUTO", job.get("orientation") or "AUTO", job.get("scale_mode") or "FIT",
            job.get("duplex") or "DEFAULT", job.get("color_mode") or "DEFAULT", spool_id_cb, spool_status_cb)
        current = job_get(job_id) or {}
        if cancel_event.is_set() or int(current.get("spool_cancelled") or 0) or (current.get("status") == JobStatus.FAILED and str(current.get("error") or "").lower().startswith("dibatalkan")):
            if current.get("status") != JobStatus.FAILED:
                job_set_status(job_id, JobStatus.FAILED, "Dibatalkan pengguna.")
            return
        if ok:
            job_update_spool(job_id, spool_id or None, "Completed"); job_set_status(job_id, JobStatus.SUCCESS)
            log.info("PRINT SUCCESS job=%s file=%s printer=%s spool=%s", job_id, source.name, printer, spool_id or "-")
            await event_publish({"type":"job","action":"success","job_id":job_id,"printer":printer,"spool_job_id":spool_id or 0})
        else:
            log.error("PRINT FAILED job=%s file=%s error=%s", job_id, source, err)
            _job_retry_or_fail(job_id, err or "Print lokal gagal.", spool_id)
            record_error("PRINT",err or "Print lokal gagal.","job=%s printer=%s"%(job_id,printer))
            await event_publish({"type":"job","action":"failed","job_id":job_id,"printer":printer,"error":err or "Print lokal gagal."})
    finally:
        _ACTIVE_CANCEL_EVENTS.pop(job_id, None)
        _ACTIVE_JOB_BY_PRINTER.pop(printer_key, None)
        _PRINT_WORKER_ACTIVE_JOB = ",".join(_ACTIVE_JOB_BY_PRINTER.values()) or None


async def direct_print_worker_loop() -> None:
    """Dispatcher multi-printer: satu job aktif per printer, printer berbeda paralel."""
    log.info("Local multi-printer dispatcher aktif.")
    while True:
        try:
            # bersihkan task selesai
            for pname, task in list(_ACTIVE_PRINTER_TASKS.items()):
                if task.done():
                    try: task.result()
                    except asyncio.CancelledError: raise
                    except Exception as exc:
                        record_error("PRINT_WORKER", str(exc), "printer=%s" % pname)
                        log.exception("Worker printer %s gagal", pname)
                    _ACTIVE_PRINTER_TASKS.pop(pname, None)
            if _PRINT_QUEUE_PAUSED:
                await asyncio.sleep(0.5); continue
            # maintenance mode menolak job baru di API, tapi job lama yang sudah antre tetap ditahan.
            if is_maintenance_mode():
                await asyncio.sleep(0.8); continue
            active = set(_ACTIVE_PRINTER_TASKS.keys())
            now = time.time()
            with _db_lock, _db_connect() as conn:
                rows = conn.execute(
                    "SELECT id,printer FROM jobs WHERE status=? AND method=? AND (scheduled_at IS NULL OR scheduled_at<=?) "
                    "ORDER BY priority DESC, created_at ASC LIMIT 100",
                    (JobStatus.QUEUED, PrintMethod.DIRECT, now)).fetchall()
                claimed = None
                for r in rows:
                    pname = str(r["printer"] or local_default_printer() or "").strip()
                    if not pname or pname in active: continue
                    cur = conn.execute("UPDATE jobs SET status=?, dispatched_at=?, started_at=?, error=NULL WHERE id=? AND status=?",
                                       (JobStatus.PROCESSING, now, now, r["id"], JobStatus.QUEUED))
                    if cur.rowcount == 1:
                        conn.commit(); claimed = (r["id"], pname); break
                if claimed is None: conn.commit()
            if not claimed:
                await asyncio.sleep(0.45); continue
            job = await _to_thread(job_get, claimed[0])
            if job:
                task = asyncio.create_task(_execute_direct_job(job), name="print-%s" % claimed[1])
                _ACTIVE_PRINTER_TASKS[claimed[1]] = task
            await asyncio.sleep(0.05)
        except asyncio.CancelledError:
            for task in list(_ACTIVE_PRINTER_TASKS.values()): task.cancel()
            raise
        except Exception as exc:
            record_error("PRINT_DISPATCHER", str(exc)); log.exception("Local multi-printer dispatcher error")
            await asyncio.sleep(1.0)


def create_test_print(printer: str) -> "tuple[bool, str, int]":
    try:
        from PIL import Image, ImageDraw, ImageFont
        out = settings.data_dir / "previews" / ("test_print_%s.png" % uuid.uuid4().hex[:8])
        img = Image.new("RGB", (1240, 1754), "white"); draw = ImageDraw.Draw(img)
        try: font = ImageFont.truetype("arial.ttf", 34); small = ImageFont.truetype("arial.ttf", 24)
        except Exception: font = small = ImageFont.load_default()
        y = 100
        for text, f in [("PrintBot WebApp - TEST PRINTER", font), ("Printer: %s" % (printer or local_default_printer()), small),
                        ("Waktu: %s" % time.strftime("%Y-%m-%d %H:%M:%S"), small),
                        ("Python: %s" % sys.version.split()[0], small), ("Mode: Printer Lokal / Windows Spooler", small)]:
            draw.text((100, y), text, fill="black", font=f); y += 70
        draw.rectangle((100, y + 50, 1140, y + 300), outline="black", width=4)
        draw.text((130, y + 130), "Jika kotak ini tercetak utuh, test berhasil.", fill="black", font=small)
        img.save(str(out), "PNG")
        try:
            return local_print_file(out, printer, 1, paper="A4", orientation="PORTRAIT", scale_mode="FIT")
        finally:
            try: out.unlink()
            except OSError: pass
    except Exception as exc:
        return False, str(exc), 0


# ============================================================================
# DIAGNOSTIC / BACKUP / STORAGE
# ============================================================================
def office_availability() -> dict:
    out = {"Word": False, "Excel": False, "PowerPoint": False}
    if not sys.platform.startswith("win"):
        return out
    try:
        import winreg
        for label, key in (("Word", r"Word.Application\CLSID"), ("Excel", r"Excel.Application\CLSID"),
                           ("PowerPoint", r"PowerPoint.Application\CLSID")):
            try:
                h = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, key); winreg.CloseKey(h); out[label] = True
            except OSError:
                pass
    except Exception:
        pass
    return out


def db_quick_check() -> dict:
    try:
        with _db_lock, _db_connect() as conn:
            row = conn.execute("PRAGMA quick_check").fetchone()
            result = str(row[0] if row else "unknown")
            conn.execute("PRAGMA wal_checkpoint(PASSIVE)")
        return {"ok": result.lower() == "ok", "result": result}
    except Exception as exc:
        return {"ok": False, "result": str(exc)}


def _dir_size(path: Path) -> int:
    total = 0
    if not path.exists(): return 0
    for f in path.rglob("*"):
        try:
            if f.is_file(): total += f.stat().st_size
        except OSError: pass
    return total


def storage_status() -> dict:
    try:
        usage = shutil.disk_usage(str(settings.data_dir))
        disk = {"total": usage.total, "used": usage.used, "free": usage.free}
    except Exception:
        disk = {"total": 0, "used": 0, "free": 0}
    return {
        "disk": disk,
        "database": settings.db_path.stat().st_size if settings.db_path.is_file() else 0,
        "archive": _dir_size(settings.print_archive_dir),
        "previews": _dir_size(settings.data_dir / "previews"),
        "logs": _dir_size(settings.data_dir / "logs"),
        "backup": _dir_size(settings.backup_dir),
    }


def storage_cleanup(category: str, max_age_days: int = 7) -> dict:
    category=(category or "").lower(); removed=0; bytes_removed=0
    targets={"previews":settings.data_dir/"previews","uploads":settings.data_dir/"uploads","exports":settings.data_dir/"exports","logs":settings.data_dir/"logs","archive":settings.print_archive_dir}
    if category not in targets: raise ValueError("Kategori storage tidak valid.")
    root=targets[category]; cutoff=time.time()-max(0,max_age_days)*86400
    for f in root.rglob("*") if root.exists() else []:
        try:
            if not f.is_file(): continue
            # log aktif jangan dihapus; hanya rotated log lama.
            if category=="logs" and f==settings.log_path: continue
            if max_age_days>0 and f.stat().st_mtime>=cutoff: continue
            size=f.stat().st_size; f.unlink(); removed+=1; bytes_removed+=size
        except OSError: pass
    return {"category":category,"removed":removed,"bytes_removed":bytes_removed}


def cleanup_old_backups() -> int:
    cutoff = time.time() - max(1, settings.backup_retention_days) * 86400
    n = 0
    for f in settings.backup_dir.glob("PrintBotWeb_Backup_*.zip"):
        try:
            if f.stat().st_mtime < cutoff: f.unlink(); n += 1
        except OSError: pass
    return n


def create_backup_archive() -> Path:
    settings.backup_dir.mkdir(parents=True, exist_ok=True)
    with _db_lock, _db_connect() as conn:
        try: conn.execute("PRAGMA wal_checkpoint(FULL)")
        except Exception: pass
    stamp = time.strftime("%Y%m%d_%H%M%S")
    out = settings.backup_dir / ("PrintBotWeb_Backup_%s.zip" % stamp)
    candidates = [settings.db_path, settings.data_dir / "roots.json", settings.data_dir / "users.json",
                  settings.data_dir / "runtime_settings.json", settings.data_dir / "webapp_secret.key", settings.data_dir / "api_token.key"]
    with zipfile.ZipFile(str(out), "w", compression=zipfile.ZIP_DEFLATED) as z:
        for f in candidates:
            if f.is_file(): z.write(str(f), arcname=f.name)
    cleanup_old_backups()
    return out


def backup_list() -> List[dict]:
    out=[]
    for f in sorted(settings.backup_dir.glob("PrintBotWeb_Backup_*.zip"),key=lambda x:x.stat().st_mtime,reverse=True):
        try: out.append({"name":f.name,"size":f.stat().st_size,"modified":f.stat().st_mtime})
        except OSError: pass
    return out


def _safe_backup_path(name: str) -> Path:
    if not name or Path(name).name!=name or not name.startswith("PrintBotWeb_Backup_") or not name.endswith(".zip"):
        raise ValueError("Nama backup tidak valid.")
    p=(settings.backup_dir/name).resolve(); p.relative_to(settings.backup_dir.resolve())
    if not p.is_file(): raise FileNotFoundError("Backup tidak ditemukan.")
    return p


def restore_backup_archive(name: str) -> dict:
    src=_safe_backup_path(name)
    allowed={"printbot.db","roots.json","users.json","runtime_settings.json","webapp_secret.key","api_token.key"}
    safety=create_backup_archive()
    temp=Path(tempfile.mkdtemp(prefix="printbot_restore_"))
    replaced=[]
    try:
        with zipfile.ZipFile(str(src),"r") as z:
            names=set(z.namelist())
            if not names or any(Path(n).name!=n or n not in allowed for n in names):
                raise ValueError("Isi backup tidak valid/tidak didukung.")
            z.extractall(str(temp))
        db=temp/"printbot.db"
        if db.is_file():
            c=sqlite3.connect(str(db)); row=c.execute("PRAGMA integrity_check").fetchone(); c.close()
            if not row or str(row[0]).lower()!="ok": raise ValueError("Database backup gagal integrity_check.")
        mapping={"printbot.db":settings.db_path,"roots.json":settings.data_dir/"roots.json","users.json":settings.data_dir/"users.json",
                 "runtime_settings.json":settings.data_dir/"runtime_settings.json","webapp_secret.key":settings.data_dir/"webapp_secret.key","api_token.key":settings.data_dir/"api_token.key"}
        with _db_lock:
            for name2,dst in mapping.items():
                f=temp/name2
                if f.is_file():
                    tmp=dst.with_suffix(dst.suffix+".restore"); shutil.copy2(str(f),str(tmp)); os.replace(str(tmp),str(dst)); replaced.append(name2)
        return {"success":True,"restored":replaced,"safety_backup":safety.name,"restart_required":True}
    except Exception:
        # safety backup tetap tersedia untuk pemulihan manual/rollback.
        raise
    finally:
        shutil.rmtree(str(temp),ignore_errors=True)


def db_full_maintenance(vacuum: bool = True) -> dict:
    with _db_lock, _db_connect() as conn:
        integrity=str(conn.execute("PRAGMA integrity_check").fetchone()[0]); conn.execute("PRAGMA wal_checkpoint(TRUNCATE)"); conn.execute("ANALYZE")
        if vacuum: conn.execute("VACUUM")
    app_setting_set("db_last_maintenance",str(time.time()))
    return {"ok":integrity.lower()=="ok","integrity":integrity,"vacuum":vacuum}


def cleanup_preview_files(max_age_sec: int = 86400) -> int:
    n = 0; cutoff = time.time() - max_age_sec
    for f in (settings.data_dir / "previews").glob("*"):
        try:
            if f.is_file() and f.stat().st_mtime < cutoff: f.unlink(); n += 1
        except OSError: pass
    return n


# ============================================================================
# REPORT EXPORT / AUTO UPDATE / ROLLBACK
# ============================================================================
def export_jobs_report(fmt: str = "xlsx", limit: int = 10000) -> Path:
    rows=job_list(limit=max(1,min(limit,50000)),offset=0); outdir=settings.data_dir/"exports"; outdir.mkdir(exist_ok=True)
    stamp=time.strftime("%Y%m%d_%H%M%S"); fields=["id","created_at","finished_at","sender","file_name","printer","copies","pages","paper","orientation","duplex","color_mode","priority","scheduled_at","status","spool_job_id","spool_status","error"]
    if fmt.lower()=="csv":
        out=outdir/("PrintBot_History_%s.csv"%stamp)
        with out.open("w",newline="",encoding="utf-8-sig") as fh:
            w=csv.DictWriter(fh,fieldnames=fields); w.writeheader()
            for r in rows: w.writerow({k:r.get(k) for k in fields})
        return out
    import openpyxl
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Print History"; ws.append(fields)
    for r in rows: ws.append([r.get(k) for k in fields])
    for c in ws[1]: c.font=openpyxl.styles.Font(bold=True)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    out=outdir/("PrintBot_History_%s.xlsx"%stamp); wb.save(str(out)); wb.close(); return out


def _version_tuple(value: str) -> tuple:
    nums=[]
    for x in re.findall(r"\d+",str(value or ""))[:4]:
        try: nums.append(int(x))
        except ValueError: nums.append(0)
    return tuple(nums or [0])


def _validate_update_code(code: bytes) -> dict:
    if len(code)<20000: raise ValueError("File update terlalu kecil/tidak valid.")
    txt=code.decode("utf-8")
    try: tree=ast.parse(txt,filename="PrintBotWeb_update.py",feature_version=(3,8))
    except (TypeError,ValueError): tree=ast.parse(txt,filename="PrintBotWeb_update.py",feature_version=8)
    compile(tree,"PrintBotWeb_update.py","exec")
    m=re.search(r'^APP_VERSION\s*=\s*["\']([^"\']+)',txt,re.M)
    if not m or "PrintBot" not in txt or "def main(" not in txt: raise ValueError("Sumber update bukan PrintBot WebApp valid.")
    return {"version":m.group(1),"sha256":hashlib.sha256(code).hexdigest()}


def check_update_available() -> dict:
    url=(settings.update_url or "").strip()
    if not url.lower().startswith(("https://","http://")): return {"configured":False,"available":False,"current":APP_VERSION}
    req=urllib.request.Request(url,headers={"User-Agent":"PrintBotWeb/%s"%APP_VERSION,**({"Authorization":"Bearer %s"%settings.update_token} if settings.update_token else {})})
    with urllib.request.urlopen(req,timeout=30) as resp: code=resp.read(8*1024*1024)
    meta=_validate_update_code(code); meta.update({"configured":True,"current":APP_VERSION,"available":_version_tuple(meta["version"])>_version_tuple(APP_VERSION),"bytes":len(code)})
    return meta


def apply_update_from_url() -> dict:
    url=(settings.update_url or "").strip()
    if not url: raise ValueError("UPDATE_URL belum dikonfigurasi.")
    req=urllib.request.Request(url,headers={"User-Agent":"PrintBotWeb/%s"%APP_VERSION,**({"Authorization":"Bearer %s"%settings.update_token} if settings.update_token else {})})
    with urllib.request.urlopen(req,timeout=45) as resp: code=resp.read(8*1024*1024)
    meta=_validate_update_code(code)
    if _version_tuple(meta["version"])<=_version_tuple(APP_VERSION): return {"updated":False,"reason":"Versi remote tidak lebih baru",**meta}
    upd=settings.data_dir/"updates"; upd.mkdir(exist_ok=True); current=Path(__file__).resolve(); stamp=time.strftime("%Y%m%d_%H%M%S")
    backup=upd/("webapp_before_%s_%s.py"%(APP_VERSION.replace("/","_"),stamp)); shutil.copy2(str(current),str(backup))
    temp=current.with_suffix(current.suffix+".update"); temp.write_bytes(code); os.replace(str(temp),str(current))
    state={"old_version":APP_VERSION,"new_version":meta["version"],"backup":str(backup),"sha256":meta["sha256"],"updated_at":time.time()}
    (upd/"update_state.json").write_text(json.dumps(state,indent=2),encoding="utf-8")
    return {"updated":True,"restart_required":True,**state}


def rollback_last_update() -> dict:
    state_path=settings.data_dir/"updates"/"update_state.json"
    if not state_path.is_file(): raise FileNotFoundError("State update tidak ditemukan.")
    state=json.loads(state_path.read_text("utf-8")); backup=Path(state.get("backup") or "")
    if not backup.is_file(): raise FileNotFoundError("File rollback tidak tersedia.")
    code=backup.read_bytes(); _validate_update_code(code); current=Path(__file__).resolve(); temp=current.with_suffix(current.suffix+".rollback"); temp.write_bytes(code); os.replace(str(temp),str(current))
    return {"rolled_back":True,"to_version":state.get("old_version"),"restart_required":True}


async def auto_update_loop() -> None:
    while True:
        try:
            if settings.update_url:
                result=await _to_thread(check_update_available)
                if result.get("available"):
                    audit_event("system","UPDATE_AVAILABLE",detail=json.dumps(result,ensure_ascii=False))
                    await event_publish({"type":"update","action":"available",**result})
                    if settings.auto_update_apply:
                        applied=await _to_thread(apply_update_from_url); await event_publish({"type":"update","action":"applied",**applied})
                        if sys.platform.startswith("win"):
                            await _to_thread(schedule_windows_task_restart,3); await asyncio.sleep(0.8); os._exit(0)
                        await asyncio.sleep(1.0); os._exit(75)
        except asyncio.CancelledError: raise
        except Exception as exc: record_error("AUTO_UPDATE",str(exc))
        await asyncio.sleep(max(1,settings.update_check_hours)*3600)


async def db_maintenance_loop() -> None:
    while True:
        try:
            now=time.time(); last_vac=float(app_setting_get("db_last_vacuum","0") or 0); do_vac=(now-last_vac)>=7*86400
            await _to_thread(db_full_maintenance,do_vac)
            if do_vac: app_setting_set("db_last_vacuum",str(now))
        except asyncio.CancelledError: raise
        except Exception as exc: record_error("DB_MAINTENANCE",str(exc))
        await asyncio.sleep(max(1,settings.db_maintenance_hours)*3600)


# ============================================================================
# USER ACCOUNTS & ROLES - akun utama (WEBAPP_USERNAME/WEBAPP_PASSWORD di .env)
# selalu berperan "admin". Akun tambahan (role "admin" atau "user") dikelola
# admin lewat WebApp, tersimpan di data/users.json dengan password di-hash
# (PBKDF2-HMAC-SHA256 + salt per akun, stdlib saja - tidak perlu dependency
# tambahan). Role "user" TIDAK bisa rename/hapus/tambah folder/upload -
# hanya browse, cari, preview, download, dan print.
# ============================================================================
ROLE_ADMIN = "admin"
ROLE_USER = "user"


def _users_path() -> Path:
    return settings.data_dir / "users.json"


def _hash_password(password: str, salt: str, iterations: int = 310_000) -> str:
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), int(iterations)).hex()


def users_load() -> dict:
    p = _users_path()
    if p.exists():
        try:
            return json.loads(p.read_text("utf-8"))
        except Exception:
            return {}
    return {}


def users_save(data: dict) -> None:
    _users_path().write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def user_add(username: str, password: str, role: str = ROLE_USER) -> dict:
    username = (username or "").strip()
    if not username or any(c in username for c in " /\\:"):
        raise FileManagerError("Username tidak valid (tidak boleh kosong atau mengandung spasi/'/'.")
    if not password or len(password) < 8:
        raise FileManagerError("Password minimal 8 karakter.")
    role = role if role in (ROLE_ADMIN, ROLE_USER) else ROLE_USER
    if constant_time_eq(username, settings.webapp_username):
        raise FileManagerError("Username ini sudah dipakai akun admin utama (dari .env).")
    data = users_load()
    salt = secrets.token_hex(16)
    data[username] = {"password_hash": _hash_password(password, salt, 310_000), "salt": salt, "iterations": 310000, "role": role}
    users_save(data)
    log.info("USER ADD username=%s role=%s", username, role)
    return {"username": username, "role": role}


def user_remove(username: str) -> bool:
    data = users_load()
    if username not in data:
        return False
    del data[username]
    users_save(data)
    log.info("USER REMOVE username=%s", username)
    return True


def user_list() -> list:
    data = users_load()
    out = [{"username": settings.webapp_username, "role": ROLE_ADMIN, "primary": True}]
    for uname, v in data.items():
        out.append({"username": uname, "role": v.get("role", ROLE_USER), "primary": False})
    return out


def verify_login(username: str, password: str) -> Optional[str]:
    """Return role ('admin'/'user') kalau kredensial valid, else None."""
    if settings.webapp_password and constant_time_eq(username, settings.webapp_username) \
            and constant_time_eq(password, settings.webapp_password):
        return ROLE_ADMIN
    entry = users_load().get(username)
    if not entry:
        return None
    salt = entry.get("salt", "")
    expected = entry.get("password_hash", "")
    if not salt or not expected:
        return None
    iterations = int(entry.get("iterations") or 100000)
    actual = _hash_password(password, salt, iterations)
    if constant_time_eq(actual, expected):
        return entry.get("role", ROLE_USER)
    return None


# ============================================================================
# REALTIME EVENT BUS (SSE)
# ============================================================================
_EVENT_SUBSCRIBERS: Set[asyncio.Queue] = set()

async def event_publish(payload: dict) -> None:
    data=dict(payload or {}); data.setdefault("ts",time.time())
    for q in list(_EVENT_SUBSCRIBERS):
        try:
            if q.qsize()>100:
                q.get_nowait()
            q.put_nowait(data)
        except Exception:
            _EVENT_SUBSCRIBERS.discard(q)

async def _event_stream() -> AsyncIterator[str]:
    q: asyncio.Queue=asyncio.Queue(maxsize=128); _EVENT_SUBSCRIBERS.add(q)
    try:
        yield "event: ready\ndata: {}\n\n"
        while True:
            try:
                item=await asyncio.wait_for(q.get(),timeout=20.0)
                yield "data: %s\n\n" % json.dumps(item,ensure_ascii=False,separators=(",",":"))
            except asyncio.TimeoutError:
                yield ": keepalive\n\n"
    finally:
        _EVENT_SUBSCRIBERS.discard(q)


def notify_admin_text(text: str) -> bool:
    if not settings.telegram_bot_token or not settings.telegram_admin_chat_id:
        return False
    try:
        url="https://api.telegram.org/bot%s/sendMessage"%settings.telegram_bot_token
        body=json.dumps({"chat_id":settings.telegram_admin_chat_id,"text":str(text)[:3900]}).encode("utf-8")
        req=urllib.request.Request(url,data=body,headers={"Content-Type":"application/json","User-Agent":"PrintBotWeb/%s"%APP_VERSION})
        with urllib.request.urlopen(req,timeout=12) as resp: return 200 <= int(getattr(resp,"status",200)) < 300
    except Exception as exc:
        record_error("TELEGRAM_NOTIFY",str(exc)); return False


# ============================================================================
# SECURITY / LOGIN THROTTLE
# ============================================================================
_LOGIN_ATTEMPTS: Dict[str, List[float]] = {}

def _client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "").split(",")[0].strip()
    return forwarded or (request.client.host if request.client else "")

def _login_key(request: Request, username: str) -> str:
    return "%s|%s" % (_client_ip(request), (username or "").lower())

def login_rate_check(request: Request, username: str) -> int:
    key = _login_key(request, username); now=time.time(); window=max(60,settings.login_lockout_sec)
    arr=[t for t in _LOGIN_ATTEMPTS.get(key,[]) if now-t<window]; _LOGIN_ATTEMPTS[key]=arr
    if len(arr) >= max(1,settings.login_max_attempts):
        return max(1, int(window-(now-arr[0])))
    return 0

def login_rate_fail(request: Request, username: str) -> None:
    key=_login_key(request,username); _LOGIN_ATTEMPTS.setdefault(key,[]).append(time.time())

def login_rate_success(request: Request, username: str) -> None:
    _LOGIN_ATTEMPTS.pop(_login_key(request,username),None)

def _ensure_csrf(request: Request) -> str:
    token=str(request.session.get("csrf") or "")
    if not token:
        token=secrets.token_urlsafe(32); request.session["csrf"]=token
    return token

def _validate_csrf(request: Request) -> None:
    if request.method.upper() in {"GET","HEAD","OPTIONS"}: return
    expected=str(request.session.get("csrf") or ""); supplied=str(request.headers.get("x-csrf-token") or "")
    if not expected or not supplied or not constant_time_eq(expected,supplied):
        raise HTTPException(status_code=403, detail="CSRF token tidak valid.")

# ============================================================================
# FASTAPI APP
# ============================================================================
log = setup_logging("webapp")

# Event handler diregistrasikan lewat add_event_handler agar kompatibel dengan
# FastAPI lama maupun baru tanpa warning decorator on_event.
app = FastAPI(title="PrintBot WebApp - Local Printer", docs_url=None, redoc_url=None)
app.add_middleware(SessionMiddleware, secret_key=settings.webapp_secret_key, same_site="lax", https_only=settings.secure_cookie, max_age=settings.session_max_age_sec)

@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"]="nosniff"
    response.headers["X-Frame-Options"]="DENY"
    response.headers["Referrer-Policy"]="same-origin"
    response.headers["Permissions-Policy"]="camera=(), microphone=(), geolocation=()"
    response.headers["Content-Security-Policy"]="default-src 'self'; img-src 'self' data: blob:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; connect-src 'self' ws: wss:; object-src 'none'; frame-ancestors 'none'; base-uri 'self'"
    if settings.secure_cookie: response.headers["Strict-Transport-Security"]="max-age=31536000; includeSubDomains"
    return response

@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    record_error("HTTP_UNHANDLED",str(exc),"%s %s"%(request.method,request.url.path)); log.exception("Unhandled HTTP error: %s %s",request.method,request.url.path)
    return JSONResponse({"detail":"Terjadi error internal. Detail dicatat di Error Dashboard."},status_code=500)

_BACKGROUND_TASKS: List[asyncio.Task] = []
_TASK_ENSURED_AT_STARTUP = False


class PrintRequest(BaseModel):
    root: str
    path: str
    copies: int = Field(1, ge=1, le=99)
    pages: str = ""
    sheet: str = ""
    printer: str = ""
    paper: str = "AUTO"
    orientation: str = "AUTO"
    scale_mode: str = "FIT"
    duplex: str = "DEFAULT"
    color_mode: str = "DEFAULT"
    priority: int = Field(0, ge=0, le=100)
    scheduled_at: Optional[float] = None


class DevicePrintRequest(BaseModel):
    token: str
    copies: int = Field(1, ge=1, le=99)
    pages: str = ""
    sheet: str = ""
    printer: str = ""
    paper: str = "AUTO"
    orientation: str = "AUTO"
    scale_mode: str = "FIT"
    duplex: str = "DEFAULT"
    color_mode: str = "DEFAULT"
    priority: int = Field(0, ge=0, le=100)
    scheduled_at: Optional[float] = None


class RootAddRequest(BaseModel):
    label: str
    path: str = ""


class RootRemoveRequest(BaseModel):
    label: str


class RenameRequest(BaseModel):
    root: str
    path: str
    new_name: str


class DeleteRequest(BaseModel):
    root: str
    path: str


class MkdirRequest(BaseModel):
    root: str
    path: str = ""
    name: str


def require_login(request: Request) -> str:
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Belum login.")
    now=time.time(); created=float(request.session.get("created_at") or now); last=float(request.session.get("last_seen") or created)
    if now-created > max(300,settings.session_max_age_sec) or now-last > max(300,settings.session_idle_sec):
        request.session.clear(); raise HTTPException(status_code=401, detail="Sesi berakhir. Silakan login kembali.")
    request.session["last_seen"]=now; _ensure_csrf(request); _validate_csrf(request)
    return str(user)


def require_admin(request: Request) -> str:
    user = require_login(request)
    if request.session.get("role") != ROLE_ADMIN:
        raise HTTPException(status_code=403, detail="Hanya admin yang boleh melakukan aksi ini.")
    return user


def require_api_scope(request: Request, scope: str) -> str:
    token=(request.headers.get("x-api-key") or "").strip()
    auth=(request.headers.get("authorization") or "").strip()
    if auth.lower().startswith("bearer "):
        token=auth[7:].strip()
    if not settings.api_token:
        raise HTTPException(status_code=503, detail="WEBAPP_API_TOKEN belum dikonfigurasi.")
    if not token or not constant_time_eq(token,settings.api_token):
        raise HTTPException(status_code=401, detail="API token tidak valid.")
    if scope not in set(settings.api_scopes) and "admin" not in set(settings.api_scopes):
        raise HTTPException(status_code=403, detail="API token tidak memiliki scope '%s'."%scope)
    return "api"


async def _startup():
    if sys.platform.startswith("win") and not acquire_single_instance():
        raise RuntimeError("PrintBot WebApp sudah berjalan pada instance lain.")
    errs = settings.validate()
    for e in errs: log.error("[CONFIG ERROR] %s", e)
    ensure_system_tools_async()
    try:
        recovered = await _to_thread(job_recover_direct_after_restart)
        if recovered: log.warning("Startup recovery memproses %d job lokal yang terputus.", recovered)
    except Exception:
        log.exception("Startup queue recovery gagal")
    # Pastikan autostart juga terpasang ketika app dijalankan via `uvicorn module:app`.
    # Jika masuk melalui main(), task sudah diperiksa sekali sehingga tidak perlu dobel.
    if sys.platform.startswith("win") and settings.auto_start_task and not _TASK_ENSURED_AT_STARTUP:
        try:
            task_result=await _to_thread(ensure_windows_startup_task,False)
            if not task_result.get("ok"): record_error("TASK_SCHEDULER",str(task_result.get("message") or "Autostart gagal"))
        except Exception as exc: record_error("TASK_SCHEDULER",str(exc))
    _BACKGROUND_TASKS[:] = [
        asyncio.create_task(direct_print_worker_loop(), name="local-print-dispatcher"),
        asyncio.create_task(_sweep_loop(), name="print-job-sweeper"),
        asyncio.create_task(file_index_loop(), name="file-indexer"),
        asyncio.create_task(printer_monitor_loop(), name="printer-monitor"),
        asyncio.create_task(db_maintenance_loop(), name="db-maintenance"),
        asyncio.create_task(auto_update_loop(), name="auto-update"),
    ]
    log.info("WebApp Local Printer siap. Default printer=%s | Root=%s",
             local_default_printer() or "(belum ada)", ", ".join(get_roots().keys()) or "(belum diset)")


async def _shutdown():
    await _to_thread(stop_file_watchers)
    for task in list(_BACKGROUND_TASKS): task.cancel()
    if _BACKGROUND_TASKS:
        await asyncio.gather(*_BACKGROUND_TASKS, return_exceptions=True)
    _BACKGROUND_TASKS.clear()


app.add_event_handler("startup", _startup)
app.add_event_handler("shutdown", _shutdown)


async def _sweep_loop():
    while True:
        try:
            n = job_sweep_stuck()
            if n:
                log.warning("SWEEP: %d job stuck di-requeue/gagalkan.", n)
            await _to_thread(cleanup_device_print_uploads)
        except Exception:
            log.exception("sweep loop error")
        await asyncio.sleep(30)


# ----------------------------------------------------------------------
# LEGACY BRIDGE/GAS DINONAKTIFKAN
# ----------------------------------------------------------------------
# Route lama dipertahankan sebagai HTTP 410 agar konfigurasi bridge lama tidak
# diam-diam mengambil job. Semua job baru diproses oleh worker printer lokal.
@app.get("/api/bridge/{token}")
@app.post("/api/bridge/{token}")
async def bridge_disabled(token: str):
    raise HTTPException(status_code=410, detail="Print Bridge/GAS sudah dinonaktifkan. Gunakan printer lokal WebApp.")


@app.get("/api/bridge/{token}/file/{job_id}")
async def bridge_file_disabled(token: str, job_id: str):
    raise HTTPException(status_code=410, detail="Print Bridge/GAS sudah dinonaktifkan.")


# ============================================================================
# HALAMAN HTML (inline - login.html & index.html, CSS+JS sudah disatukan agar
# webapp.py tetap satu file mandiri, tanpa folder static/template terpisah)
# ============================================================================
LOGIN_HTML = r"""<!doctype html>
<html lang="id">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>Masuk - PrintBot</title>
<link rel="manifest" href="/manifest.webmanifest"><meta name="theme-color" content="#1e293b">
<script>(function(){try{var p=localStorage.getItem("printbot-theme")||"system";var d=p==="light"||p==="dark"?p:(window.matchMedia&&window.matchMedia("(prefers-color-scheme: light)").matches?"light":"dark");document.documentElement.setAttribute("data-theme",d);document.documentElement.setAttribute("data-theme-pref",p);}catch(e){}})();</script>
<style>
:root{
  --bg:#0f172a; --bg-soft:#111827; --panel:#1e293b; --panel-2:#243244;
  --border:#2d3b4f; --text:#e5e7eb; --muted:#94a3b8; --primary:#3b82f6;
  --primary-dark:#2563eb; --success:#22c55e; --warn:#f59e0b; --danger:#ef4444;
  --radius:12px; --sidebar-w:240px; --topbar-h:56px;
}
html[data-theme="light"]{
  --bg:#f3f6fb; --bg-soft:#eef2f7; --panel:#ffffff; --panel-2:#e8eef6;
  --border:#d5dde8; --text:#172033; --muted:#667085; --primary:#2563eb;
  --primary-dark:#1d4ed8; --success:#16a34a; --warn:#d97706; --danger:#dc2626;
  color-scheme:light;
}
html[data-theme="dark"]{color-scheme:dark}
*{box-sizing:border-box}
html,body{height:100%}
body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
  background:var(--bg);color:var(--text);-webkit-font-smoothing:antialiased}
img,svg{max-width:100%}
button,input,select{font:inherit;color:inherit}
a{color:inherit}
.muted{color:var(--muted)}
.small{font-size:.8rem}

/* ---------- layout ---------- */
.app{display:flex;flex-direction:column;height:100vh;min-height:100dvh;overflow:hidden}
.topbar{
  height:var(--topbar-h);flex:0 0 auto;display:flex;align-items:center;gap:.6rem;
  padding:0 .75rem;background:var(--panel);border-bottom:1px solid var(--border);
  padding-top:env(safe-area-inset-top);min-width:0;
}
.brand{display:flex;align-items:center;gap:.4rem;font-weight:700;white-space:nowrap;flex:0 0 auto}
.brand span{display:none}
@media(min-width:640px){.brand span{display:inline}}
.search-wrap{flex:1 1 auto;min-width:0}
#searchInput{
  width:100%;padding:.5rem .75rem;border-radius:999px;border:1px solid var(--border);
  background:var(--bg-soft);color:var(--text);min-width:0;
}
.icon-btn{
  background:transparent;border:none;color:var(--text);font-size:1.1rem;
  padding:.4rem .5rem;border-radius:8px;cursor:pointer;flex:0 0 auto;
  text-decoration:none;display:inline-flex;align-items:center;justify-content:center;
  transition:background-color .15s ease,transform .1s ease;
}
.icon-btn:hover{background:var(--panel-2)}
.icon-btn:active{transform:scale(.9)}
.only-mobile{display:inline-flex}
@media(min-width:900px){.only-mobile{display:none}}

.body{flex:1 1 auto;display:flex;min-height:0}
.sidebar{
  width:var(--sidebar-w);flex:0 0 auto;background:var(--panel);border-right:1px solid var(--border);
  padding:.75rem;overflow-y:auto;transition:transform .2s ease;
}
.sidebar-title{font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin:.75rem 0 .4rem}
.sidebar-title:first-child{margin-top:0}
.root-list,.jobs-mini{display:flex;flex-direction:column;gap:.25rem}
.root-item{
  display:flex;align-items:center;gap:.5rem;padding:.5rem .6rem;border-radius:8px;cursor:pointer;
  background:transparent;border:none;color:var(--text);text-align:left;width:100%;min-width:0;
  transition:background-color .15s ease;
}
.root-item span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.root-item:hover,.root-item.active{background:var(--panel-2)}
.job-mini{font-size:.78rem;padding:.35rem .5rem;border-radius:8px;background:var(--bg-soft);
  display:flex;justify-content:space-between;align-items:center;gap:.4rem;min-width:0}
.job-mini span:first-child{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}

@media(max-width:899px){
  .sidebar{position:fixed;top:var(--topbar-h);bottom:0;left:0;z-index:40;transform:translateX(-100%);
    box-shadow:2px 0 12px rgba(0,0,0,.4);width:min(80vw,300px)}
  .sidebar.open{transform:translateX(0)}
  .drawer-overlay{position:fixed;inset:0;top:var(--topbar-h);background:rgba(0,0,0,0);z-index:30;display:block;visibility:hidden;pointer-events:none;transition:background-color .2s ease,visibility 0s linear .2s}
  .drawer-overlay.open{visibility:visible;pointer-events:auto;background:rgba(0,0,0,.5);transition:background-color .2s ease,visibility 0s linear 0s}
}

.main{flex:1 1 auto;min-width:0;display:flex;flex-direction:column;padding:.75rem 1rem;overflow:hidden}
.breadcrumb{display:flex;flex-wrap:wrap;gap:.25rem;font-size:.85rem;color:var(--muted);min-height:1.4em;word-break:break-word}
.breadcrumb button{background:none;border:none;color:var(--primary);cursor:pointer;padding:0;font:inherit}
.breadcrumb .sep{color:var(--muted)}

.toolbar{display:flex;justify-content:space-between;align-items:center;gap:.5rem;margin:.5rem 0;flex-wrap:wrap}
.sort-group{display:flex;align-items:center;gap:.4rem;font-size:.85rem}
.sort-group select{background:var(--panel);border:1px solid var(--border);border-radius:8px;padding:.3rem .5rem}

.list-container{flex:1 1 auto;overflow-y:auto;border:1px solid var(--border);border-radius:var(--radius);
  background:var(--panel);min-height:0}
.empty-state{padding:2rem 1rem;text-align:center;color:var(--muted)}
.file-table{width:100%;border-collapse:collapse;table-layout:fixed}
.file-table th{position:sticky;top:0;background:var(--panel);text-align:left;font-size:.72rem;
  color:var(--muted);text-transform:uppercase;padding:.5rem .6rem;border-bottom:1px solid var(--border);z-index:1}
.file-table td{padding:.55rem .6rem;border-bottom:1px solid var(--border);vertical-align:middle;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:0}
.file-table th:first-child,.file-table td:first-child{width:2rem;max-width:2rem}
.file-row{cursor:pointer;transition:background-color .12s ease}
.file-row:hover{background:var(--panel-2)}
.file-row:active{background:var(--panel-2);filter:brightness(1.08)}
.file-name{display:flex;align-items:center;gap:.5rem;min-width:0}
.file-name span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.hide-sm{display:table-cell;width:110px}
@media(max-width:640px){.hide-sm{display:none}}

.pager{display:flex;justify-content:center;gap:.5rem;padding:.6rem 0;flex:0 0 auto}
.pager button{background:var(--panel);border:1px solid var(--border);border-radius:8px;padding:.4rem .8rem;
  color:var(--text);cursor:pointer}
.pager button:disabled{opacity:.4;cursor:not-allowed}

/* ---------- badges / toast ---------- */
.badge{font-size:.72rem;padding:.3rem .6rem;border-radius:999px;white-space:nowrap;flex:0 0 auto}
.badge-ok{background:rgba(34,197,94,.15);color:var(--success)}
.badge-danger{background:rgba(239,68,68,.15);color:var(--danger)}
.badge-muted{background:rgba(148,163,184,.15);color:var(--muted)}
.toast-host{position:fixed;bottom:1rem;left:50%;transform:translateX(-50%);z-index:100;
  display:flex;flex-direction:column;gap:.4rem;width:min(92vw,420px);pointer-events:none}
.toast{background:var(--panel-2);border:1px solid var(--border);border-radius:10px;padding:.6rem .9rem;
  font-size:.85rem;box-shadow:0 4px 16px rgba(0,0,0,.4);word-break:break-word;
  animation:toastIn .28s cubic-bezier(.34,1.56,.64,1)}
.toast.ok{border-color:var(--success)}
.toast.err{border-color:var(--danger)}
.toast.out{animation:toastOut .18s ease forwards}
@keyframes toastIn{from{opacity:0;transform:translateY(14px) scale(.96)} to{opacity:1;transform:translateY(0) scale(1)}}
@keyframes toastOut{to{opacity:0;transform:translateY(8px) scale(.95)}}

/* ---------- sheets / modals ---------- */
.sheet-overlay{position:fixed;inset:0;background:rgba(0,0,0,0);display:flex;align-items:flex-end;
  justify-content:center;z-index:60;padding:0;visibility:hidden;opacity:0;pointer-events:none;
  transition:opacity .22s ease,background-color .22s ease,visibility 0s linear .22s}
.sheet-overlay.open{visibility:visible;opacity:1;background:rgba(0,0,0,.55);pointer-events:auto;
  transition:opacity .22s ease,background-color .22s ease,visibility 0s linear 0s}
@media(min-width:640px){.sheet-overlay{align-items:center;padding:1rem}}
.sheet{
  background:var(--panel);width:100%;max-width:480px;border-radius:16px 16px 0 0;
  padding:1rem;max-height:88vh;overflow-y:auto;padding-bottom:calc(1rem + env(safe-area-inset-bottom));
  transform:translateY(28px);transition:transform .26s cubic-bezier(.22,1,.36,1);
}
.sheet-overlay.open .sheet{transform:translateY(0)}
@media(min-width:640px){
  .sheet{border-radius:16px;transform:translateY(10px) scale(.97)}
  .sheet-overlay.open .sheet{transform:translateY(0) scale(1)}
}
.sheet-tall{max-height:80vh}
.sheet-header{display:flex;justify-content:space-between;align-items:center;gap:.5rem;margin-bottom:.5rem}
.sheet-filename{font-weight:600;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.sheet-info{font-size:.85rem;color:var(--muted);margin-bottom:.75rem;word-break:break-word}
.sheet-actions{display:flex;gap:.5rem;flex-wrap:wrap}

.btn{background:var(--panel-2);border:1px solid var(--border);color:var(--text);border-radius:10px;
  padding:.6rem .9rem;cursor:pointer;flex:1 1 auto;min-width:0;
  transition:background-color .15s ease,filter .15s ease,transform .1s ease}
.btn:hover{filter:brightness(1.15)}
.btn:active{transform:scale(.96)}
.btn-primary{background:var(--primary);border-color:var(--primary-dark)}
.btn-primary:hover{background:var(--primary-dark)}
.btn-block{width:100%;margin-top:.75rem}

.print-form label{display:block;font-size:.8rem;color:var(--muted);margin:.7rem 0 .3rem}
.print-form input[type=text],.print-form input[type=number],.print-form select{
  width:100%;padding:.55rem .7rem;border-radius:10px;border:1px solid var(--border);background:var(--bg-soft);
}
.form-grid-2{display:grid;grid-template-columns:1fr 1fr;gap:.55rem}
.form-grid-2>div{display:flex;flex-direction:column;gap:.2rem}
.form-grid-2 label{margin:.2rem 0 .25rem}
@media(max-width:520px){.form-grid-2{grid-template-columns:1fr}}
.stepper{display:flex;align-items:center;gap:.5rem}
.stepper input{width:5rem;text-align:center;padding:.5rem;border-radius:10px;border:1px solid var(--border);
  background:var(--bg-soft)}
.stepper .btn{flex:0 0 auto;padding:.5rem .9rem}
.quick-copies{display:flex;flex-wrap:wrap;gap:.4rem;margin-top:.5rem}
.quick-copies button{flex:0 0 auto;padding:.35rem .7rem;border-radius:999px;border:1px solid var(--border);
  background:var(--bg-soft);color:var(--text);cursor:pointer}

.jobs-list{display:flex;flex-direction:column;gap:.5rem}
.job-card{background:var(--bg-soft);border:1px solid var(--border);border-radius:10px;padding:.6rem .75rem;min-width:0}
.job-card .row1{display:flex;justify-content:space-between;gap:.5rem;align-items:center;min-width:0}
.job-card .fname{font-weight:600;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.status-pill{font-size:.7rem;padding:.15rem .55rem;border-radius:999px;white-space:nowrap;flex:0 0 auto}
.status-QUEUED{background:rgba(148,163,184,.2);color:#cbd5e1}
.status-PROCESSING{background:rgba(245,158,11,.2);color:#fbbf24}
.status-PRINTING{background:rgba(59,130,246,.2);color:#93c5fd}
.status-SUCCESS{background:rgba(34,197,94,.2);color:#4ade80}
.status-FAILED{background:rgba(239,68,68,.2);color:#f87171}
.job-card .meta{font-size:.75rem;color:var(--muted);margin-top:.25rem;word-break:break-word}

.fab{position:fixed;right:1rem;bottom:calc(1rem + env(safe-area-inset-bottom));width:52px;height:52px;
  border-radius:50%;background:var(--primary);border:none;color:#fff;font-size:1.3rem;cursor:pointer;
  box-shadow:0 6px 16px rgba(0,0,0,.4);z-index:50}

/* ---------- login ---------- */
.login-body{display:flex;align-items:center;justify-content:center;min-height:100dvh;padding:1rem}
.login-card{background:var(--panel);border:1px solid var(--border);border-radius:16px;padding:2rem;
  width:100%;max-width:360px;display:flex;flex-direction:column;gap:.35rem}
.login-logo{font-size:2.5rem;text-align:center}
.login-card h1{margin:.25rem 0 0;text-align:center}
.login-card p{text-align:center;margin:0 0 .5rem}
.login-card label{font-size:.8rem;color:var(--muted);margin-top:.4rem}
.login-card input{padding:.6rem .7rem;border-radius:10px;border:1px solid var(--border);background:var(--bg-soft);
  color:var(--text)}
.login-card button{margin-top:1rem;padding:.65rem;border-radius:10px;border:none;background:var(--primary);
  color:#fff;font-weight:600;cursor:pointer}
.login-card .err{color:var(--danger);font-size:.85rem;text-align:center;margin:0}

</style>
</head>
<body class="login-body">
  <button id="loginThemeBtn" type="button" title="Ganti tema" style="position:fixed;right:1rem;top:1rem;border:1px solid var(--border);background:var(--panel);color:var(--text);width:42px;height:42px;border-radius:50%;cursor:pointer;font-size:1.1rem">◐</button>
  <form class="login-card" method="post" action="/login">
    <div class="login-logo">&#128424;&#65039;</div>
    <h1>PrintBot</h1>
    <p class="muted">Masuk untuk mengelola file &amp; print</p>
    <!--ERROR-->
    <label>Username</label>
    <input type="text" name="username" autocomplete="username" required autofocus>
    <label>Password</label>
    <input type="password" name="password" autocomplete="current-password" required>
    <button type="submit">Masuk</button>
  </form>
<script>
(function(){
  const media=window.matchMedia?window.matchMedia("(prefers-color-scheme: light)"):null;
  const btn=document.getElementById("loginThemeBtn");
  const prefs=["system","light","dark"];
  function apply(pref){const actual=pref==="light"||pref==="dark"?pref:(media&&media.matches?"light":"dark");document.documentElement.setAttribute("data-theme",actual);document.documentElement.setAttribute("data-theme-pref",pref);localStorage.setItem("printbot-theme",pref);btn.textContent=pref==="light"?"☀":pref==="dark"?"☾":"◐";btn.title="Tema: "+(pref==="system"?"System":pref==="light"?"Light":"Dark");}
  let pref=localStorage.getItem("printbot-theme")||"system";apply(pref);
  btn.onclick=()=>{const i=prefs.indexOf(localStorage.getItem("printbot-theme")||"system");apply(prefs[(i+1)%prefs.length]);};
  if(media){const f=()=>{if((localStorage.getItem("printbot-theme")||"system")==="system")apply("system");};if(media.addEventListener)media.addEventListener("change",f);else if(media.addListener)media.addListener(f);}
})();
</script>
</body>
</html>
"""
INDEX_HTML = r"""<!doctype html>
<html lang="id">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>PrintBot - File Manager</title>
<link rel="manifest" href="/manifest.webmanifest"><meta name="theme-color" content="#1e293b">
<script>(function(){try{var p=localStorage.getItem("printbot-theme")||"system";var d=p==="light"||p==="dark"?p:(window.matchMedia&&window.matchMedia("(prefers-color-scheme: light)").matches?"light":"dark");document.documentElement.setAttribute("data-theme",d);document.documentElement.setAttribute("data-theme-pref",p);}catch(e){}})();</script>
<style>
:root{
  --bg:#0f172a; --bg-soft:#111827; --panel:#1e293b; --panel-2:#243244;
  --border:#2d3b4f; --text:#e5e7eb; --muted:#94a3b8; --primary:#3b82f6;
  --primary-dark:#2563eb; --success:#22c55e; --warn:#f59e0b; --danger:#ef4444;
  --radius:12px; --sidebar-w:240px; --topbar-h:56px;
}
html[data-theme="light"]{
  --bg:#f3f6fb; --bg-soft:#eef2f7; --panel:#ffffff; --panel-2:#e8eef6;
  --border:#d5dde8; --text:#172033; --muted:#667085; --primary:#2563eb;
  --primary-dark:#1d4ed8; --success:#16a34a; --warn:#d97706; --danger:#dc2626;
  color-scheme:light;
}
html[data-theme="dark"]{color-scheme:dark}
*{box-sizing:border-box}
html,body{height:100%}
body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
  background:var(--bg);color:var(--text);-webkit-font-smoothing:antialiased}
img,svg{max-width:100%}
button,input,select{font:inherit;color:inherit}
a{color:inherit}
.muted{color:var(--muted)}
.small{font-size:.8rem}

/* ---------- layout ---------- */
.app{display:flex;flex-direction:column;height:100vh;min-height:100dvh;overflow:hidden}
.topbar{
  height:var(--topbar-h);flex:0 0 auto;display:flex;align-items:center;gap:.6rem;
  padding:0 .75rem;background:var(--panel);border-bottom:1px solid var(--border);
  padding-top:env(safe-area-inset-top);min-width:0;
}
.brand{display:flex;align-items:center;gap:.4rem;font-weight:700;white-space:nowrap;flex:0 0 auto}
.brand span{display:none}
@media(min-width:640px){.brand span{display:inline}}
.search-wrap{flex:1 1 auto;min-width:0}
#searchInput{
  width:100%;padding:.5rem .75rem;border-radius:999px;border:1px solid var(--border);
  background:var(--bg-soft);color:var(--text);min-width:0;
}
.icon-btn{
  background:transparent;border:none;color:var(--text);font-size:1.1rem;
  padding:.4rem .5rem;border-radius:8px;cursor:pointer;flex:0 0 auto;
  text-decoration:none;display:inline-flex;align-items:center;justify-content:center;
  transition:background-color .15s ease,transform .1s ease;
}
.icon-btn:hover{background:var(--panel-2)}
.icon-btn:active{transform:scale(.9)}
.only-mobile{display:inline-flex}
@media(min-width:900px){.only-mobile{display:none}}

.body{flex:1 1 auto;display:flex;min-height:0}
.sidebar{
  width:var(--sidebar-w);flex:0 0 auto;background:var(--panel);border-right:1px solid var(--border);
  padding:.75rem;overflow-y:auto;transition:transform .2s ease;
}
.sidebar-title{font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin:.75rem 0 .4rem}
.sidebar-title:first-child{margin-top:0}
.root-list,.jobs-mini{display:flex;flex-direction:column;gap:.25rem}
.root-item{
  display:flex;align-items:center;gap:.5rem;padding:.5rem .6rem;border-radius:8px;cursor:pointer;
  background:transparent;border:none;color:var(--text);text-align:left;width:100%;min-width:0;
  transition:background-color .15s ease;
}
.root-item span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.root-item:hover,.root-item.active{background:var(--panel-2)}
.job-mini{font-size:.78rem;padding:.35rem .5rem;border-radius:8px;background:var(--bg-soft);
  display:flex;justify-content:space-between;align-items:center;gap:.4rem;min-width:0}
.job-mini span:first-child{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}

@media(max-width:899px){
  .sidebar{position:fixed;top:var(--topbar-h);bottom:0;left:0;z-index:40;transform:translateX(-100%);
    box-shadow:2px 0 12px rgba(0,0,0,.4);width:min(80vw,300px)}
  .sidebar.open{transform:translateX(0)}
  .drawer-overlay{position:fixed;inset:0;top:var(--topbar-h);background:rgba(0,0,0,0);z-index:30;display:block;visibility:hidden;pointer-events:none;transition:background-color .2s ease,visibility 0s linear .2s}
  .drawer-overlay.open{visibility:visible;pointer-events:auto;background:rgba(0,0,0,.5);transition:background-color .2s ease,visibility 0s linear 0s}
}

.main{flex:1 1 auto;min-width:0;display:flex;flex-direction:column;padding:.75rem 1rem;overflow:hidden}
.breadcrumb{display:flex;flex-wrap:wrap;gap:.25rem;font-size:.85rem;color:var(--muted);min-height:1.4em;word-break:break-word}
.breadcrumb button{background:none;border:none;color:var(--primary);cursor:pointer;padding:0;font:inherit}
.breadcrumb .sep{color:var(--muted)}

.toolbar{display:flex;justify-content:space-between;align-items:center;gap:.5rem;margin:.5rem 0;flex-wrap:wrap}
.sort-group{display:flex;align-items:center;gap:.4rem;font-size:.85rem}
.sort-group select{background:var(--panel);border:1px solid var(--border);border-radius:8px;padding:.3rem .5rem}

.list-container{flex:1 1 auto;overflow-y:auto;border:1px solid var(--border);border-radius:var(--radius);
  background:var(--panel);min-height:0}
.empty-state{padding:2rem 1rem;text-align:center;color:var(--muted)}
.file-table{width:100%;border-collapse:collapse;table-layout:fixed}
.file-table th{position:sticky;top:0;background:var(--panel);text-align:left;font-size:.72rem;
  color:var(--muted);text-transform:uppercase;padding:.5rem .6rem;border-bottom:1px solid var(--border);z-index:1}
.file-table td{padding:.55rem .6rem;border-bottom:1px solid var(--border);vertical-align:middle;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:0}
.file-table th:first-child,.file-table td:first-child{width:2rem;max-width:2rem}
.file-row{cursor:pointer;transition:background-color .12s ease}
.file-row:hover{background:var(--panel-2)}
.file-row:active{background:var(--panel-2);filter:brightness(1.08)}
.file-name{display:flex;align-items:center;gap:.5rem;min-width:0}
.file-name span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.hide-sm{display:table-cell;width:110px}
@media(max-width:640px){.hide-sm{display:none}}

.pager{display:flex;justify-content:center;gap:.5rem;padding:.6rem 0;flex:0 0 auto}
.pager button{background:var(--panel);border:1px solid var(--border);border-radius:8px;padding:.4rem .8rem;
  color:var(--text);cursor:pointer}
.pager button:disabled{opacity:.4;cursor:not-allowed}

/* ---------- badges / toast ---------- */
.badge{font-size:.72rem;padding:.3rem .6rem;border-radius:999px;white-space:nowrap;flex:0 0 auto}
.badge-ok{background:rgba(34,197,94,.15);color:var(--success)}
.badge-danger{background:rgba(239,68,68,.15);color:var(--danger)}
.badge-muted{background:rgba(148,163,184,.15);color:var(--muted)}
.printer-dot{font-size:.95rem;line-height:1;display:inline-flex;align-items:center;justify-content:center}
.printer-status-text{margin-left:.25rem}
.theme-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:.55rem;margin-top:.8rem}
.theme-choice{display:flex;flex-direction:column;align-items:center;gap:.35rem;padding:.8rem .35rem!important;min-width:0}
.theme-choice .theme-icon{font-size:1.45rem}.theme-choice.active{border-color:var(--primary);box-shadow:0 0 0 2px rgba(59,130,246,.22)}
.confirm-body{padding:.35rem 0 .2rem}.confirm-message{white-space:pre-line;line-height:1.5;color:var(--text);margin:.4rem 0 1rem}
.confirm-actions{display:grid;grid-template-columns:1fr 1fr;gap:.55rem}.btn-danger{background:var(--danger);border-color:var(--danger);color:#fff}
@media(max-width:520px){
  .topbar{gap:.25rem;padding-left:.4rem;padding-right:.4rem}
  .topbar .brand{display:none}
  .topbar .icon-btn{padding:.38rem .4rem}
  #bridgeBadge{padding:.18rem;background:transparent!important;border-radius:50%;font-size:1rem;min-width:24px;text-align:center}
  #bridgeBadge .printer-status-text{display:none}
  #searchInput{padding:.48rem .65rem}
}

.toast-host{position:fixed;bottom:1rem;left:50%;transform:translateX(-50%);z-index:100;
  display:flex;flex-direction:column;gap:.4rem;width:min(92vw,420px);pointer-events:none}
.toast{background:var(--panel-2);border:1px solid var(--border);border-radius:10px;padding:.6rem .9rem;
  font-size:.85rem;box-shadow:0 4px 16px rgba(0,0,0,.4);word-break:break-word;
  animation:toastIn .28s cubic-bezier(.34,1.56,.64,1)}
.toast.ok{border-color:var(--success)}
.toast.err{border-color:var(--danger)}
.toast.out{animation:toastOut .18s ease forwards}
@keyframes toastIn{from{opacity:0;transform:translateY(14px) scale(.96)} to{opacity:1;transform:translateY(0) scale(1)}}
@keyframes toastOut{to{opacity:0;transform:translateY(8px) scale(.95)}}

/* ---------- sheets / modals ---------- */
.sheet-overlay{position:fixed;inset:0;background:rgba(0,0,0,0);display:flex;align-items:flex-end;
  justify-content:center;z-index:60;padding:0;visibility:hidden;opacity:0;pointer-events:none;
  transition:opacity .22s ease,background-color .22s ease,visibility 0s linear .22s}
.sheet-overlay.open{visibility:visible;opacity:1;background:rgba(0,0,0,.55);pointer-events:auto;
  transition:opacity .22s ease,background-color .22s ease,visibility 0s linear 0s}
@media(min-width:640px){.sheet-overlay{align-items:center;padding:1rem}}
.sheet{
  background:var(--panel);width:100%;max-width:480px;border-radius:16px 16px 0 0;
  padding:1rem;max-height:88vh;overflow-y:auto;padding-bottom:calc(1rem + env(safe-area-inset-bottom));
  transform:translateY(28px);transition:transform .26s cubic-bezier(.22,1,.36,1);
}
.sheet-overlay.open .sheet{transform:translateY(0)}
@media(min-width:640px){
  .sheet{border-radius:16px;transform:translateY(10px) scale(.97)}
  .sheet-overlay.open .sheet{transform:translateY(0) scale(1)}
}
.sheet-tall{max-height:80vh}
.sheet-header{display:flex;justify-content:space-between;align-items:center;gap:.5rem;margin-bottom:.5rem}
.sheet-filename{font-weight:600;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.sheet-info{font-size:.85rem;color:var(--muted);margin-bottom:.75rem;word-break:break-word}
.sheet-actions{display:flex;gap:.5rem;flex-wrap:wrap}

.btn{background:var(--panel-2);border:1px solid var(--border);color:var(--text);border-radius:10px;
  padding:.6rem .9rem;cursor:pointer;flex:1 1 auto;min-width:0;
  transition:background-color .15s ease,filter .15s ease,transform .1s ease}
.btn:hover{filter:brightness(1.15)}
.btn:active{transform:scale(.96)}
.btn-primary{background:var(--primary);border-color:var(--primary-dark)}
.btn-primary:hover{background:var(--primary-dark)}
.btn-block{width:100%;margin-top:.75rem}

.print-form label{display:block;font-size:.8rem;color:var(--muted);margin:.7rem 0 .3rem}
.print-form input[type=text],.print-form input[type=number],.print-form select{
  width:100%;padding:.55rem .7rem;border-radius:10px;border:1px solid var(--border);background:var(--bg-soft);
}
.form-grid-2{display:grid;grid-template-columns:1fr 1fr;gap:.55rem}
.form-grid-2>div{display:flex;flex-direction:column;gap:.2rem}
.form-grid-2 label{margin:.2rem 0 .25rem}
@media(max-width:520px){.form-grid-2{grid-template-columns:1fr}}
.stepper{display:flex;align-items:center;gap:.5rem}
.stepper input{width:5rem;text-align:center;padding:.5rem;border-radius:10px;border:1px solid var(--border);
  background:var(--bg-soft)}
.stepper .btn{flex:0 0 auto;padding:.5rem .9rem}
.quick-copies{display:flex;flex-wrap:wrap;gap:.4rem;margin-top:.5rem}
.quick-copies button{flex:0 0 auto;padding:.35rem .7rem;border-radius:999px;border:1px solid var(--border);
  background:var(--bg-soft);color:var(--text);cursor:pointer}

.jobs-list{display:flex;flex-direction:column;gap:.5rem}
.job-card{background:var(--bg-soft);border:1px solid var(--border);border-radius:10px;padding:.6rem .75rem;min-width:0}
.job-card .row1{display:flex;justify-content:space-between;gap:.5rem;align-items:center;min-width:0}
.job-card .fname{font-weight:600;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;min-width:0}
.status-pill{font-size:.7rem;padding:.15rem .55rem;border-radius:999px;white-space:nowrap;flex:0 0 auto}
.status-QUEUED{background:rgba(148,163,184,.2);color:#cbd5e1}
.status-PROCESSING{background:rgba(245,158,11,.2);color:#fbbf24}
.status-PRINTING{background:rgba(59,130,246,.2);color:#93c5fd}
.status-SUCCESS{background:rgba(34,197,94,.2);color:#4ade80}
.status-FAILED{background:rgba(239,68,68,.2);color:#f87171}
.job-card .meta{font-size:.75rem;color:var(--muted);margin-top:.25rem;word-break:break-word}

.fab{position:fixed;right:1rem;bottom:calc(1rem + env(safe-area-inset-bottom));width:52px;height:52px;
  border-radius:50%;background:var(--primary);border:none;color:#fff;font-size:1.3rem;cursor:pointer;
  box-shadow:0 6px 16px rgba(0,0,0,.4);z-index:50}

/* ---------- login ---------- */
.login-body{display:flex;align-items:center;justify-content:center;min-height:100dvh;padding:1rem}
.login-card{background:var(--panel);border:1px solid var(--border);border-radius:16px;padding:2rem;
  width:100%;max-width:360px;display:flex;flex-direction:column;gap:.35rem}
.login-logo{font-size:2.5rem;text-align:center}
.login-card h1{margin:.25rem 0 0;text-align:center}
.login-card p{text-align:center;margin:0 0 .5rem}
.login-card label{font-size:.8rem;color:var(--muted);margin-top:.4rem}
.login-card input{padding:.6rem .7rem;border-radius:10px;border:1px solid var(--border);background:var(--bg-soft);
  color:var(--text)}
.login-card button{margin-top:1rem;padding:.65rem;border-radius:10px;border:none;background:var(--primary);
  color:#fff;font-weight:600;cursor:pointer}
.login-card .err{color:var(--danger);font-size:.85rem;text-align:center;margin:0}
/* ---------- root management, method toggle, preview iframe, stats ---------- */
.sidebar-title-row{display:flex;align-items:center;justify-content:space-between;margin:.75rem 0 .4rem}
.sidebar-title-row .sidebar-title{margin:0}
.small-icon-btn{width:1.6rem;height:1.6rem;padding:0;border:1px solid var(--border);border-radius:6px;
  font-size:1rem;line-height:1;transition:background-color .15s ease,transform .1s ease}
.small-icon-btn:active{transform:scale(.9)}
.root-row{display:flex;align-items:center;gap:.15rem;min-width:0}
.root-row .root-item{flex:1 1 auto;min-width:0}
.root-del{flex:0 0 auto;background:transparent;border:none;color:var(--muted);cursor:pointer;
  padding:.3rem .5rem;border-radius:6px;font-size:.8rem;transition:background-color .15s ease,color .15s ease}
.root-del:hover{background:var(--panel-2);color:var(--danger)}
.small-btn{padding:.45rem;font-size:.82rem;margin-top:.4rem}
.badge-inline{font-size:.68rem;color:var(--muted);background:var(--bg-soft);border-radius:6px;
  padding:.05rem .4rem;margin-left:.4rem;flex:0 0 auto}
.pager-label{align-self:center;font-size:.82rem;color:var(--muted)}

.method-toggle{display:flex;gap:.5rem;margin-top:.3rem}
.method-btn{flex:1 1 0;padding:.55rem .4rem;border-radius:10px;border:1px solid var(--border);
  background:var(--bg-soft);color:var(--text);cursor:pointer;font-size:.85rem;
  transition:background-color .15s ease,border-color .15s ease,transform .1s ease}
.method-btn:active{transform:scale(.97)}
.method-btn.active{background:var(--primary);border-color:var(--primary-dark);color:#fff}

.sheet-wide{max-width:720px}
.preview-frame-wrap{background:#111;border-radius:10px;overflow:hidden;min-height:55vh}
.preview-frame{width:100%;height:65vh;border:0;background:#0b0b0b;display:block}

.stats-mini{display:flex;flex-direction:column;gap:.15rem;font-size:.78rem;color:var(--muted)}
.stats-mini-row{padding:.1rem 0}
.stats-content{display:flex;flex-direction:column;gap:.3rem}
.stats-summary{display:flex;gap:.5rem;margin-bottom:.5rem;flex-wrap:wrap}
.stat-card{flex:1 1 90px;background:var(--bg-soft);border:1px solid var(--border);border-radius:10px;
  padding:.6rem;text-align:center;min-width:90px;transition:transform .15s ease}
.stat-card:hover{transform:translateY(-2px)}
.stat-card-icon{font-size:1.1rem}
.stat-card-value{font-size:1.3rem;font-weight:700;margin:.1rem 0}
.stat-card-label{font-size:.72rem;color:var(--muted)}
.stats-section-title{font-size:.78rem;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);
  margin:.6rem 0 .2rem}
.stats-list-row{font-size:.85rem;padding:.15rem 0;word-break:break-word}

/* ---------- file CRUD: rename/delete/mkdir/upload ---------- */
.btn-danger{background:rgba(239,68,68,.12);border-color:var(--danger);color:#fca5a5}
.btn-danger:hover{background:rgba(239,68,68,.22)}
.toolbar-right{display:flex;align-items:center;gap:.4rem;flex-wrap:wrap}
.col-menu{width:2.1rem}
.row-actions-cell{width:2.1rem;text-align:center;padding:.2rem !important}
.row-menu-btn{background:transparent;border:none;color:var(--muted);font-size:1.05rem;cursor:pointer;
  padding:.25rem .5rem;border-radius:6px;line-height:1;transition:background-color .15s ease,color .15s ease}
.row-menu-btn:hover{background:var(--panel-2);color:var(--text)}
#textPromptInput{margin-top:.3rem}

/* ---------- loading indicator, list fade-in, fab press ---------- */
.list-container{position:relative}
.loading-bar{position:absolute;top:0;left:0;right:0;height:2px;overflow:hidden;z-index:2;
  opacity:0;transition:opacity .15s ease}
.loading-bar.active{opacity:1}
.loading-bar::after{content:"";position:absolute;top:0;left:-35%;height:100%;width:35%;
  background:var(--primary);animation:loadingSlide 1s ease-in-out infinite}
@keyframes loadingSlide{0%{left:-35%} 100%{left:100%}}
#fileTbody{animation:fadeInRows .18s ease}
@keyframes fadeInRows{from{opacity:.35} to{opacity:1}}
.fab{transition:transform .15s ease,box-shadow .15s ease}
.fab:active{transform:scale(.92)}
.action-progress{position:fixed;inset:0;z-index:120;display:flex;align-items:center;justify-content:center;
  background:rgba(2,6,23,.5);backdrop-filter:blur(2px);visibility:hidden;opacity:0;pointer-events:none;
  transition:opacity .16s ease,visibility 0s linear .16s}
.action-progress.open{visibility:visible;opacity:1;pointer-events:auto;transition:opacity .16s ease}
.action-progress-card{min-width:190px;max-width:86vw;background:var(--panel);border:1px solid var(--border);
  border-radius:15px;padding:1rem 1.25rem;text-align:center;box-shadow:0 16px 50px rgba(0,0,0,.38)}
.action-spinner{width:32px;height:32px;margin:0 auto .65rem;border:4px solid var(--panel-2);
  border-top-color:var(--primary);border-radius:50%;animation:actionSpin .7s linear infinite}
@keyframes actionSpin{to{transform:rotate(360deg)}}
.btn.is-busy,.icon-btn.is-busy{position:relative;pointer-events:none;opacity:.75}
.job-actions{display:flex;gap:.45rem;flex-wrap:wrap;margin-top:.55rem}
.job-action-btn{border:0;border-radius:8px;padding:.42rem .65rem;background:var(--primary);color:#fff;
  font-size:.75rem;font-weight:600;cursor:pointer;transition:transform .1s ease,filter .15s ease}
.job-action-btn:active{transform:scale(.96)}.job-action-btn:hover{filter:brightness(1.12)}
.job-action-btn.cancel{background:var(--danger)}
.device-drop-overlay{position:fixed;inset:0;z-index:9999;background:rgba(15,23,42,.88);display:flex;
  align-items:center;justify-content:center;opacity:0;visibility:hidden;pointer-events:none;transition:.15s ease}
.device-drop-overlay.open{opacity:1;visibility:visible}
.device-drop-card{border:2px dashed var(--primary);border-radius:18px;padding:2.3rem;max-width:520px;width:calc(100% - 2rem);
  background:var(--panel);text-align:center;box-shadow:0 20px 70px rgba(0,0,0,.45)}
.device-drop-card .drop-icon{font-size:3rem;margin-bottom:.65rem}.device-drop-card b{font-size:1.12rem}

</style>
</head>
<body>
<div id="app" class="app">

  <header class="topbar">
    <button id="btnMenu" class="icon-btn only-mobile" aria-label="Menu">&#9776;</button>
    <div class="brand">&#128424;&#65039; <span>PrintBot</span></div>
    <div class="search-wrap">
      <input id="searchInput" type="search" placeholder="Cari dokumen..." autocomplete="off">
    </div>
    <div id="bridgeBadge" class="badge badge-muted" title="Status printer"><span class="printer-dot">⚪</span><span class="printer-status-text">Printer: memeriksa...</span></div>
    <button id="btnTheme" class="icon-btn" type="button" title="Tema tampilan" aria-label="Tema tampilan">◐</button>
    <button id="btnLogout" class="icon-btn" type="button" title="Keluar" aria-label="Keluar"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg></button>
  </header>

  <div class="body">
    <nav id="sidebar" class="sidebar">
      <div class="sidebar-title-row">
        <div class="sidebar-title">Root Folder</div>
        <button id="btnAddRoot" class="icon-btn small-icon-btn" title="Tambah Folder">+</button>
      </div>
      <div id="rootList" class="root-list"></div>
      <div class="sheet-actions"><button id="btnFavorites" class="btn small-btn">⭐ Favorit</button><button id="btnRecent" class="btn small-btn">🕘 Terbaru</button></div>
      <div class="sidebar-title">Antrean Print</div>
      <div id="jobsMini" class="jobs-mini"></div>
      <div class="sidebar-title">Statistik</div>
      <div id="statsMini" class="stats-mini"></div>
      <button id="btnStatsDetail" class="btn btn-block small-btn">&#128200; Lihat Detail</button>
      <div class="sidebar-title" id="adminSectionTitle">Admin</div>
      <button id="btnManageUsers" class="btn btn-block small-btn" type="button">&#128100; Kelola User</button>
      <button id="btnPrinterTools" class="btn btn-block small-btn" type="button">&#128424;&#65039; Printer &amp; Sistem</button>
    </nav>

    <main class="main">
      <div class="breadcrumb" id="breadcrumb"></div>
      <div class="toolbar">
        <div class="sort-group">
          <label for="sortSelect">Urutkan:</label>
          <select id="sortSelect">
            <option value="name">Nama</option>
            <option value="size">Ukuran</option>
            <option value="type">Jenis</option>
            <option value="modified">Diubah</option>
          </select>
        </div>
        <div class="toolbar-right">
          <button id="btnNewFolder" class="btn small-btn" type="button">&#128193;+ Folder</button>
          <button id="btnDevicePrint" class="btn btn-primary small-btn" type="button">&#128424;&#65039; Print dari Perangkat</button>
          <input id="devicePrintInput" type="file" accept=".pdf,.doc,.docx,.xls,.xlsx,.xlsm,.ppt,.pptx,.txt,.rtf,.jpg,.jpeg,.png,.bmp,.gif,.tif,.tiff" style="display:none">
          <button id="btnUpload" class="btn small-btn" type="button">&#11014;&#65039; Upload</button>
          <input id="uploadFileInput" type="file" style="display:none">
          <span id="itemCount" class="muted"></span>
        </div>
      </div>

      <div id="listContainer" class="list-container">
        <div class="loading-bar" id="loadingBar"></div>
        <div class="empty-state" id="emptyState">Pilih folder di sisi kiri untuk mulai.</div>
        <table class="file-table" id="fileTable" style="display:none">
          <thead><tr><th></th><th>Nama</th><th class="hide-sm">Ukuran</th><th class="hide-sm">Diubah</th><th class="col-menu"></th></tr></thead>
          <tbody id="fileTbody"></tbody>
        </table>
      </div>

      <div class="pager" id="pager"></div>
    </main>
  </div>

  <div id="drawerOverlay" class="drawer-overlay"></div>
  <div id="deviceDropOverlay" class="device-drop-overlay">
    <div class="device-drop-card"><div class="drop-icon">&#128229;</div><b>Lepaskan file untuk dicetak</b>
      <div class="muted small" style="margin-top:.55rem">PDF, Word, Excel, PowerPoint, gambar, TXT/RTF</div></div>
  </div>

  <div id="fileSheet" class="sheet-overlay">
    <div class="sheet">
      <div class="sheet-header">
        <div id="sheetFileName" class="sheet-filename"></div>
        <button class="icon-btn" id="sheetClose">&#10005;</button>
      </div>
      <div id="sheetInfo" class="sheet-info"></div>
      <div class="sheet-actions">
        <button id="actOpenFolder" class="btn" style="display:none">&#128193; Buka</button>
        <button id="actPreview" class="btn">&#128065;&#65039; Preview</button>
        <button id="actDownload" class="btn">&#11015;&#65039; Download</button>
        <button id="actPrint" class="btn btn-primary">&#128424;&#65039; Print</button>
        <button id="actFavorite" class="btn">⭐ Favorit</button>
      </div>
      <div class="sheet-actions">
        <button id="actRename" class="btn">&#9999;&#65039; Ganti Nama</button>
        <button id="actDelete" class="btn btn-danger">&#128465;&#65039; Hapus</button>
      </div>
    </div>
  </div>

  <div id="confirmModal" class="sheet-overlay" role="dialog" aria-modal="true" aria-labelledby="confirmTitle">
    <div class="sheet">
      <div class="sheet-header">
        <div id="confirmTitle" class="sheet-filename">Konfirmasi</div>
        <button class="icon-btn" id="confirmClose" type="button">&#10005;</button>
      </div>
      <div class="confirm-body">
        <div id="confirmMessage" class="confirm-message"></div>
        <div class="confirm-actions">
          <button id="confirmCancel" class="btn" type="button">Batal</button>
          <button id="confirmOk" class="btn btn-primary" type="button">Lanjutkan</button>
        </div>
      </div>
    </div>
  </div>

  <div id="themeModal" class="sheet-overlay" role="dialog" aria-modal="true" aria-labelledby="themeTitle">
    <div class="sheet">
      <div class="sheet-header">
        <div id="themeTitle" class="sheet-filename">Tema Tampilan</div>
        <button class="icon-btn" id="themeClose" type="button">&#10005;</button>
      </div>
      <div class="muted small">Pilih tema. Preferensi disimpan di perangkat ini.</div>
      <div class="theme-grid">
        <button class="btn theme-choice" type="button" data-theme-choice="light"><span class="theme-icon">☀️</span><span>Light</span></button>
        <button class="btn theme-choice" type="button" data-theme-choice="dark"><span class="theme-icon">🌙</span><span>Dark</span></button>
        <button class="btn theme-choice" type="button" data-theme-choice="system"><span class="theme-icon">💻</span><span>System</span></button>
      </div>
    </div>
  </div>

  <div id="textPromptModal" class="sheet-overlay">
    <div class="sheet">
      <div class="sheet-header">
        <div id="textPromptTitle" class="sheet-filename">Judul</div>
        <button class="icon-btn" id="textPromptClose">&#10005;</button>
      </div>
      <div class="print-form">
        <input id="textPromptInput" type="text">
        <button id="textPromptSubmit" class="btn btn-primary btn-block">OK</button>
      </div>
    </div>
  </div>

  <div id="printModal" class="sheet-overlay">
    <div class="sheet">
      <div class="sheet-header">
        <div class="sheet-filename">Cetak Dokumen</div>
        <button class="icon-btn" id="printClose">&#10005;</button>
      </div>
      <div class="print-form">
        <div id="printFileName" class="muted"></div>

        <div id="printerPickWrap">
          <label for="printerSelect">Printer Lokal</label>
          <select id="printerSelect"></select>
          <div id="printerStatusHint" class="muted small"></div>
        </div>

        <div class="form-grid-2">
          <div><label for="paperSelect">Ukuran Kertas</label><select id="paperSelect">
            <option value="AUTO">Auto/Driver</option><option value="A4">A4</option><option value="F4">F4/Folio</option>
            <option value="LEGAL">Legal</option><option value="LETTER">Letter</option>
          </select></div>
          <div><label for="orientationSelect">Orientasi</label><select id="orientationSelect">
            <option value="AUTO">Auto</option><option value="PORTRAIT">Portrait</option><option value="LANDSCAPE">Landscape</option>
          </select></div>
          <div><label for="scaleSelect">Skala</label><select id="scaleSelect">
            <option value="FIT">Fit to Page</option><option value="ACTUAL">Actual Size</option>
          </select></div>
          <div><label for="duplexSelect">Duplex</label><select id="duplexSelect">
            <option value="DEFAULT">Default Driver</option><option value="OFF">Simplex</option>
            <option value="LONG">Duplex Long Edge</option><option value="SHORT">Duplex Short Edge</option>
          </select></div>
          <div><label for="colorSelect">Warna</label><select id="colorSelect">
            <option value="DEFAULT">Default Driver</option><option value="COLOR">Color</option><option value="MONO">Hitam Putih</option>
          </select></div>
        </div>
        <label for="profileSelect">Print Profile</label>
        <div class="form-grid-2"><select id="profileSelect"><option value="">(tanpa preset)</option></select><div class="sheet-actions"><button id="saveProfileBtn" class="btn" type="button">Simpan</button><button id="deleteProfileBtn" class="btn btn-danger" type="button">Hapus</button></div></div>
        <div class="form-grid-2">
          <div><label for="prioritySelect">Prioritas</label><select id="prioritySelect"><option value="0">Normal</option><option value="50">High</option><option value="100">Urgent</option></select></div>
          <div><label for="scheduleInput">Jadwal (opsional)</label><input id="scheduleInput" type="datetime-local"></div>
        </div>

        <label for="copiesInput">Jumlah Salinan</label>
        <div class="stepper">
          <button id="copiesMinus" class="btn" type="button">-</button>
          <input id="copiesInput" type="number" min="1" max="99" value="1">
          <button id="copiesPlus" class="btn" type="button">+</button>
        </div>
        <div class="quick-copies" id="quickCopies"></div>
        <label for="pagesInput">Halaman (opsional)</label>
        <input id="pagesInput" type="text" placeholder="contoh: 1-5 atau 2,4,6">
        <div id="sheetPickWrap" style="display:none">
          <label for="sheetSelect">Sheet Excel (opsional)</label>
          <select id="sheetSelect"></select>
        </div>
        <p class="muted small">Print diproses berurutan oleh printer lokal. PDF/gambar dikirim sebagai satu Windows spool job;
          Word/Excel/PowerPoint memakai Microsoft Office COM bila tersedia. Untuk Excel/PowerPoint, pilihan halaman non-berurutan
          seperti 2,4,6 ditolak agar tidak mencetak halaman yang salah.</p>
        <button id="printSubmit" class="btn btn-primary btn-block">Tambahkan ke Antrean Lokal</button>
      </div>
    </div>
  </div>

  <div id="previewModal" class="sheet-overlay">
    <div class="sheet sheet-tall sheet-wide">
      <div class="sheet-header">
        <div id="previewFileName" class="sheet-filename"></div>
        <div class="sheet-actions"><button id="previewPrev" class="btn small-btn">←</button><span id="previewPageLabel" class="muted small">1 / 1</span><button id="previewNext" class="btn small-btn">→</button></div>
        <button class="icon-btn" id="previewClose">&#10005;</button>
      </div>
      <div class="preview-frame-wrap">
        <iframe id="previewFrame" class="preview-frame" title="Preview dokumen"></iframe>
      </div>
    </div>
  </div>

  <div id="rootModal" class="sheet-overlay">
    <div class="sheet">
      <div class="sheet-header">
        <div class="sheet-filename">Tambah Folder Root</div>
        <button class="icon-btn" id="rootModalClose">&#10005;</button>
      </div>
      <div class="print-form">
        <label for="rootLabelInput">Nama Folder</label>
        <input id="rootLabelInput" type="text" placeholder="contoh: Arsip 2026">
        <label for="rootPathInput">Path di Server (opsional)</label>
        <input id="rootPathInput" type="text" placeholder="Kosongkan untuk membuat folder otomatis">
        <p class="muted small">Cukup isi nama folder - folder akan <b>dibuat otomatis</b> di server dan langsung
          bisa diakses, tanpa perlu tahu/isi path absolut. Isi path hanya jika ingin menghubungkan ke folder
          tertentu (folder akan dibuat otomatis juga kalau belum ada). Setelah ditambahkan, langsung muncul di
          sini maupun di Telegram Bot - tanpa restart.</p>
        <button id="rootAddSubmit" class="btn btn-primary btn-block">Tambahkan Folder</button>
      </div>
    </div>
  </div>

  <div id="jobsPanel" class="sheet-overlay">
    <div class="sheet sheet-tall">
      <div class="sheet-header">
        <div class="sheet-filename">Antrean &amp; Riwayat Print</div>
        <button class="icon-btn" id="jobsClose">&#10005;</button>
      </div>
      <div id="jobsList" class="jobs-list"></div>
    </div>
  </div>

  <div id="statsPanel" class="sheet-overlay">
    <div class="sheet sheet-tall">
      <div class="sheet-header">
        <div class="sheet-filename">&#128200; Statistik Print</div>
        <button class="icon-btn" id="statsClose">&#10005;</button>
      </div>
      <div id="statsContent" class="stats-content"></div>
    </div>
  </div>

  <div id="printerToolsPanel" class="sheet-overlay">
    <div class="sheet sheet-tall">
      <div class="sheet-header">
        <div class="sheet-filename">&#128424;&#65039; Printer &amp; Diagnostic</div>
        <button class="icon-btn" id="printerToolsClose">&#10005;</button>
      </div>
      <div id="diagnosticSummary" class="stats-content"></div>
      <div class="sheet-actions">
        <button id="queuePauseBtn" class="btn">&#9208; Pause Queue</button>
        <button id="queueResumeBtn" class="btn btn-primary">&#9654; Resume Queue</button>
      </div>
      <div id="printerToolsList" class="jobs-list"></div>
      <div class="sheet-actions">
        <button id="taskRepairBtn" class="btn">Task Scheduler Repair</button>
        <button id="restartBtn" class="btn btn-danger">Restart WebApp</button>
        <button id="backupNowBtn" class="btn">Backup Config/DB</button>
        <button id="cleanupNowBtn" class="btn">Cleanup</button>
      </div>
      <div class="sheet-actions">
        <button id="maintenanceBtn" class="btn">Maintenance Mode</button>
        <button id="reindexBtn" class="btn">Reindex File</button>
        <button id="dbMaintBtn" class="btn">DB Maintenance</button>
      </div>
      <div class="sheet-actions">
        <button id="backupManagerBtn" class="btn">Backup / Restore</button>
        <button id="exportXlsxBtn" class="btn">Export XLSX</button>
        <button id="exportCsvBtn" class="btn">Export CSV</button>
        <button id="auditBtn" class="btn">Audit & Error</button>
        <button id="storageBtn" class="btn">Storage</button>
        <button id="apiTokenBtn" class="btn">API Token</button>
        <button id="notifBtn" class="btn">Notifikasi Printer</button>
      </div>
      <div class="sheet-actions">
        <button id="updateCheckBtn" class="btn">Check Update</button>
        <button id="updateApplyBtn" class="btn">Apply Update</button>
        <button id="updateRollbackBtn" class="btn">Rollback</button>
      </div>
      <div id="advancedAdminOutput" class="stats-content"></div>
    </div>
  </div>

  <div id="usersPanel" class="sheet-overlay">
    <div class="sheet sheet-tall">
      <div class="sheet-header">
        <div class="sheet-filename">&#128100; Kelola User</div>
        <button class="icon-btn" id="usersClose">&#10005;</button>
      </div>
      <div id="usersList" class="jobs-list"></div>
      <button id="btnAddUser" class="btn btn-primary btn-block" type="button">+ Tambah User</button>
    </div>
  </div>

  <div id="addUserModal" class="sheet-overlay">
    <div class="sheet">
      <div class="sheet-header">
        <div class="sheet-filename">Tambah User Baru</div>
        <button class="icon-btn" id="addUserClose">&#10005;</button>
      </div>
      <div class="print-form">
        <label for="newUserUsername">Username</label>
        <input id="newUserUsername" type="text" autocomplete="off">
        <label for="newUserPassword">Password</label>
        <input id="newUserPassword" type="text" autocomplete="off">
        <label for="newUserRole">Role</label>
        <select id="newUserRole">
          <option value="user">User (baca, cari, preview, download, print)</option>
          <option value="admin">Admin (akses penuh + kelola folder &amp; user)</option>
        </select>
        <p class="muted small">Role <b>User</b> tidak bisa ganti nama, hapus, tambah folder, atau upload file.</p>
        <button id="newUserSubmit" class="btn btn-primary btn-block">Tambah User</button>
      </div>
    </div>
  </div>

  <button id="fabJobs" class="fab" title="Riwayat Print">&#128203;</button>
  <div id="actionProgress" class="action-progress" role="status" aria-live="polite">
    <div class="action-progress-card"><div class="action-spinner"></div><div id="actionProgressText">Memproses...</div></div>
  </div>
  <div id="toastHost" class="toast-host"></div>
</div>
<script>
(() => {
  "use strict";
  const state = {
    root: null, path: "", page: 0, sort: "name", selected: null,
    mode: "browse", searchResults: [], searchPage: 0, searchQuery: "",
    pageSize: 20, localPrintAvailable: false, localPrinters: [], defaultPrinter: "",
    role: "user", username: "", csrf: "", maintenance: false, previewObjectUrl: null, previewPage: 1, previewPages: 1,
  };

  const $ = (sel) => document.querySelector(sel);
  const el = (tag, cls, txt) => {
    const e = document.createElement(tag);
    if (cls) e.className = cls;
    if (txt !== undefined) e.textContent = txt;
    return e;
  };

  async function api(path, opts = {}) {
    const { silent = false, ...fetchOpts } = opts;
    if (!silent) showLoading();
    try {
      const res = await fetch(path, {
        credentials: "same-origin",
        headers: { "Content-Type": "application/json", "X-CSRF-Token": state.csrf || "" },
        ...fetchOpts,
      });
      if (res.status === 401) { window.location.href = "/login"; throw new Error("unauth"); }
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.detail || "Terjadi kesalahan.");
      return data;
    } finally { if (!silent) hideLoading(); }
  }

  function toast(msg, kind = "ok") {
    const host = $("#toastHost");
    const t = el("div", `toast ${kind}`, msg);
    host.appendChild(t);
    setTimeout(() => {
      t.classList.add("out");
      setTimeout(() => t.remove(), 200);
    }, 3800);
  }

  let _confirmResolve = null;
  function closeConfirm(result = false) {
    $("#confirmModal").classList.remove("open");
    const resolve = _confirmResolve; _confirmResolve = null;
    if (resolve) resolve(!!result);
  }
  function uiConfirm(message, options = {}) {
    if (_confirmResolve) closeConfirm(false);
    $("#confirmTitle").textContent = options.title || "Konfirmasi";
    $("#confirmMessage").textContent = message || "Lanjutkan tindakan ini?";
    $("#confirmOk").textContent = options.okText || "Lanjutkan";
    $("#confirmCancel").textContent = options.cancelText || "Batal";
    $("#confirmOk").className = options.danger ? "btn btn-danger" : "btn btn-primary";
    $("#confirmModal").classList.add("open");
    setTimeout(() => $("#confirmOk").focus(), 60);
    return new Promise((resolve) => { _confirmResolve = resolve; });
  }
  $("#confirmClose").onclick = () => closeConfirm(false);
  $("#confirmCancel").onclick = () => closeConfirm(false);
  $("#confirmOk").onclick = () => closeConfirm(true);
  $("#confirmModal").addEventListener("click", (e) => { if (e.target === $("#confirmModal")) closeConfirm(false); });

  const themeMedia = window.matchMedia ? window.matchMedia("(prefers-color-scheme: light)") : null;
  function effectiveTheme(pref) {
    return pref === "light" || pref === "dark" ? pref : (themeMedia && themeMedia.matches ? "light" : "dark");
  }
  function applyTheme(pref, persist = true) {
    pref = ["light", "dark", "system"].includes(pref) ? pref : "system";
    const actual = effectiveTheme(pref);
    document.documentElement.setAttribute("data-theme", actual);
    document.documentElement.setAttribute("data-theme-pref", pref);
    if (persist) localStorage.setItem("printbot-theme", pref);
    const meta = document.querySelector('meta[name="theme-color"]');
    if (meta) meta.setAttribute("content", actual === "light" ? "#ffffff" : "#1e293b");
    $("#btnTheme").textContent = pref === "light" ? "☀" : (pref === "dark" ? "☾" : "◐");
    document.querySelectorAll("[data-theme-choice]").forEach((b) => b.classList.toggle("active", b.dataset.themeChoice === pref));
  }
  function currentThemePref(){ return localStorage.getItem("printbot-theme") || "system"; }
  $("#btnTheme").onclick = () => { applyTheme(currentThemePref(), false); $("#themeModal").classList.add("open"); };
  $("#themeClose").onclick = () => $("#themeModal").classList.remove("open");
  $("#themeModal").addEventListener("click", (e) => { if (e.target === $("#themeModal")) $("#themeModal").classList.remove("open"); });
  document.querySelectorAll("[data-theme-choice]").forEach((b) => b.onclick = () => { applyTheme(b.dataset.themeChoice); $("#themeModal").classList.remove("open"); toast(`Tema: ${b.textContent.trim()}`, "ok"); });
  if (themeMedia) { const fn = () => { if (currentThemePref() === "system") applyTheme("system", false); }; if (themeMedia.addEventListener) themeMedia.addEventListener("change", fn); else if (themeMedia.addListener) themeMedia.addListener(fn); }
  applyTheme(currentThemePref(), false);
  $("#btnLogout").onclick = async () => { if (await uiConfirm("Keluar dari PrintBot sekarang?", {title:"Logout", okText:"Keluar", danger:true})) window.location.href = "/logout"; };

  let loadingDepth = 0;
  function showLoading() { loadingDepth += 1; $("#loadingBar").classList.add("active"); }
  function hideLoading() {
    loadingDepth = Math.max(0, loadingDepth - 1);
    if (!loadingDepth) $("#loadingBar").classList.remove("active");
  }
  function showActionProgress(message, button = null) {
    $("#actionProgressText").textContent = message || "Memproses...";
    $("#actionProgress").classList.add("open");
    if (button) button.classList.add("is-busy");
  }
  function hideActionProgress(button = null) {
    $("#actionProgress").classList.remove("open");
    if (button) button.classList.remove("is-busy");
  }

  function fmtSize(n) {
    if (n < 1024) return n + " B";
    const units = ["KB", "MB", "GB"];
    let i = -1;
    do { n /= 1024; i++; } while (n >= 1024 && i < units.length - 1);
    return n.toFixed(1) + " " + units[i];
  }
  function fmtDate(ts) {
    if (!ts) return "-";
    const d = new Date(ts * 1000);
    return d.toLocaleDateString("id-ID") + " " + d.toLocaleTimeString("id-ID", { hour: "2-digit", minute: "2-digit" });
  }
  function iconFor(item) {
    if (item.is_dir) return "\u{1F4C1}";
    const ext = (item.ext || "").toLowerCase();
    if (ext === ".pdf") return "\u{1F4D5}";
    if ([".doc", ".docx"].includes(ext)) return "\u{1F4D8}";
    if ([".xls", ".xlsx", ".xlsm"].includes(ext)) return "\u{1F4D7}";
    if ([".jpg", ".jpeg", ".png", ".gif", ".bmp"].includes(ext)) return "\u{1F5BC}\uFE0F";
    return "\u{1F4C4}";
  }

  // -------------------------------------------------------------- config --
  async function loadConfig() {
    try {
      const data = await api("/api/config");
      state.pageSize = data.page_size || 20;
    } catch (e) { /* keep default */ }
  }

  // -------------------------------------------------------------- roots --
  async function loadRoots() {
    const { roots } = await api("/api/roots");
    const box = $("#rootList");
    box.innerHTML = "";
    roots.forEach((r) => {
      const row = el("div", "root-row");
      const b = el("button", "root-item");
      b.dataset.root = r.label;
      b.innerHTML = `\u{1F4C1} <span>${r.label}</span>`;
      b.onclick = () => openRoot(r.label);
      row.appendChild(b);
      if (r.dynamic && state.role === "admin") {
        const del = el("button", "root-del", "\u2715");
        del.title = "Hapus folder ini";
        del.onclick = async (ev) => {
          ev.stopPropagation();
          if (!(await uiConfirm(`Hapus folder root "${r.label}" dari daftar?
File di dalamnya tidak akan dihapus.`, {title:"Hapus Root Folder", okText:"Hapus", danger:true}))) return;
          try {
            await api("/api/roots/remove", { method: "POST", body: JSON.stringify({ label: r.label }) });
            toast(`Folder "${r.label}" dihapus.`, "ok");
            await loadRoots();
            if (state.root === r.label) { state.root = null; loadRoots(); }
          } catch (e) { toast(e.message, "err"); }
        };
        row.appendChild(del);
      }
      box.appendChild(row);
    });
    if (roots.length && !state.root) openRoot(roots[0].label);
    else if (!roots.length) {
      $("#emptyState").style.display = "block";
      $("#emptyState").textContent = "Belum ada folder root. Klik tombol + di atas untuk menambahkan.";
      $("#fileTable").style.display = "none";
    }
  }

  function openRoot(root) {
    state.root = root; state.path = ""; state.page = 0; state.mode = "browse";
    document.querySelectorAll(".root-item").forEach((b) => b.classList.toggle("active", b.dataset.root === root));
    closeSidebarMobile();
    $("#searchInput").value = "";
    loadBrowse();
  }

  $("#btnAddRoot").onclick = () => {
    $("#rootLabelInput").value = "";
    $("#rootPathInput").value = "";
    $("#rootModal").classList.add("open");
  };
  $("#rootModalClose").onclick = () => $("#rootModal").classList.remove("open");
  $("#rootAddSubmit").onclick = async () => {
    const label = $("#rootLabelInput").value.trim();
    const path = $("#rootPathInput").value.trim();
    if (!label) { toast("Nama folder wajib diisi.", "err"); return; }
    try {
      const info = await api("/api/roots/add", { method: "POST", body: JSON.stringify({ label, path }) });
      toast(info.auto_created ? `Folder "${label}" dibuat otomatis & ditambahkan.` : `Folder "${label}" ditambahkan.`, "ok");
      $("#rootModal").classList.remove("open");
      await loadRoots();
    } catch (e) { toast(e.message, "err"); }
  };

  // ------------------------------------------------------------- browse --
  async function loadBrowse() {
    if (!state.root) return;
    showLoading();
    try {
      const qs = `root=${encodeURIComponent(state.root)}&path=${encodeURIComponent(state.path)}&page=${state.page}&sort=${state.sort}`;
      const data = await api(`/api/browse?${qs}`);
      renderBreadcrumb(data.breadcrumb);
      renderList(data.items, data.total, data.page_size, data.has_prev, data.has_next);
    } catch (e) { toast(e.message, "err"); } finally { hideLoading(); }
  }

  function renderBreadcrumb(crumbs) {
    const box = $("#breadcrumb");
    box.innerHTML = "";
    crumbs.forEach((c, i) => {
      if (i > 0) box.appendChild(el("span", "sep", "/"));
      const b = el("button", null, c.label);
      b.onclick = () => { state.mode = "browse"; state.path = c.rel_path; state.page = 0; loadBrowse(); };
      box.appendChild(b);
    });
  }

  function renderList(items, total, pageSize, hasPrev, hasNext) {
    $("#emptyState").style.display = items.length ? "none" : "block";
    $("#emptyState").textContent = "Tidak ada file/folder di sini.";
    $("#fileTable").style.display = items.length ? "table" : "none";
    $("#itemCount").textContent = `${total} item`;
    const tbody = $("#fileTbody");
    tbody.innerHTML = "";
    items.forEach((item) => {
      const tr = el("tr", "file-row");
      const tdIcon = el("td", null, iconFor(item));
      const tdName = el("td");
      const nameWrap = el("div", "file-name");
      const nameSpan = el("span", null, item.name);
      nameWrap.appendChild(nameSpan);
      if (state.mode === "search" && item.root) {
        nameWrap.appendChild(el("span", "badge-inline", item.root));
      }
      tdName.appendChild(nameWrap);
      const tdSize = el("td", "hide-sm", item.is_dir ? "-" : fmtSize(item.size));
      const tdMod = el("td", "hide-sm", fmtDate(item.modified));
      const tdMenu = el("td", "row-actions-cell");
      const menuBtn = el("button", "row-menu-btn", "\u22EE");
      menuBtn.type = "button";
      menuBtn.title = "Opsi lainnya";
      menuBtn.onclick = (ev) => { ev.stopPropagation(); openFileSheet(item); };
      tdMenu.appendChild(menuBtn);
      tr.append(tdIcon, tdName, tdSize, tdMod, tdMenu);
      tr.onclick = () => (item.is_dir ? enterDir(item) : openFileSheet(item));
      tbody.appendChild(tr);
    });
    renderPager(total, pageSize, hasPrev, hasNext);
  }

  function refreshCurrentView() {
    if (state.mode === "search") doSearch(state.searchQuery);
    else loadBrowse();
  }

  function enterDir(item) {
    state.mode = "browse"; state.path = item.rel_path; state.page = 0; loadBrowse();
  }

  function renderPager(total, pageSize, hasPrev, hasNext) {
    const box = $("#pager");
    box.innerHTML = "";
    if (total <= pageSize) return;
    const curPage = state.mode === "search" ? state.searchPage : state.page;
    const totalPages = Math.max(1, Math.ceil(total / pageSize));
    const prev = el("button", null, "\u25C0 Sebelumnya");
    prev.disabled = !hasPrev;
    prev.onclick = () => { pageStep(-1); };
    const label = el("span", "pager-label", `Hal ${curPage + 1}/${totalPages}`);
    const next = el("button", null, "Berikutnya \u25B6");
    next.disabled = !hasNext;
    next.onclick = () => { pageStep(1); };
    box.append(prev, label, next);
  }

  function pageStep(delta) {
    if (state.mode === "search") {
      state.searchPage += delta;
      renderSearchPage();
    } else {
      state.page += delta;
      loadBrowse();
    }
  }

  // ------------------------------------------------------------- search --
  let searchTimer = null;
  $("#searchInput").addEventListener("input", (e) => {
    clearTimeout(searchTimer);
    const q = e.target.value.trim();
    if (!q) { state.mode = "browse"; $("#pager").innerHTML = ""; loadBrowse(); return; }
    searchTimer = setTimeout(() => doSearch(q), 350);
  });

  async function doSearch(q) {
    if (q.length < 2) return;
    showLoading();
    try {
      const data = await api(`/api/search?q=${encodeURIComponent(q)}`);
      state.mode = "search";
      state.searchResults = data.results;
      state.searchPage = 0;
      state.searchQuery = q;
      $("#breadcrumb").innerHTML = "";
      $("#breadcrumb").appendChild(el("span", null, `Hasil cari: "${q}" (${data.count})`));
      renderSearchPage();
    } catch (e) { toast(e.message, "err"); } finally { hideLoading(); }
  }

  function renderSearchPage() {
    const total = state.searchResults.length;
    const start = state.searchPage * state.pageSize;
    const pageItems = state.searchResults.slice(start, start + state.pageSize);
    const hasPrev = state.searchPage > 0;
    const hasNext = start + state.pageSize < total;
    renderList(pageItems, total, state.pageSize, hasPrev, hasNext);
  }

  // ---------------------------------------------------------- file sheet --
  function openFileSheet(item) {
    state.selected = item;
    $("#sheetFileName").textContent = item.name;
    $("#sheetInfo").textContent = item.is_dir ? "Folder" : `${fmtSize(item.size)} \u2022 Diubah ${fmtDate(item.modified)}`;
    const isDir = !!item.is_dir;
    const isAdmin = state.role === "admin";
    $("#actOpenFolder").style.display = isDir ? "" : "none";
    $("#actPreview").style.display = isDir ? "none" : "";
    $("#actDownload").style.display = isDir ? "none" : "";
    $("#actPrint").style.display = isDir ? "none" : "";
    $("#actFavorite").style.display = isDir ? "none" : "";
    $("#actRename").style.display = isAdmin ? "" : "none";
    $("#actDelete").style.display = isAdmin ? "" : "none";
    $("#fileSheet").classList.add("open");
  }
  $("#sheetClose").onclick = () => $("#fileSheet").classList.remove("open");
  $("#actFavorite").onclick = async () => {
    const it=state.selected;if(!it)return;
    try{const d=await api("/api/favorites/toggle",{method:"POST",body:JSON.stringify({root:it.root||state.root,path:it.rel_path})});toast(d.favorite?"Ditambahkan ke Favorit.":"Dihapus dari Favorit.","ok");}catch(e){toast(e.message,"err");}
  };

  $("#actOpenFolder").onclick = () => {
    const it = state.selected;
    if (!it) return;
    $("#fileSheet").classList.remove("open");
    enterDir(it);
  };

  $("#actDownload").onclick = () => {
    const it = state.selected;
    if (!it) return;
    const url = `/api/download?root=${encodeURIComponent(it.root || state.root)}&path=${encodeURIComponent(it.rel_path)}`;
    showActionProgress("Menyiapkan download...", $("#actDownload"));
    window.open(url, "_blank");
    setTimeout(() => hideActionProgress($("#actDownload")), 900);
  };

  // ------------------------------------------------ preview (gesture zoom/pan) --
  function buildPreviewSrcdoc(imageUrl) {
    return [
      '<!doctype html><html><head><meta charset="utf-8">',
      '<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1, user-scalable=no">',
      '<style>',
      'html,body{margin:0;padding:0;height:100%;background:#0b0b0b;overflow:hidden;touch-action:none;',
      '  -webkit-user-select:none;user-select:none}',
      '#stage{width:100%;height:100%;display:flex;align-items:center;justify-content:center;',
      '  overflow:hidden;position:relative}',
      '#img{max-width:none;max-height:none;-webkit-user-drag:none;transform-origin:0 0;',
      '  will-change:transform;cursor:grab;transition:transform .06s linear;opacity:0}',
      '#img.ready{opacity:1;transition:opacity .25s ease,transform .06s linear}',
      '#img.dragging{cursor:grabbing;transition:none}',
      '#ctrl{position:fixed;left:50%;bottom:14px;transform:translateX(-50%);display:flex;gap:8px;',
      '  background:rgba(20,20,20,.72);backdrop-filter:blur(4px);padding:6px;border-radius:999px;',
      '  box-shadow:0 4px 14px rgba(0,0,0,.4)}',
      '#ctrl button{width:36px;height:36px;border-radius:50%;border:none;background:rgba(255,255,255,.12);',
      '  color:#fff;font-size:18px;cursor:pointer;display:flex;align-items:center;justify-content:center}',
      '#ctrl button:active{background:rgba(255,255,255,.28)}',
      '#hint{position:fixed;top:10px;left:50%;transform:translateX(-50%);color:rgba(255,255,255,.55);',
      '  font:11px -apple-system,sans-serif;pointer-events:none;text-align:center;transition:opacity .4s ease}',
      '</style></head><body>',
      '<div id="stage"><img id="img" src="' + imageUrl + '" draggable="false"></div>',
      '<div id="hint">Cubit untuk zoom &#8226; Geser untuk pindah &#8226; Ketuk 2x untuk reset</div>',
      '<div id="ctrl">',
      '<button id="zOut" type="button">&#8722;</button>',
      '<button id="zReset" type="button">&#8226;</button>',
      '<button id="zIn" type="button">&#43;</button>',
      '</div>',
      '<script>',
      '(function(){',
      '  var img = document.getElementById("img");',
      '  var stage = document.getElementById("stage");',
      '  var hint = document.getElementById("hint");',
      '  var scale = 1, tx = 0, ty = 0, minScale = 1, maxScale = 6;',
      '  var fitAttempts = 0, fitTimer = 0;',
      '  var pointers = {}, lastDist = 0, dragging = false, lastX = 0, lastY = 0, lastTapTime = 0;',
      '  function apply(animated){',
      '    img.style.transition = animated ? "transform .18s ease" : "none";',
      '    img.style.transform = "translate(" + tx + "px," + ty + "px) scale(" + scale + ")";',
      '  }',
      '  function clampScale(s){ return Math.max(minScale, Math.min(maxScale, s)); }',
      '  function center(){',
      '    var r = stage.getBoundingClientRect();',
      '    tx = (r.width - img.naturalWidth * scale) / 2;',
      '    ty = (r.height - img.naturalHeight * scale) / 2;',
      '  }',
      '  function fitAndShow(){',
      '    var r = stage.getBoundingClientRect();',
      '    if(!img.naturalWidth || r.width < 2 || r.height < 2){',
      '      if(fitAttempts++ < 50){ clearTimeout(fitTimer); fitTimer = setTimeout(fitAndShow, 40); }',
      '      return;',
      '    }',
      '    fitAttempts = 0;',
      '    var fit = Math.min(r.width / img.naturalWidth, r.height / img.naturalHeight, 1);',
      '    minScale = fit; scale = fit;',
      '    center();',
      '    apply(false);',
      '    img.classList.add("ready");',
      '    setTimeout(function(){ hint.style.opacity = "0"; }, 2200);',
      '  }',
      '  if (img.complete && img.naturalWidth) fitAndShow(); else img.addEventListener("load", fitAndShow);',
      '  if(window.ResizeObserver){ new ResizeObserver(function(){ if(img.naturalWidth) fitAndShow(); }).observe(stage); }',
      '  requestAnimationFrame(function(){ requestAnimationFrame(fitAndShow); });',
      '  function dist(a,b){ var dx=a.x-b.x, dy=a.y-b.y; return Math.sqrt(dx*dx+dy*dy); }',
      '  function mid(a,b){ return {x:(a.x+b.x)/2, y:(a.y+b.y)/2}; }',
      '  function zoomAt(mx, my, factor){',
      '    var newScale = clampScale(scale * factor);',
      '    var imgX = (mx - tx) / scale, imgY = (my - ty) / scale;',
      '    tx = mx - imgX * newScale; ty = my - imgY * newScale;',
      '    scale = newScale;',
      '    apply(false);',
      '  }',
      '  stage.addEventListener("pointerdown", function(e){',
      '    stage.setPointerCapture(e.pointerId);',
      '    pointers[e.pointerId] = {x:e.clientX, y:e.clientY};',
      '    var pts = Object.values(pointers);',
      '    if(pts.length === 1){',
      '      dragging = true; lastX = e.clientX; lastY = e.clientY; img.classList.add("dragging");',
      '    } else if(pts.length === 2){',
      '      dragging = false; lastDist = dist(pts[0], pts[1]);',
      '    }',
      '  });',
      '  stage.addEventListener("pointermove", function(e){',
      '    if(!(e.pointerId in pointers)) return;',
      '    pointers[e.pointerId] = {x:e.clientX, y:e.clientY};',
      '    var pts = Object.values(pointers);',
      '    if(pts.length === 2){',
      '      var d = dist(pts[0], pts[1]);',
      '      var m = mid(pts[0], pts[1]);',
      '      var r = stage.getBoundingClientRect();',
      '      if(lastDist > 0) zoomAt(m.x - r.left, m.y - r.top, d / lastDist);',
      '      lastDist = d;',
      '    } else if(pts.length === 1 && dragging){',
      '      tx += e.clientX - lastX; ty += e.clientY - lastY;',
      '      lastX = e.clientX; lastY = e.clientY;',
      '      apply(false);',
      '    }',
      '  });',
      '  function endPointer(e){',
      '    delete pointers[e.pointerId];',
      '    var pts = Object.values(pointers);',
      '    if(pts.length < 2) lastDist = 0;',
      '    if(pts.length === 0){',
      '      dragging = false; img.classList.remove("dragging");',
      '      var now = Date.now();',
      '      if(now - lastTapTime < 280){',
      '        var r = stage.getBoundingClientRect();',
      '        var mx = e.clientX - r.left, my = e.clientY - r.top;',
      '        if(scale > minScale * 1.4){ scale = minScale; center(); apply(true); }',
      '        else zoomAtAnimated(mx, my, 2.5 / scale);',
      '      }',
      '      lastTapTime = now;',
      '    }',
      '  }',
      '  function zoomAtAnimated(mx, my, factor){',
      '    var newScale = clampScale(scale * factor);',
      '    var imgX = (mx - tx) / scale, imgY = (my - ty) / scale;',
      '    tx = mx - imgX * newScale; ty = my - imgY * newScale;',
      '    scale = newScale;',
      '    apply(true);',
      '  }',
      '  stage.addEventListener("pointerup", endPointer);',
      '  stage.addEventListener("pointercancel", endPointer);',
      '  stage.addEventListener("wheel", function(e){',
      '    e.preventDefault();',
      '    var r = stage.getBoundingClientRect();',
      '    zoomAt(e.clientX - r.left, e.clientY - r.top, e.deltaY < 0 ? 1.18 : 1/1.18);',
      '  }, {passive:false});',
      '  document.getElementById("zIn").addEventListener("click", function(){',
      '    var r = stage.getBoundingClientRect();',
      '    zoomAtAnimated(r.width/2, r.height/2, 1.5);',
      '  });',
      '  document.getElementById("zOut").addEventListener("click", function(){',
      '    var r = stage.getBoundingClientRect();',
      '    zoomAtAnimated(r.width/2, r.height/2, 1/1.5);',
      '  });',
      '  document.getElementById("zReset").addEventListener("click", function(){',
      '    scale = minScale; center(); apply(true);',
      '  });',
      '  window.addEventListener("resize", function(){ minScale = 0; fitAndShow(); });',
      '  window.addEventListener("pageshow", fitAndShow);',
      '})();',
      '</' + 'script>',
      '</body></html>',
    ].join("\n");
  }

  async function loadPreviewPage(pageNum) {
    const it=state.selected;if(!it)return;
    const root=it.root||state.root; state.previewPage=Math.max(1,Math.min(state.previewPages,pageNum));
    const url=`/api/preview/page?root=${encodeURIComponent(root)}&path=${encodeURIComponent(it.rel_path)}&page=${state.previewPage}&t=${Date.now()}`;
    const res=await fetch(url,{credentials:"same-origin"});if(!res.ok){const d=await res.json().catch(()=>({}));throw new Error(d.detail||"Preview tidak tersedia.");}
    const blob=await res.blob();if(state.previewObjectUrl)URL.revokeObjectURL(state.previewObjectUrl);state.previewObjectUrl=URL.createObjectURL(blob);
    $("#previewPageLabel").textContent=`${state.previewPage} / ${state.previewPages}`;$("#previewPrev").disabled=state.previewPage<=1;$("#previewNext").disabled=state.previewPage>=state.previewPages;
    $("#previewFrame").srcdoc=buildPreviewSrcdoc(state.previewObjectUrl);
  }
  $("#actPreview").onclick = async () => {
    const it = state.selected;if(!it)return;const root=it.root||state.root;const button=$("#actPreview");showActionProgress("Menyiapkan preview...",button);
    try{
      const meta=await api(`/api/preview/meta?root=${encodeURIComponent(root)}&path=${encodeURIComponent(it.rel_path)}`,{silent:true});state.previewPages=Math.max(1,+meta.pages||1);state.previewPage=1;
      $("#previewFileName").textContent=it.name;$("#fileSheet").classList.remove("open");$("#previewModal").classList.add("open");await new Promise(r=>requestAnimationFrame(()=>requestAnimationFrame(r)));await loadPreviewPage(1);
    }catch(e){toast(e.message,"err");$("#previewModal").classList.remove("open");}finally{hideActionProgress(button);}
  };
  $("#previewPrev").onclick=async()=>{try{await loadPreviewPage(state.previewPage-1);}catch(e){toast(e.message,"err");}};
  $("#previewNext").onclick=async()=>{try{await loadPreviewPage(state.previewPage+1);}catch(e){toast(e.message,"err");}};
  $("#previewClose").onclick = () => {
    $("#previewModal").classList.remove("open");
    $("#previewFrame").removeAttribute("srcdoc");
    $("#previewFrame").src = "about:blank";
    if (state.previewObjectUrl) {
      URL.revokeObjectURL(state.previewObjectUrl);
      state.previewObjectUrl = null;
    }
  };

  $("#actPrint").onclick = () => { $("#fileSheet").classList.remove("open"); openPrintModal(state.selected); };

  // ------------------------------------------------------- rename/delete --
  let _textPromptCallback = null;
  function openTextPrompt(title, initialValue, onSubmit) {
    $("#textPromptTitle").textContent = title;
    $("#textPromptInput").value = initialValue || "";
    _textPromptCallback = onSubmit;
    $("#textPromptModal").classList.add("open");
    setTimeout(() => $("#textPromptInput").focus(), 50);
  }
  $("#textPromptClose").onclick = () => $("#textPromptModal").classList.remove("open");
  $("#textPromptSubmit").onclick = () => {
    const val = $("#textPromptInput").value.trim();
    $("#textPromptModal").classList.remove("open");
    if (_textPromptCallback) _textPromptCallback(val);
  };
  $("#textPromptInput").addEventListener("keydown", (e) => {
    if (e.key === "Enter") { e.preventDefault(); $("#textPromptSubmit").click(); }
  });

  $("#actRename").onclick = () => {
    const it = state.selected;
    if (!it) return;
    openTextPrompt("Ganti Nama", it.name, async (newName) => {
      if (!newName || newName === it.name) return;
      try {
        await api("/api/rename", {
          method: "POST",
          body: JSON.stringify({ root: it.root || state.root, path: it.rel_path, new_name: newName }),
        });
        toast(`Diganti nama jadi "${newName}".`, "ok");
        $("#fileSheet").classList.remove("open");
        refreshCurrentView();
      } catch (e) { toast(e.message, "err"); }
    });
  };

  $("#actDelete").onclick = async () => {
    const it = state.selected;
    if (!it) return;
    const warn = it.is_dir ? " Folder beserta SEMUA ISI di dalamnya akan ikut terhapus." : "";
    if (!(await uiConfirm(`Hapus "${it.name}"?${warn}\n\nTindakan ini tidak bisa dibatalkan.`, {title:"Hapus Permanen", okText:"Hapus", danger:true}))) return;
    try {
      await api("/api/delete", {
        method: "POST",
        body: JSON.stringify({ root: it.root || state.root, path: it.rel_path }),
      });
      toast(`"${it.name}" dihapus.`, "ok");
      $("#fileSheet").classList.remove("open");
      refreshCurrentView();
    } catch (e) { toast(e.message, "err"); }
  };

  // ------------------------------------------ print file dari perangkat --
  async function stageDevicePrintFile(file) {
    if (!file) return;
    const fd = new FormData(); fd.append("file", file);
    const btn = $("#btnDevicePrint"); showActionProgress(`Menyiapkan "${file.name}"...`, btn);
    try {
      const res = await fetch("/api/device-print/upload", {method:"POST", credentials:"same-origin",
        headers:{"X-CSRF-Token":state.csrf||""}, body:fd});
      if (res.status === 401) { window.location.href="/login"; return; }
      const data = await res.json().catch(()=>({}));
      if (!res.ok) throw new Error(data.detail || "Gagal mengambil file dari perangkat.");
      const item={name:data.name,ext:data.ext,size:data.size||0,modified:Date.now()/1000,is_dir:false,
        source_type:"device",device_token:data.token,sheets:data.sheets||[]};
      toast(`File "${data.name}" siap diatur untuk print.`,"ok");
      await openPrintModal(item);
    } catch(e) { toast(e.message,"err"); }
    finally { hideActionProgress(btn); }
  }
  $("#btnDevicePrint").onclick=()=>$("#devicePrintInput").click();
  $("#devicePrintInput").addEventListener("change",async(e)=>{const f=e.target.files&&e.target.files[0];e.target.value="";if(f)await stageDevicePrintFile(f);});
  let deviceDragDepth=0; const deviceDrop=$("#deviceDropOverlay");
  document.addEventListener("dragenter",(e)=>{e.preventDefault();deviceDragDepth++;deviceDrop.classList.add("open");});
  document.addEventListener("dragover",(e)=>{e.preventDefault();deviceDrop.classList.add("open");});
  document.addEventListener("dragleave",(e)=>{e.preventDefault();deviceDragDepth=Math.max(0,deviceDragDepth-1);if(!deviceDragDepth)deviceDrop.classList.remove("open");});
  document.addEventListener("drop",async(e)=>{e.preventDefault();deviceDragDepth=0;deviceDrop.classList.remove("open");const files=e.dataTransfer&&e.dataTransfer.files;if(files&&files.length){if(files.length>1)toast("Gunakan satu file per proses print. File pertama dipilih.","ok");await stageDevicePrintFile(files[0]);}});

  // ------------------------------------------------------- mkdir/upload --
  $("#btnNewFolder").onclick = () => {
    if (!state.root) { toast("Pilih folder root dulu.", "err"); return; }
    if (state.mode === "search") { toast("Tutup pencarian dulu untuk membuat folder di sini.", "err"); return; }
    openTextPrompt("Buat Folder Baru", "", async (name) => {
      if (!name) return;
      try {
        await api("/api/mkdir", { method: "POST", body: JSON.stringify({ root: state.root, path: state.path, name }) });
        toast(`Folder "${name}" dibuat.`, "ok");
        loadBrowse();
      } catch (e) { toast(e.message, "err"); }
    });
  };

  $("#btnUpload").onclick = () => {
    if (!state.root) { toast("Pilih folder root dulu.", "err"); return; }
    if (state.mode === "search") { toast("Tutup pencarian dulu untuk mengunggah ke sini.", "err"); return; }
    $("#uploadFileInput").click();
  };
  $("#uploadFileInput").addEventListener("change", async (e) => {
    const f = e.target.files[0];
    e.target.value = "";
    if (!f) return;
    const fd = new FormData();
    fd.append("root", state.root);
    fd.append("path", state.path);
    fd.append("file", f);
    toast(`Mengunggah "${f.name}"...`, "ok");
    const uploadButton = $("#btnUpload");
    showActionProgress(`Mengunggah "${f.name}"...`, uploadButton);
    try {
      const res = await fetch("/api/upload", { method: "POST", credentials: "same-origin", headers:{"X-CSRF-Token":state.csrf||""}, body: fd });
      if (res.status === 401) { window.location.href = "/login"; return; }
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.detail || "Upload gagal.");
      toast(`File "${data.name}" berhasil diunggah.`, "ok");
      loadBrowse();
    } catch (err) { toast(err.message, "err"); }
    finally { hideActionProgress(uploadButton); }
  });

  // --------------------------------------------------------- print modal --
  async function loadPrinterOptions() {
    const sel = $("#printerSelect");
    sel.innerHTML = "";
    try {
      const data = await api("/api/local-printers", { silent: true });
      state.localPrintAvailable = !!data.available;
      state.localPrinters = data.printers || [];
      state.defaultPrinter = data.default_printer || "";
      if (!state.localPrinters.length) {
        const o = el("option", null, "(Tidak ada printer terdeteksi)"); o.value = ""; sel.appendChild(o);
        $("#printerStatusHint").textContent = "Printer lokal tidak ditemukan. Pastikan driver printer terpasang.";
        return;
      }
      state.localPrinters.forEach((p) => {
        const o = el("option", null, p + (p === state.defaultPrinter ? " (Default)" : ""));
        o.value = p; if (p === state.defaultPrinter) o.selected = true; sel.appendChild(o);
      });
      const details = data.details || [];
      const current = details.find((d) => d.printer === sel.value);
      $("#printerStatusHint").textContent = current ? `Status: ${current.message} • ${current.jobs || 0} job di spooler` : "Printer siap dipilih.";
    } catch (e) {
      const o = el("option", null, "(Gagal membaca printer)"); o.value = ""; sel.appendChild(o);
      $("#printerStatusHint").textContent = e.message || "Gagal membaca printer lokal.";
    }
  }

  $("#printerSelect").addEventListener("change", async () => {
    try {
      const data = await api(`/api/printers/status?name=${encodeURIComponent($("#printerSelect").value)}`, { silent: true });
      $("#printerStatusHint").textContent = `Status: ${data.message} • ${data.jobs || 0} job di spooler`;
    } catch (e) { $("#printerStatusHint").textContent = e.message; }
  });

  async function loadProfiles() {
    const sel=$("#profileSelect");sel.innerHTML='<option value="">(tanpa preset)</option>';
    try{const d=await api("/api/profiles",{silent:true});(d.profiles||[]).forEach(p=>{const o=el("option",null,p.name);o.value=p.name;o.dataset.profile=JSON.stringify(p);sel.appendChild(o);});}catch(e){}
  }
  $("#profileSelect").onchange=()=>{const o=$("#profileSelect").selectedOptions[0];if(!o||!o.dataset.profile)return;const p=JSON.parse(o.dataset.profile);if(p.printer&&state.localPrinters.includes(p.printer))$("#printerSelect").value=p.printer;$("#paperSelect").value=p.paper||"AUTO";$("#orientationSelect").value=p.orientation||"AUTO";$("#scaleSelect").value=p.scale_mode||"FIT";$("#duplexSelect").value=p.duplex||"DEFAULT";$("#colorSelect").value=p.color_mode||"DEFAULT";$("#copiesInput").value=p.copies||1;};
  $("#saveProfileBtn").onclick=()=>{if(state.role!=="admin"){toast("Hanya admin yang dapat menyimpan profile.","err");return;}openTextPrompt("Nama Print Profile","",async(name)=>{if(!name)return;const payload={name,printer:$("#printerSelect").value,paper:$("#paperSelect").value,orientation:$("#orientationSelect").value,scale_mode:$("#scaleSelect").value,duplex:$("#duplexSelect").value,color_mode:$("#colorSelect").value,copies:+$("#copiesInput").value||1};try{await api("/api/profiles",{method:"POST",body:JSON.stringify(payload)});toast("Profile tersimpan.","ok");await loadProfiles();$("#profileSelect").value=name;}catch(e){toast(e.message,"err");}});};
  $("#deleteProfileBtn").onclick=async()=>{if(state.role!=="admin")return;const name=$("#profileSelect").value;if(!name){toast("Pilih profile dahulu.","err");return;}if(!(await uiConfirm(`Hapus profile "${name}"?`,{title:"Hapus Print Profile",okText:"Hapus",danger:true})))return;try{await api(`/api/profiles/${encodeURIComponent(name)}`,{method:"DELETE"});toast("Profile dihapus.","ok");await loadProfiles();}catch(e){toast(e.message,"err");}};

  async function openPrintModal(item) {
    state.selected = item;
    $("#printFileName").textContent = item.name;
    $("#copiesInput").value = 1;
    $("#pagesInput").value = "";
    $("#paperSelect").value = "AUTO";
    $("#orientationSelect").value = "AUTO";
    $("#scaleSelect").value = "FIT";
    $("#duplexSelect").value = "DEFAULT";
    $("#colorSelect").value = "DEFAULT";
    $("#prioritySelect").value = "0"; $("#scheduleInput").value = "";
    await loadPrinterOptions(); await loadProfiles();
    $("#saveProfileBtn").style.display = state.role === "admin" ? "" : "none";
    $("#deleteProfileBtn").style.display = state.role === "admin" ? "" : "none";
    $("#prioritySelect").disabled = state.role !== "admin";

    const quick = $("#quickCopies");
    quick.innerHTML = "";
    [1, 2, 3, 4, 5, 10].forEach((n) => {
      const b = el("button", null, n + "x"); b.type = "button";
      b.onclick = () => { $("#copiesInput").value = n; }; quick.appendChild(b);
    });
    const ext = (item.ext || "").toLowerCase();
    const sheetWrap = $("#sheetPickWrap");
    if ([".xlsx", ".xlsm", ".xls"].includes(ext)) {
      try {
        let sheets=[];
        if(item.source_type==="device") sheets=item.sheets||[];
        else { const info = await api(`/api/file-info?root=${encodeURIComponent(item.root || state.root)}&path=${encodeURIComponent(item.rel_path)}`); sheets=info.sheets||[]; }
        const sel = $("#sheetSelect"); sel.innerHTML = '<option value="">(semua sheet)</option>';
        sheets.forEach((name) => { const o = el("option", null, name); o.value = name; sel.appendChild(o); });
        sheetWrap.style.display = sheets.length ? "block" : "none";
      } catch { sheetWrap.style.display = "none"; }
    } else { sheetWrap.style.display = "none"; }
    $("#printModal").classList.add("open");
  }
  $("#printClose").onclick = () => $("#printModal").classList.remove("open");
  $("#copiesMinus").onclick = () => { const i = $("#copiesInput"); i.value = Math.max(1, (+i.value || 1) - 1); };
  $("#copiesPlus").onclick = () => { const i = $("#copiesInput"); i.value = Math.min(99, (+i.value || 1) + 1); };

  $("#printSubmit").onclick = async () => {
    const item = state.selected; if (!item) return;
    const copies = Math.max(1, Math.min(99, +$("#copiesInput").value || 1));
    const pages = $("#pagesInput").value.trim();
    const sheetSel = $("#sheetSelect");
    const sheet = sheetSel && $("#sheetPickWrap").style.display !== "none" ? sheetSel.value : "";
    const printer = $("#printerSelect").value;
    if (!printer) { toast("Pilih printer lokal terlebih dahulu.", "err"); return; }
    const payload = {
      root: item.root || state.root, path: item.rel_path, copies, pages, sheet, printer,
      paper: $("#paperSelect").value, orientation: $("#orientationSelect").value,
      scale_mode: $("#scaleSelect").value, duplex: $("#duplexSelect").value,
      color_mode: $("#colorSelect").value, priority:+$("#prioritySelect").value||0,
      scheduled_at: $("#scheduleInput").value ? Math.floor(new Date($("#scheduleInput").value).getTime()/1000) : null,
    };
    const submitButton = $("#printSubmit"); showActionProgress("Menambahkan ke antrean printer lokal...", submitButton);
    try {
      let result;
      if(item.source_type==="device") {
        const devicePayload={...payload,token:item.device_token}; delete devicePayload.root; delete devicePayload.path;
        result=await api("/api/device-print/queue",{method:"POST",body:JSON.stringify(devicePayload)});
      } else {
        result=await api("/api/print", { method: "POST", body: JSON.stringify(payload) });
      }
      toast(`Masuk antrean lokal: ${item.name} (Job #${result.id})`, "ok");
      $("#printModal").classList.remove("open"); loadJobs();
    } catch (e) { toast(e.message, "err"); }
    finally { hideActionProgress(submitButton); }
  };

  async function showSpecialList(kind){
    try{const d=await api(kind==="favorites"?"/api/favorites":"/api/recent");state.mode="search";state.searchResults=(d.items||[]).map(x=>({...x,is_dir:false}));state.searchPage=0;state.searchQuery=kind==="favorites"?"Favorit":"Terbaru";$("#breadcrumb").innerHTML=`<span>${kind==="favorites"?"⭐ Favorit":"🕘 Terbaru"}</span>`;renderSearchPage();closeSidebarMobile();}catch(e){toast(e.message,"err");}
  }
  $("#btnFavorites").onclick=()=>showSpecialList("favorites");
  $("#btnRecent").onclick=()=>showSpecialList("recent");

  // --------------------------------------------------------------- jobs --
  function statusBadgeHtml(job) {
    const span = document.createElement("span");
    span.className = `status-pill status-${job.status}`;
    span.textContent = job.status_label;
    return span;
  }

  async function loadJobs(silent = false) {
    try {
      const { jobs } = await api("/api/jobs?limit=30", { silent });
      renderJobsMini(jobs);
      renderJobsPanel(jobs);
    } catch (e) { /* silent - non-critical widget */ }
  }

  function renderJobsMini(jobs) {
    const box = $("#jobsMini");
    box.innerHTML = "";
    jobs.slice(0, 5).forEach((j) => {
      const d = el("div", "job-mini");
      const span = el("span", null, j.label);
      d.appendChild(span);
      d.appendChild(statusBadgeHtml(j));
      box.appendChild(d);
    });
    if (!jobs.length) box.appendChild(el("div", "muted small", "Belum ada job."));
  }

  function renderJobsPanel(jobs) {
    const box = $("#jobsList");
    box.innerHTML = "";
    jobs.forEach((j) => {
      const c = el("div", "job-card");
      const row1 = el("div", "row1");
      row1.appendChild(el("span", "fname", j.label));
      row1.appendChild(statusBadgeHtml(j));
      c.appendChild(row1);
      const methodLabel = `\u{1F5A8}\uFE0F Lokal${j.printer ? ` (${j.printer})` : ""}`;
      c.appendChild(el("div", "meta", `${j.copies}x \u2022 ${j.sender} \u2022 ${methodLabel} \u2022 ${fmtDate(j.created_at)}`));
      const opts = [j.paper || "AUTO", j.orientation || "AUTO", j.scale_mode || "FIT", j.duplex || "DEFAULT", j.color_mode || "DEFAULT"].join(" / ");
      c.appendChild(el("div", "meta", opts));
      if (j.scheduled_at && j.scheduled_at > Date.now()/1000) c.appendChild(el("div","meta",`⏰ Dijadwalkan: ${fmtDate(j.scheduled_at)}`));
      if (j.priority) c.appendChild(el("div","meta",`Prioritas: ${j.priority>=100?"Urgent":j.priority>=50?"High":"Normal"}`));
      if (j.spool_job_id) c.appendChild(el("div", "meta", `Windows Job #${j.spool_job_id}${j.spool_status ? ` \u2022 ${j.spool_status}` : ""}`));
      c.appendChild(el("div", "meta", `File ini sudah berhasil diprint ${j.print_count || 0} kali.`));
      if (j.error) {
        const e = el("div", "meta", j.error);
        e.style.color = "var(--danger)";
        c.appendChild(e);
      }
      const canManage = state.role === "admin" || j.sender === state.username;
      if (canManage) {
        const actions = el("div", "job-actions");
        const cancelable = ["QUEUED", "PROCESSING", "PRINTING"].includes(j.status);
        if (cancelable) {
          const cancel = el("button", "job-action-btn cancel", "Batalkan");
          cancel.type = "button";
          cancel.onclick = async () => {
            if (!(await uiConfirm(`Batalkan antrean print "${j.label}"?`, {title:"Batalkan Print", okText:"Batalkan Job", danger:true}))) return;
            showActionProgress("Membatalkan antrean print...", cancel);
            try {
              await api(`/api/jobs/${encodeURIComponent(j.id)}/cancel`, { method: "POST" });
              toast("Antrean print dibatalkan.", "ok");
              await loadJobs();
            } catch (e) { toast(e.message, "err"); }
            finally { hideActionProgress(cancel); }
          };
          actions.appendChild(cancel);
        }
        const reprint = el("button", "job-action-btn", "Print Ulang");
        reprint.type = "button";
        reprint.onclick = async () => {
          if (!(await uiConfirm(`Print ulang "${j.label}" dengan pengaturan sebelumnya?`, {title:"Print Ulang", okText:"Print Ulang"}))) return;
          showActionProgress("Mengirim print ulang...", reprint);
          try {
            const result = await api(`/api/jobs/${encodeURIComponent(j.id)}/reprint`, { method: "POST" });
            toast(`Print ulang masuk antrean lokal. Job #${result.id}`, "ok");
            await loadJobs();
          } catch (e) { toast(e.message, "err"); }
          finally { hideActionProgress(reprint); }
        };
        actions.appendChild(reprint);
        c.appendChild(actions);
      }
      box.appendChild(c);
    });
    if (!jobs.length) box.appendChild(el("div", "muted", "Belum ada riwayat print."));
  }

  $("#fabJobs").onclick = () => { $("#jobsPanel").classList.add("open"); loadJobs(); };
  $("#jobsClose").onclick = () => $("#jobsPanel").classList.remove("open");

  // ------------------------------------------------------------- stats --
  async function loadStats(silent = false) {
    try {
      const s = await api("/api/stats", { silent });
      renderStatsMini(s);
      renderStatsDetail(s);
    } catch (e) { /* silent */ }
  }

  function renderStatsMini(s) {
    const box = $("#statsMini");
    box.innerHTML = "";
    box.appendChild(el("div", "stats-mini-row",
      `\u2705 ${s.total_success} sukses \u2022 \u274C ${s.total_failed} gagal`));
    box.appendChild(el("div", "stats-mini-row muted small", `${s.total_copies_printed} lembar tercetak`));
  }

  function renderStatsDetail(s) {
    const box = $("#statsContent");
    box.innerHTML = "";

    const summary = el("div", "stats-summary");
    summary.appendChild(statCard("\u2705", s.total_success, "Sukses"));
    summary.appendChild(statCard("\u274C", s.total_failed, "Gagal"));
    summary.appendChild(statCard("\u{1F4C4}", s.total_copies_printed, "Lembar Tercetak"));
    box.appendChild(summary);

    const bm = s.by_method || {};
    box.appendChild(el("div", "stats-section-title", "Metode Print"));
    box.appendChild(el("div", "meta", `\u{1F5A8}\uFE0F Printer Lokal: ${bm.direct || 0}x${(bm.gas || 0) + (bm.bridge || 0) ? ` \u2022 Legacy: ${(bm.gas || 0) + (bm.bridge || 0)}x` : ""}`));

    box.appendChild(el("div", "stats-section-title", "File Paling Sering Diprint"));
    if (s.top_files && s.top_files.length) {
      s.top_files.forEach((f, i) => {
        box.appendChild(el("div", "stats-list-row", `${i + 1}. ${f.name} \u2014 ${f.count}x (${f.copies} lembar)`));
      });
    } else {
      box.appendChild(el("div", "muted small", "Belum ada riwayat print sukses."));
    }

    box.appendChild(el("div", "stats-section-title", "Pengguna Paling Aktif"));
    if (s.top_senders && s.top_senders.length) {
      s.top_senders.forEach((sd, i) => {
        box.appendChild(el("div", "stats-list-row", `${i + 1}. ${sd.sender} \u2014 ${sd.count}x`));
      });
    } else {
      box.appendChild(el("div", "muted small", "Belum ada data."));
    }
  }

  function statCard(icon, value, label) {
    const c = el("div", "stat-card");
    c.appendChild(el("div", "stat-card-icon", icon));
    c.appendChild(el("div", "stat-card-value", String(value)));
    c.appendChild(el("div", "stat-card-label", label));
    return c;
  }

  $("#btnStatsDetail").onclick = () => { $("#statsPanel").classList.add("open"); loadStats(); };
  $("#statsClose").onclick = () => $("#statsPanel").classList.remove("open");

  // ------------------------------------------------ printer/system tools --
  async function loadPrinterTools() {
    if (state.role !== "admin") return;
    try {
      const d = await api("/api/diagnostics");
      const sum = $("#diagnosticSummary"); sum.innerHTML = "";
      const db = d.database || {}; const st = d.storage || {}; const task = d.task_scheduler || {};
      [
        `Python: ${d.python} • ${d.platform}`,
        `Default printer: ${d.default_printer || "(belum diset)"}`,
        `Queue: ${d.queue_paused ? "PAUSED" : "RUNNING"}${d.active_job ? ` • aktif #${d.active_job}` : ""}`,
        `Maintenance: ${d.maintenance_mode ? "AKTIF" : "Normal"} • Multi-printer aktif: ${Object.keys(d.active_jobs_by_printer||{}).length}`,
        `Database: ${db.ok ? "OK" : "ERROR"} (${db.result || "-"}) • Schema v${d.db_schema_version || "-"}`,
        `File Index: ${d.file_index_count || 0} file • Config v${d.config_version || "-"}`,
        `Task Scheduler: ${task.ok ? "OK" : (task.exists ? "Perlu repair" : "Belum ada")}`,
        `Data: DB ${fmtSize(st.database || 0)} • Archive ${fmtSize(st.archive || 0)} • Backup ${fmtSize(st.backup || 0)}`,
      ].forEach((txt) => sum.appendChild(el("div", "stats-list-row", txt)));

      const box = $("#printerToolsList"); box.innerHTML = "";
      (d.printers || []).forEach((p) => {
        const c = el("div", "job-card"); const r = el("div", "row1");
        r.appendChild(el("span", "fname", p.printer + (p.default ? " (Default)" : "")));
        r.appendChild(el("span", `status-pill ${p.problem ? "status-FAILED" : "status-SUCCESS"}`, p.message || "Unknown"));
        c.appendChild(r); c.appendChild(el("div", "meta", `${p.jobs || 0} job di Windows spooler`));
        const a = el("div", "job-actions");
        const setDef = el("button", "job-action-btn", "Jadikan Default"); setDef.type = "button";
        setDef.onclick = async () => { try { await api("/api/printers/default", {method:"POST", body:JSON.stringify({printer:p.printer})}); toast("Default printer diperbarui.","ok"); loadPrinterTools(); loadStatus(); } catch(e){toast(e.message,"err");} };
        const test = el("button", "job-action-btn", "Test Print"); test.type = "button";
        test.onclick = async () => { showActionProgress(`Test print ke ${p.printer}...`, test); try { const x=await api("/api/printers/test", {method:"POST", body:JSON.stringify({printer:p.printer})}); toast(`Test print dikirim${x.spool_job_id ? ` (Windows Job #${x.spool_job_id})` : ""}.`,"ok"); } catch(e){toast(e.message,"err");} finally {hideActionProgress(test);} };
        a.append(setDef, test); c.appendChild(a); box.appendChild(c);
      });
      if (!(d.printers || []).length) box.appendChild(el("div", "muted", "Tidak ada printer lokal terdeteksi."));
    } catch(e) { toast(e.message, "err"); }
  }

  $("#btnPrinterTools").onclick = () => { $("#printerToolsPanel").classList.add("open"); loadPrinterTools(); };
  $("#printerToolsClose").onclick = () => $("#printerToolsPanel").classList.remove("open");
  $("#queuePauseBtn").onclick = async () => { try { await api("/api/queue/pause", {method:"POST"}); toast("Queue dipause.","ok"); loadPrinterTools(); } catch(e){toast(e.message,"err");} };
  $("#queueResumeBtn").onclick = async () => { try { await api("/api/queue/resume", {method:"POST"}); toast("Queue dilanjutkan.","ok"); loadPrinterTools(); } catch(e){toast(e.message,"err");} };
  $("#taskRepairBtn").onclick = async () => { showActionProgress("Memperbaiki Task Scheduler...", $("#taskRepairBtn")); try { await api("/api/task/repair", {method:"POST"}); toast("Task Scheduler siap.","ok"); loadPrinterTools(); } catch(e){toast(e.message,"err");} finally {hideActionProgress($("#taskRepairBtn"));} };
  $("#restartBtn").onclick=async()=>{if(!(await uiConfirm("Restart PrintBot WebApp sekarang?",{title:"Restart WebApp",okText:"Restart",danger:true})))return;try{await api("/api/restart",{method:"POST"});toast("Restart dijadwalkan...","ok");}catch(e){toast(e.message,"err");}};
  $("#backupNowBtn").onclick = async () => { showActionProgress("Membuat backup...", $("#backupNowBtn")); try { const d=await api("/api/backup", {method:"POST"}); toast(`Backup dibuat: ${d.name}`,"ok"); loadPrinterTools(); } catch(e){toast(e.message,"err");} finally {hideActionProgress($("#backupNowBtn"));} };
  $("#cleanupNowBtn").onclick = async () => { try { const d=await api("/api/maintenance/cleanup", {method:"POST"}); toast(`Cleanup: ${d.previews_removed} preview, ${d.backups_removed} backup lama.`,"ok"); loadPrinterTools(); } catch(e){toast(e.message,"err");} };

  // ---------------------------------------------------- role & user mgmt --
  async function loadWhoami() {
    try {
      const data = await api("/api/whoami");
      state.role = data.role || "user";
      state.username = data.username || "";
      state.csrf = data.csrf || "";
      state.maintenance = !!data.maintenance;
    } catch (e) {
      state.role = "user";
    }
    applyRoleUI();
  }

  function applyRoleUI() {
    const isAdmin = state.role === "admin";
    $("#btnAddRoot").style.display = isAdmin ? "" : "none";
    $("#btnNewFolder").style.display = isAdmin ? "" : "none";
    $("#btnUpload").style.display = isAdmin ? "" : "none";
    $("#adminSectionTitle").style.display = isAdmin ? "" : "none";
    $("#btnManageUsers").style.display = isAdmin ? "" : "none";
    $("#btnPrinterTools").style.display = isAdmin ? "" : "none";
  }

  async function loadUsers() {
    if (state.role !== "admin") return;
    try {
      const { users } = await api("/api/users");
      renderUsersList(users);
    } catch (e) { toast(e.message, "err"); }
  }

  function renderUsersList(users) {
    const box = $("#usersList");
    box.innerHTML = "";
    users.forEach((u) => {
      const c = el("div", "job-card");
      const row1 = el("div", "row1");
      const label = u.username + (u.username === state.username ? " (Anda)" : "");
      row1.appendChild(el("span", "fname", label));
      const badgeCls = u.role === "admin" ? "status-SUCCESS" : "status-QUEUED";
      row1.appendChild(el("span", `status-pill ${badgeCls}`, u.role === "admin" ? "Admin" : "User"));
      c.appendChild(row1);
      if (!u.primary && u.username !== state.username) {
        const delBtn = el("button", "btn btn-danger small-btn", "Hapus User");
        delBtn.type = "button";
        delBtn.style.marginTop = ".4rem";
        delBtn.onclick = async () => {
          if (!(await uiConfirm(`Hapus user "${u.username}"?`, {title:"Hapus User", okText:"Hapus", danger:true}))) return;
          try {
            await api("/api/users/remove", { method: "POST", body: JSON.stringify({ username: u.username }) });
            toast(`User "${u.username}" dihapus.`, "ok");
            loadUsers();
          } catch (e) { toast(e.message, "err"); }
        };
        c.appendChild(delBtn);
      } else {
        c.appendChild(el("div", "meta", u.primary ? "Akun admin utama (.env)" : "Akun Anda saat ini"));
      }
      box.appendChild(c);
    });
    if (!users.length) box.appendChild(el("div", "muted", "Belum ada user."));
  }

  $("#btnManageUsers").onclick = () => { $("#usersPanel").classList.add("open"); loadUsers(); };
  $("#usersClose").onclick = () => $("#usersPanel").classList.remove("open");
  $("#btnAddUser").onclick = () => {
    $("#newUserUsername").value = "";
    $("#newUserPassword").value = "";
    $("#newUserRole").value = "user";
    $("#addUserModal").classList.add("open");
  };
  $("#addUserClose").onclick = () => $("#addUserModal").classList.remove("open");
  $("#newUserSubmit").onclick = async () => {
    const username = $("#newUserUsername").value.trim();
    const password = $("#newUserPassword").value;
    const role = $("#newUserRole").value;
    if (!username || !password) { toast("Username & password wajib diisi.", "err"); return; }
    try {
      await api("/api/users/add", { method: "POST", body: JSON.stringify({ username, password, role }) });
      toast(`User "${username}" ditambahkan.`, "ok");
      $("#addUserModal").classList.remove("open");
      loadUsers();
    } catch (e) { toast(e.message, "err"); }
  };

  // -------------------------------------------------------------- status --
  async function loadStatus(silent = false) {
    try {
      const data = await api("/api/status", { silent });
      state.localPrintAvailable = !!data.local_print_available;
      const badge = $("#bridgeBadge");
      const ps = data.printer_system || {};
      const st = ps.default_status || {};
      const flags = Array.isArray(st.flags) ? st.flags : [];
      let dot = "⚪", kind = "muted", stateText = "Tidak diketahui";
      if (!ps.available) { dot = "🔴"; kind = "danger"; stateText = "Printer tidak tersedia"; }
      else if (st.problem || flags.some(x => ["OFFLINE","ERROR","PAPER_JAM","PAPER_OUT","PAPER_PROBLEM","NO_TONER","USER_INTERVENTION","DOOR_OPEN","NOT_AVAILABLE"].includes(x))) {
        dot = "🔴"; kind = "danger"; stateText = st.message || "Printer bermasalah";
      } else if (flags.includes("PRINTING")) { dot = "🔵"; kind = "ok"; stateText = "Sedang mencetak"; }
      else if (flags.some(x => ["TONER_LOW","OUTPUT_BIN_FULL"].includes(x))) { dot = "🟠"; kind = "muted"; stateText = st.message || "Perlu perhatian"; }
      else if (flags.includes("POWER_SAVE")) { dot = "🟡"; kind = "muted"; stateText = "Power save"; }
      else { dot = "🟢"; kind = "ok"; stateText = "Online / Ready"; }
      const printerName = ps.default_printer || "Printer lokal";
      badge.innerHTML = `<span class="printer-dot">${dot}</span><span class="printer-status-text">Printer: ${printerName}</span>`;
      badge.className = `badge badge-${kind}`;
      badge.title = `${printerName} — ${stateText}`;
      badge.setAttribute("aria-label", `${printerName}: ${stateText}`);
    } catch (e) { /* silent */ }
  }

  // -------------------------------------------------------------- sort --
  $("#sortSelect").onchange = (e) => {
    state.sort = e.target.value;
    if (state.mode === "search") {
      state.searchResults.sort((a, b) => {
        if (state.sort === "size") return b.size - a.size;
        if (state.sort === "modified") return b.modified - a.modified;
        if (state.sort === "type") return (a.ext || "").localeCompare(b.ext || "") || a.name.localeCompare(b.name);
        return a.name.localeCompare(b.name);
      });
      state.searchPage = 0;
      renderSearchPage();
    } else {
      state.page = 0;
      loadBrowse();
    }
  };

  // ------------------------------------------------------------ mobile --
  function closeSidebarMobile() {
    $("#sidebar").classList.remove("open");
    $("#drawerOverlay").classList.remove("open");
  }
  $("#btnMenu").onclick = () => {
    $("#sidebar").classList.toggle("open");
    $("#drawerOverlay").classList.toggle("open");
  };
  $("#drawerOverlay").onclick = closeSidebarMobile;

  // ------------------------------------------------------ realtime/PWA --
  function initRealtime(){
    try{
      const es=new EventSource("/api/events");
      es.onmessage=(ev)=>{try{const d=JSON.parse(ev.data||"{}");if(d.type==="job"){loadJobs(true);loadStatus(true);loadStats(true);}if(d.type==="printer"){loadStatus(true);if(state.role==="admin"){const msg=`${d.printer||"Printer"}: ${d.message||d.action||"status berubah"}`;toast(msg,d.problem?"err":"ok");if(Notification.permission==="granted")new Notification("PrintBot Printer",{body:msg});}}if(d.type==="update"&&state.role==="admin")toast(`Update: ${d.action||"tersedia"} ${d.version||d.new_version||""}`,"ok");if(d.type==="system"&&d.action==="maintenance"){state.maintenance=!!d.enabled;toast(`Maintenance ${d.enabled?"aktif":"nonaktif"}.`,d.enabled?"err":"ok");}}catch(e){}};
      es.onerror=()=>{};
    }catch(e){}
  }
  if("serviceWorker" in navigator){window.addEventListener("load",()=>navigator.serviceWorker.register("/sw.js").catch(()=>{}));}

  async function renderAdvanced(kind){
    const out=$("#advancedAdminOutput");out.textContent="Memuat...";
    try{
      if(kind==="backup"){
        const d=await api("/api/backups");out.innerHTML="";(d.backups||[]).forEach(b=>{const c=el("div","job-card");c.appendChild(el("div","fname",b.name));c.appendChild(el("div","meta",`${fmtSize(b.size)} • ${fmtDate(b.modified)}`));const a=el("a","btn small-btn","Download");a.href=`/api/backups/${encodeURIComponent(b.name)}/download`;a.target="_blank";const r=el("button","btn btn-danger small-btn","Restore");r.onclick=async()=>{if(!(await uiConfirm(`Restore backup "${b.name}"?
Data saat ini akan dibackup otomatis sebelum restore.`,{title:"Restore Backup",okText:"Restore",danger:true})))return;try{const x=await api(`/api/backups/${encodeURIComponent(b.name)}/restore`,{method:"POST"});toast(`Restore selesai. Safety: ${x.safety_backup}. Restart WebApp diperlukan.`,"ok");}catch(e){toast(e.message,"err");}};const acts=el("div","sheet-actions");acts.append(a,r);c.appendChild(acts);out.appendChild(c);});if(!(d.backups||[]).length)out.textContent="Belum ada backup.";
      }else if(kind==="audit"){
        const [a,e]=await Promise.all([api("/api/audit?limit=80"),api("/api/errors?limit=80")]);out.innerHTML="<b>Audit terbaru</b>";(a.events||[]).slice(0,30).forEach(x=>out.appendChild(el("div","meta",`${fmtDate(x.ts)} • ${x.username||"system"} • ${x.action} • ${x.target||""}`)));out.appendChild(el("div","sidebar-title","Error terbaru"));(e.errors||[]).slice(0,30).forEach(x=>{const d=el("div","meta",`${fmtDate(x.ts)} • ${x.category} • ${x.message}`);d.style.color="var(--danger)";out.appendChild(d);});
      }else if(kind==="storage"){
        const d=await api("/api/storage");out.innerHTML="<b>Storage</b>";["database","archive","previews","logs","backup"].forEach(k=>out.appendChild(el("div","meta",`${k}: ${fmtSize(d[k]||0)}`)));out.appendChild(el("div","meta",`Disk free: ${fmtSize((d.disk||{}).free||0)}`));["previews","uploads","exports","archive","logs"].forEach(cat=>{const b=el("button","btn small-btn",`Cleanup ${cat}`);b.onclick=async()=>{if(!(await uiConfirm(`Hapus file ${cat} yang lebih lama dari 7 hari?`,{title:"Cleanup Storage",okText:"Hapus",danger:true})))return;try{const x=await api("/api/storage/cleanup",{method:"POST",body:JSON.stringify({category:cat,max_age_days:7})});toast(`${x.removed} file dihapus (${fmtSize(x.bytes_removed)}).`,"ok");renderAdvanced("storage");}catch(e){toast(e.message,"err");}};out.appendChild(b);});
      }
    }catch(e){out.textContent=e.message;}
  }
  $("#maintenanceBtn").onclick=async()=>{try{const cur=await api("/api/maintenance");const d=await api("/api/maintenance",{method:"POST",body:JSON.stringify({enabled:!cur.enabled})});toast(`Maintenance ${d.enabled?"aktif":"nonaktif"}.`,d.enabled?"err":"ok");loadPrinterTools();}catch(e){toast(e.message,"err");}};
  $("#reindexBtn").onclick=async()=>{showActionProgress("Membangun index file...",$("#reindexBtn"));try{const d=await api("/api/index/rebuild",{method:"POST",body:"{}"});toast(`Index selesai: ${d.indexed} file.`,"ok");}catch(e){toast(e.message,"err");}finally{hideActionProgress($("#reindexBtn"));}};
  $("#dbMaintBtn").onclick=async()=>{showActionProgress("Maintenance database...",$("#dbMaintBtn"));try{const d=await api("/api/db/maintenance",{method:"POST",body:JSON.stringify({vacuum:true})});toast(`DB: ${d.integrity}`,d.ok?"ok":"err");}catch(e){toast(e.message,"err");}finally{hideActionProgress($("#dbMaintBtn"));}};
  $("#backupManagerBtn").onclick=()=>renderAdvanced("backup");
  $("#auditBtn").onclick=()=>renderAdvanced("audit");
  $("#storageBtn").onclick=()=>renderAdvanced("storage");
  $("#apiTokenBtn").onclick=async()=>{try{const d=await api("/api/api-token");$("#advancedAdminOutput").innerHTML=`<b>API Token</b><div class="meta" style="word-break:break-all">${d.token}</div><div class="meta">Scopes: ${(d.scopes||[]).join(", ")}</div>`;if(navigator.clipboard)await navigator.clipboard.writeText(d.token);toast("API token ditampilkan dan disalin ke clipboard.","ok");}catch(e){toast(e.message,"err");}};
  $("#exportXlsxBtn").onclick=()=>window.open("/api/export?fmt=xlsx","_blank");
  $("#exportCsvBtn").onclick=()=>window.open("/api/export?fmt=csv","_blank");
  $("#notifBtn").onclick=async()=>{if(!("Notification" in window)){toast("Browser tidak mendukung notifikasi.","err");return;}const p=await Notification.requestPermission();toast(`Notifikasi: ${p}`,p==="granted"?"ok":"err");};
  $("#updateCheckBtn").onclick=async()=>{try{const d=await api("/api/update/check");toast(d.configured?(d.available?`Update ${d.version} tersedia.`:"Sudah versi terbaru."):"UPDATE_URL belum diset.",d.available?"ok":"err");$("#advancedAdminOutput").textContent=JSON.stringify(d,null,2);}catch(e){toast(e.message,"err");}};
  $("#updateApplyBtn").onclick=async()=>{if(!(await uiConfirm("Terapkan update dari UPDATE_URL?",{title:"Update PrintBot",okText:"Terapkan Update"})))return;try{const d=await api("/api/update/apply",{method:"POST"});toast(d.updated?"Update diterapkan. Restart diperlukan.":(d.reason||"Tidak ada update."),"ok");}catch(e){toast(e.message,"err");}};
  $("#updateRollbackBtn").onclick=async()=>{if(!(await uiConfirm("Rollback ke versi sebelum update terakhir?",{title:"Rollback PrintBot",okText:"Rollback",danger:true})))return;try{const d=await api("/api/update/rollback",{method:"POST"});toast(`Rollback ke ${d.to_version||"versi lama"}. Restart diperlukan.`,"ok");}catch(e){toast(e.message,"err");}};

  document.addEventListener("keydown", (e) => {
    if (e.key !== "Escape") return;
    if ($("#confirmModal").classList.contains("open")) { closeConfirm(false); return; }
    if ($("#themeModal").classList.contains("open")) { $("#themeModal").classList.remove("open"); return; }
  });

  // -------------------------------------------------------------- init --
  (async () => {
    await loadConfig();
    await loadWhoami();
    initRealtime();
    await loadRoots();
    loadJobs();
    loadStatus();
    loadStats();
  })();
  setInterval(() => loadJobs(true), 15000);
  setInterval(() => loadStatus(true), 30000);
  setInterval(() => loadStats(true), 45000);
})();

</script>
</body>
</html>
"""

# ============================================================================
# PWA ROUTES
# ============================================================================
@app.get("/manifest.webmanifest")
async def pwa_manifest():
    return JSONResponse({"name":"PrintBot WebApp","short_name":"PrintBot","start_url":"/","display":"standalone","background_color":"#0f172a","theme_color":"#1e293b","icons":[{"src":"/pwa-icon-192.png","sizes":"192x192","type":"image/png"},{"src":"/pwa-icon-512.png","sizes":"512x512","type":"image/png","purpose":"any maskable"}]},media_type="application/manifest+json")

def _pwa_icon_bytes(size: int) -> bytes:
    from PIL import Image,ImageDraw,ImageFont
    img=Image.new("RGB",(size,size),(30,41,59));d=ImageDraw.Draw(img)
    try: f=ImageFont.truetype("arial.ttf",int(size*0.42))
    except Exception: f=ImageFont.load_default()
    text="P"; box=d.textbbox((0,0),text,font=f); w=box[2]-box[0]; h=box[3]-box[1]; d.rounded_rectangle((size*.12,size*.12,size*.88,size*.88),radius=size*.15,outline=(59,130,246),width=max(4,int(size*.04))); d.text(((size-w)/2,(size-h)/2-size*.03),text,fill=(229,231,235),font=f)
    buf=io.BytesIO(); img.save(buf,"PNG"); return buf.getvalue()

@app.get("/pwa-icon-{size}.png")
async def pwa_icon(size: int):
    if size not in (192,512): raise HTTPException(status_code=404)
    return Response(_pwa_icon_bytes(size),media_type="image/png",headers={"Cache-Control":"public,max-age=86400"})

@app.get("/sw.js")
async def pwa_service_worker():
    code="""self.addEventListener('install',e=>self.skipWaiting());self.addEventListener('activate',e=>e.waitUntil(self.clients.claim()));self.addEventListener('fetch',e=>{if(e.request.method==='GET'&&!e.request.url.includes('/api/'))e.respondWith(fetch(e.request).catch(()=>new Response('Offline',{status:503})));});"""
    return Response(code,media_type="application/javascript",headers={"Cache-Control":"no-cache"})


# ============================================================================
# PAGE ROUTES
# ============================================================================
@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if request.session.get("user"):
        return RedirectResponse("/")
    return LOGIN_HTML.replace("<!--ERROR-->", "")


@app.post("/login")
async def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    remaining=login_rate_check(request,username)
    if remaining:
        audit_event(username,"LOGIN_BLOCKED",detail="lockout %ss"%remaining,success=False,ip=_client_ip(request))
        html_=LOGIN_HTML.replace("<!--ERROR-->", '<p class="err">Terlalu banyak percobaan. Coba lagi beberapa menit.</p>')
        return HTMLResponse(html_,status_code=429)
    role = verify_login(username, password)
    if role:
        login_rate_success(request,username); now=time.time(); request.session.clear()
        request.session["user"] = username; request.session["role"] = role; request.session["created_at"]=now; request.session["last_seen"]=now; request.session["csrf"]=secrets.token_urlsafe(32)
        audit_event(username,"LOGIN",detail="role=%s"%role,ip=_client_ip(request))
        return RedirectResponse("/", status_code=303)
    login_rate_fail(request,username); audit_event(username,"LOGIN_FAILED",success=False,ip=_client_ip(request))
    html_ = LOGIN_HTML.replace("<!--ERROR-->", '<p class="err">Username atau password salah.</p>')
    return HTMLResponse(html_, status_code=401)


@app.get("/logout")
async def logout(request: Request):
    user=str(request.session.get("user") or ""); audit_event(user,"LOGOUT",ip=_client_ip(request)); request.session.clear()
    return RedirectResponse("/login")


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login")
    return INDEX_HTML


# ============================================================================
# REST API
# ============================================================================
@app.get("/api/health")
async def api_health():
    return {"status": "ok", "version": APP_VERSION, "printer_local": local_print_available(),
            "queue_paused": _PRINT_QUEUE_PAUSED, "active_job": _PRINT_WORKER_ACTIVE_JOB}


@app.get("/api/whoami")
async def api_whoami(request: Request, user: str = Depends(require_login)):
    return {"username": user, "role": request.session.get("role", ROLE_USER), "csrf": _ensure_csrf(request), "maintenance": is_maintenance_mode()}


@app.get("/api/events")
async def api_events(user: str = Depends(require_login)):
    return StreamingResponse(_event_stream(),media_type="text/event-stream",headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


@app.get("/api/quota")
async def api_quota(user: str = Depends(require_login)):
    return print_quota_status(user)


@app.get("/api/maintenance")
async def api_maintenance_get(user: str = Depends(require_login)):
    return {"enabled":is_maintenance_mode(),"queue_paused":_PRINT_QUEUE_PAUSED}


@app.post("/api/maintenance")
async def api_maintenance_set(payload: dict, request: Request, user: str = Depends(require_admin)):
    enabled=bool(payload.get("enabled")); set_maintenance_mode(enabled); audit_event(user,"MAINTENANCE_ON" if enabled else "MAINTENANCE_OFF",ip=_client_ip(request)); await event_publish({"type":"system","action":"maintenance","enabled":enabled}); return {"enabled":enabled}


@app.get("/api/storage")
async def api_storage(user: str = Depends(require_admin)):
    return await _to_thread(storage_status)


@app.post("/api/storage/cleanup")
async def api_storage_cleanup(payload: dict, request: Request, user: str = Depends(require_admin)):
    category=str(payload.get("category") or ""); days=int(payload.get("max_age_days") or 7)
    try: result=await _to_thread(storage_cleanup,category,days)
    except Exception as exc: raise HTTPException(status_code=400,detail=str(exc))
    audit_event(user,"STORAGE_CLEANUP",target=category,detail=json.dumps(result),ip=_client_ip(request)); return result


@app.get("/api/audit")
async def api_audit(limit: int = 200, user: str = Depends(require_admin)):
    return {"events":await _to_thread(audit_list,limit)}


@app.get("/api/errors")
async def api_errors(limit: int = 200, user: str = Depends(require_admin)):
    return {"errors":await _to_thread(error_list,limit)}


@app.get("/api/backups")
async def api_backups(user: str = Depends(require_admin)):
    return {"backups":await _to_thread(backup_list)}


@app.get("/api/backups/{name}/download")
async def api_backup_download(name: str, user: str = Depends(require_admin)):
    try: p=await _to_thread(_safe_backup_path,name)
    except Exception as exc: raise HTTPException(status_code=404,detail=str(exc))
    return FileResponse(p,filename=p.name,media_type="application/zip")


@app.post("/api/backups/{name}/restore")
async def api_backup_restore(name: str, request: Request, user: str = Depends(require_admin)):
    global _PRINT_QUEUE_PAUSED
    if _ACTIVE_JOB_BY_PRINTER:
        raise HTTPException(status_code=409,detail="Tidak bisa restore saat masih ada job print aktif.")
    _PRINT_QUEUE_PAUSED=True; app_setting_set("queue_paused","1"); set_maintenance_mode(True)
    try: result=await _to_thread(restore_backup_archive,name)
    except Exception as exc: record_error("RESTORE",str(exc)); audit_event(user,"RESTORE_FAILED",target=name,detail=str(exc),success=False,ip=_client_ip(request)); raise HTTPException(status_code=500,detail=str(exc))
    audit_event(user,"RESTORE",target=name,detail=json.dumps(result),ip=_client_ip(request)); return result


@app.post("/api/db/maintenance")
async def api_db_maintenance(payload: dict, request: Request, user: str = Depends(require_admin)):
    result=await _to_thread(db_full_maintenance,bool(payload.get("vacuum",True))); audit_event(user,"DB_MAINTENANCE",detail=json.dumps(result),ip=_client_ip(request)); return result


@app.get("/api/export")
async def api_export(fmt: str = "xlsx", user: str = Depends(require_admin)):
    if fmt.lower() not in {"xlsx","csv"}: raise HTTPException(status_code=400,detail="Format harus xlsx/csv.")
    p=await _to_thread(export_jobs_report,fmt.lower()); return FileResponse(p,filename=p.name,background=BackgroundTask(lambda: p.unlink(missing_ok=True)))


@app.get("/api/update/check")
async def api_update_check(user: str = Depends(require_admin)):
    try: return await _to_thread(check_update_available)
    except Exception as exc: raise HTTPException(status_code=502,detail=str(exc))


@app.post("/api/update/apply")
async def api_update_apply(request: Request, user: str = Depends(require_admin)):
    try: result=await _to_thread(apply_update_from_url)
    except Exception as exc: record_error("UPDATE",str(exc)); raise HTTPException(status_code=500,detail=str(exc))
    audit_event(user,"UPDATE_APPLY",detail=json.dumps(result),ip=_client_ip(request)); return result


@app.post("/api/update/rollback")
async def api_update_rollback(request: Request, user: str = Depends(require_admin)):
    try: result=await _to_thread(rollback_last_update)
    except Exception as exc: raise HTTPException(status_code=500,detail=str(exc))
    audit_event(user,"UPDATE_ROLLBACK",detail=json.dumps(result),ip=_client_ip(request)); return result


@app.post("/api/index/rebuild")
async def api_index_rebuild(payload: dict, request: Request, user: str = Depends(require_admin)):
    result=await _to_thread(rebuild_file_index,payload.get("root") or None); audit_event(user,"INDEX_REBUILD",detail=json.dumps(result),ip=_client_ip(request)); return result


@app.get("/api/index/status")
async def api_index_status(user: str = Depends(require_admin)):
    with _db_lock, _db_connect() as conn: count=conn.execute("SELECT COUNT(*) c FROM file_index").fetchone()["c"]
    return {"count":count,"last_build":float(app_setting_get("file_index_last_build","0") or 0)}


@app.get("/api/favorites")
async def api_favorites(user: str = Depends(require_login)):
    return {"items":favorite_list(user)}


@app.post("/api/favorites/toggle")
async def api_favorite_toggle(payload: dict, user: str = Depends(require_login)):
    root=str(payload.get("root") or ""); path=str(payload.get("path") or ""); fm_resolve(root,path); return {"favorite":favorite_toggle(user,root,path)}


@app.get("/api/recent")
async def api_recent(user: str = Depends(require_login)):
    return {"items":recent_list(user)}


@app.get("/api/profiles")
async def api_profiles(user: str = Depends(require_login)):
    return {"profiles":printer_profile_list()}


@app.post("/api/profiles")
async def api_profile_save(payload: dict, request: Request, user: str = Depends(require_admin)):
    try: printer_profile_save(str(payload.get("name") or ""),payload)
    except Exception as exc: raise HTTPException(status_code=400,detail=str(exc))
    audit_event(user,"PROFILE_SAVE",target=str(payload.get("name") or ""),ip=_client_ip(request)); return {"success":True}


@app.delete("/api/profiles/{name}")
async def api_profile_delete(name: str, request: Request, user: str = Depends(require_admin)):
    ok=printer_profile_remove(name); audit_event(user,"PROFILE_DELETE",target=name,success=ok,ip=_client_ip(request)); return {"success":ok}


@app.get("/api/preview/meta")
async def api_preview_meta(root: str, path: str, user: str = Depends(require_login)):
    try: p=fm_resolve(root,path)
    except FileManagerError as exc: raise HTTPException(status_code=400,detail=str(exc))
    if not p.is_file(): raise HTTPException(status_code=404,detail="File tidak ditemukan.")
    recent_touch(user,root,path,"preview"); return await _to_thread(preview_document_meta,p)


@app.get("/api/preview/page")
async def api_preview_page(root: str, path: str, page: int = 1, user: str = Depends(require_login)):
    try: p=fm_resolve(root,path)
    except FileManagerError as exc: raise HTTPException(status_code=400,detail=str(exc))
    if not p.is_file(): raise HTTPException(status_code=404,detail="File tidak ditemukan.")
    img=await _to_thread(generate_preview_page,p,max(1,page))
    if not img or not img.is_file(): raise HTTPException(status_code=422,detail="Halaman preview tidak tersedia.")
    return FileResponse(img,media_type="image/jpeg",headers={"Cache-Control":"no-store"},background=BackgroundTask(lambda: img.unlink(missing_ok=True)))


@app.get("/api/api-token")
async def api_token_info(user: str = Depends(require_admin)):
    return {"token":settings.api_token,"scopes":settings.api_scopes}

@app.post("/api/api-token/regenerate")
async def api_token_regenerate(request: Request, user: str = Depends(require_admin)):
    settings.api_token=secrets.token_urlsafe(36); (settings.data_dir/"api_token.key").write_text(settings.api_token,encoding="utf-8")
    audit_event(user,"API_TOKEN_REGENERATE",ip=_client_ip(request)); return {"token":settings.api_token,"scopes":settings.api_scopes}


@app.get("/api/external/status")
async def api_external_status(request: Request):
    require_api_scope(request,"read")
    return {"version":APP_VERSION,"counts":job_counts(),"maintenance":is_maintenance_mode(),"queue_paused":_PRINT_QUEUE_PAUSED,"default_printer":local_default_printer(),"printers":await _to_thread(local_printer_details)}


@app.get("/api/external/search")
async def api_external_search(request: Request, q: str, root: Optional[str] = None):
    require_api_scope(request,"read")
    try: rows=await _to_thread(indexed_search,root,q)
    except FileManagerError as exc: raise HTTPException(status_code=400,detail=str(exc))
    return {"query":q,"results":rows,"count":len(rows)}


@app.post("/api/external/print")
async def api_external_print(payload: PrintRequest, request: Request):
    require_api_scope(request,"print")
    if is_maintenance_mode(): raise HTTPException(status_code=503,detail="Maintenance Mode aktif.")
    try: enforce_print_quota("api",payload.copies,False)
    except FileManagerError as exc: raise HTTPException(status_code=429,detail=str(exc))
    if payload.scheduled_at is not None and (payload.scheduled_at < time.time()-30 or payload.scheduled_at > time.time()+366*86400):
        raise HTTPException(status_code=400,detail="Jadwal print tidak valid.")
    if (payload.paper or "AUTO").upper() not in {"AUTO","A4","F4","LEGAL","LETTER"}: raise HTTPException(status_code=400,detail="Paper tidak valid.")
    if (payload.orientation or "AUTO").upper() not in {"AUTO","PORTRAIT","LANDSCAPE"}: raise HTTPException(status_code=400,detail="Orientasi tidak valid.")
    if (payload.scale_mode or "FIT").upper() not in {"FIT","ACTUAL"}: raise HTTPException(status_code=400,detail="Scale tidak valid.")
    if (payload.duplex or "DEFAULT").upper() not in {"DEFAULT","OFF","LONG","SHORT"}: raise HTTPException(status_code=400,detail="Duplex tidak valid.")
    if (payload.color_mode or "DEFAULT").upper() not in {"DEFAULT","COLOR","MONO"}: raise HTTPException(status_code=400,detail="Color mode tidak valid.")
    try: _parse_page_numbers(payload.pages)
    except ValueError as exc: raise HTTPException(status_code=400,detail=str(exc))
    try: abs_path=fm_resolve(payload.root,payload.path)
    except FileManagerError as exc: raise HTTPException(status_code=400,detail=str(exc))
    if not abs_path.is_file() or abs_path.suffix.lower() not in PRINTABLE_EXTS: raise HTTPException(status_code=404,detail="File tidak ditemukan/tidak printable.")
    printer=(payload.printer or local_default_printer()).strip()
    if not printer: raise HTTPException(status_code=400,detail="Printer belum ditentukan.")
    printers=await _to_thread(local_list_printers)
    if printers and printer not in printers: raise HTTPException(status_code=404,detail="Printer tidak ditemukan.")
    pstat=await _to_thread(printer_status_details,printer)
    if pstat.get("problem"): raise HTTPException(status_code=409,detail="Printer belum siap: %s"%pstat.get("message","Unknown"))
    try: archive=await _to_thread(archive_print_source,abs_path)
    except OSError as exc: raise HTTPException(status_code=500,detail=str(exc))
    job=job_create(source="api",sender="api",sender_ref="api-token",file_path=str(abs_path),file_name=abs_path.name,label=abs_path.name,
                   copies=payload.copies,pages=payload.pages,sheet=payload.sheet,printer=printer,archive_path=str(archive),paper=payload.paper,
                   orientation=payload.orientation,scale_mode=payload.scale_mode,duplex=payload.duplex,color_mode=payload.color_mode,priority=payload.priority,scheduled_at=payload.scheduled_at)
    audit_event("api","PRINT_QUEUE",target=job["id"],detail=abs_path.name,ip=_client_ip(request)); await event_publish({"type":"job","action":"queued","job":job}); return job


@app.get("/api/users")
async def api_users(user: str = Depends(require_admin)):
    return {"users": user_list()}


class UserAddRequest(BaseModel):
    username: str
    password: str
    role: str = ROLE_USER


class UserRemoveRequest(BaseModel):
    username: str


@app.post("/api/users/add")
async def api_users_add(payload: UserAddRequest, request: Request, user: str = Depends(require_admin)):
    try:
        info = user_add(payload.username, payload.password, payload.role)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    audit_event(user,"USER_ADD",target=info["username"],detail="role=%s"%info["role"],ip=_client_ip(request))
    return {"status": "success", **info}


@app.post("/api/users/remove")
async def api_users_remove(payload: UserRemoveRequest, request: Request, user: str = Depends(require_admin)):
    if constant_time_eq(payload.username, user):
        raise HTTPException(status_code=400, detail="Tidak bisa menghapus akun yang sedang login.")
    ok = user_remove(payload.username)
    if not ok:
        raise HTTPException(status_code=400, detail="User tidak ditemukan (atau itu akun admin utama dari .env).")
    audit_event(user,"USER_REMOVE",target=payload.username,ip=_client_ip(request))
    return {"status": "success"}


@app.get("/api/roots")
async def api_roots(user: str = Depends(require_login)):
    roots = get_roots()
    return {"roots": [{"label": k, "path": v, "dynamic": root_is_dynamic(k)} for k, v in roots.items()]}


@app.post("/api/roots/add")
async def api_roots_add(payload: RootAddRequest, request: Request, user: str = Depends(require_admin)):
    try:
        info = root_add(payload.label, payload.path)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    log.info("ROOT ADD (web) label=%s path=%s user=%s", info["label"], info["path"], user)
    audit_event(user,"ROOT_ADD",target=info["label"],detail=info["path"],ip=_client_ip(request)); asyncio.create_task(_to_thread(rebuild_file_index,info["label"])); asyncio.create_task(_to_thread(restart_file_watchers))
    return {"status": "success", **info}


@app.post("/api/roots/remove")
async def api_roots_remove(payload: RootRemoveRequest, request: Request, user: str = Depends(require_admin)):
    ok = root_remove(payload.label)
    if not ok:
        raise HTTPException(status_code=400, detail="Folder ini bukan folder dinamis (dari FILE_MANAGER_ROOTS "
                                                       "di .env) atau tidak ditemukan.")
    log.info("ROOT REMOVE (web) label=%s user=%s", payload.label, user)
    audit_event(user,"ROOT_REMOVE",target=payload.label,ip=_client_ip(request)); asyncio.create_task(_to_thread(restart_file_watchers))
    return {"status": "success"}


@app.post("/api/rename")
async def api_rename(payload: RenameRequest, request: Request, user: str = Depends(require_admin)):
    try:
        new_path = fm_rename(payload.root, payload.path, payload.new_name)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    log.info("RENAME (web) root=%s path=%s -> %s user=%s", payload.root, payload.path, new_path.name, user)
    audit_event(user,"RENAME",target=payload.path,detail=new_path.name,ip=_client_ip(request)); asyncio.create_task(_to_thread(rebuild_file_index,payload.root))
    return {"status": "success", "new_name": new_path.name}


@app.post("/api/delete")
async def api_delete(payload: DeleteRequest, request: Request, user: str = Depends(require_admin)):
    try:
        fm_delete(payload.root, payload.path)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    log.info("DELETE (web) root=%s path=%s user=%s", payload.root, payload.path, user)
    audit_event(user,"DELETE",target=payload.path,detail="root=%s"%payload.root,ip=_client_ip(request)); asyncio.create_task(_to_thread(rebuild_file_index,payload.root))
    return {"status": "success"}


@app.post("/api/mkdir")
async def api_mkdir(payload: MkdirRequest, request: Request, user: str = Depends(require_admin)):
    try:
        new_dir = fm_mkdir(payload.root, payload.path, payload.name)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    log.info("MKDIR (web) root=%s path=%s name=%s user=%s", payload.root, payload.path, payload.name, user)
    audit_event(user,"MKDIR",target=str(new_dir),ip=_client_ip(request)); asyncio.create_task(_to_thread(rebuild_file_index,payload.root))
    return {"status": "success", "name": new_dir.name}


@app.post("/api/upload")
async def api_upload(request: Request, root: str = Form(...), path: str = Form(""), file: UploadFile = File(...),
                      user: str = Depends(require_admin)):
    max_bytes = settings.max_upload_mb * 1024 * 1024
    chunks = []
    size = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        size += len(chunk)
        if size > max_bytes:
            raise HTTPException(status_code=400, detail=f"File terlalu besar. Maks {settings.max_upload_mb} MB.")
        chunks.append(chunk)
    try:
        dest = fm_save_upload(root, path, file.filename or "upload.bin", b"".join(chunks))
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    log.info("UPLOAD (web) -> %s user=%s", dest, user)
    audit_event(user,"UPLOAD",target=str(dest),detail="%s bytes"%size,ip=_client_ip(request)); asyncio.create_task(_to_thread(rebuild_file_index,root))
    return {"status": "success", "name": dest.name}


@app.get("/api/config")
async def api_config(user: str = Depends(require_login)):
    return {"page_size": settings.items_per_page}


@app.get("/api/local-printers")
async def api_local_printers(user: str = Depends(require_login)):
    available = local_print_available()
    printers = await _to_thread(local_list_printers) if available else []
    default = local_default_printer()
    details = await asyncio.gather(*(_to_thread(printer_status_details, p) for p in printers)) if printers else []
    return {"available": available, "printers": printers, "default_printer": default, "details": details}


@app.get("/api/printers/status")
async def api_printer_status(name: str, user: str = Depends(require_login)):
    printers = await _to_thread(local_list_printers)
    if name not in printers:
        raise HTTPException(status_code=404, detail="Printer tidak ditemukan.")
    return await _to_thread(printer_status_details, name)


@app.post("/api/printers/test")
async def api_printer_test(payload: dict, request: Request, user: str = Depends(require_admin)):
    printer = str(payload.get("printer") or local_default_printer()).strip()
    if not printer:
        raise HTTPException(status_code=400, detail="Pilih printer terlebih dahulu.")
    ok, msg, spool = await _to_thread(create_test_print, printer)
    if not ok:
        audit_event(user,"TEST_PRINT",target=printer,detail=msg or "gagal",success=False,ip=_client_ip(request)); raise HTTPException(status_code=500, detail=msg or "Test print gagal.")
    audit_event(user,"TEST_PRINT",target=printer,detail="spool=%s"%spool,ip=_client_ip(request))
    return {"success": True, "printer": printer, "spool_job_id": spool, "message": msg}


@app.post("/api/queue/pause")
async def api_queue_pause(request: Request, user: str = Depends(require_admin)):
    global _PRINT_QUEUE_PAUSED
    _PRINT_QUEUE_PAUSED = True; app_setting_set("queue_paused","1")
    log.warning("QUEUE PAUSED user=%s", user); audit_event(user,"QUEUE_PAUSE",ip=_client_ip(request)); await event_publish({"type":"system","action":"queue_pause"})
    return {"paused": True}


@app.post("/api/queue/resume")
async def api_queue_resume(request: Request, user: str = Depends(require_admin)):
    global _PRINT_QUEUE_PAUSED
    _PRINT_QUEUE_PAUSED = False; app_setting_set("queue_paused","0")
    log.info("QUEUE RESUMED user=%s", user); audit_event(user,"QUEUE_RESUME",ip=_client_ip(request)); await event_publish({"type":"system","action":"queue_resume"})
    return {"paused": False}


@app.post("/api/printers/default")
async def api_printer_default(payload: dict, request: Request, user: str = Depends(require_admin)):
    printer = str(payload.get("printer") or "").strip()
    printers = await _to_thread(local_list_printers)
    if printer not in printers:
        raise HTTPException(status_code=404, detail="Printer tidak ditemukan.")
    runtime_settings_update(default_printer=printer)
    log.info("DEFAULT PRINTER -> %s user=%s", printer, user); audit_event(user,"DEFAULT_PRINTER",target=printer,ip=_client_ip(request))
    return {"success": True, "default_printer": printer}


@app.get("/api/diagnostics")
async def api_diagnostics(user: str = Depends(require_admin)):
    printers = await _to_thread(local_list_printers)
    details = await asyncio.gather(*(_to_thread(printer_status_details, p) for p in printers)) if printers else []
    return {
        "version": APP_VERSION, "python": sys.version.split()[0], "platform": platform.platform(),
        "local_print_available": local_print_available(), "default_printer": local_default_printer(),
        "printers": details, "office": office_availability(), "preview_tools": preview_available_tools(),
        "database": await _to_thread(db_quick_check), "storage": await _to_thread(storage_status),
        "task_scheduler": startup_task_status(), "queue_paused": _PRINT_QUEUE_PAUSED,
        "active_job": _PRINT_WORKER_ACTIVE_JOB, "active_jobs_by_printer": dict(_ACTIVE_JOB_BY_PRINTER), "data_dir": str(settings.data_dir),
        "maintenance_mode": is_maintenance_mode(), "config_version": CONFIG_VERSION, "db_schema_version": int(app_setting_get("db_schema_version",str(DB_SCHEMA_VERSION)) or DB_SCHEMA_VERSION),
        "file_index_count": file_index_count(),
        "update": {"configured": bool(settings.update_url), "auto_apply": settings.auto_update_apply, "check_hours": settings.update_check_hours},
        "security": {"session_max_age_sec": settings.session_max_age_sec, "session_idle_sec": settings.session_idle_sec, "secure_cookie": settings.secure_cookie},
    }


@app.post("/api/restart")
async def api_restart(request: Request, user: str = Depends(require_admin)):
    if not sys.platform.startswith("win"): raise HTTPException(status_code=400,detail="Restart otomatis hanya tersedia di Windows.")
    result=await _to_thread(schedule_windows_task_restart,3)
    if not result.get("ok"): raise HTTPException(status_code=500,detail=result.get("message") or "Gagal menjadwalkan restart.")
    audit_event(user,"RESTART",detail=result.get("message") or "",ip=_client_ip(request))
    async def _exit_later():
        await asyncio.sleep(0.8); os._exit(0)
    asyncio.create_task(_exit_later()); return result


@app.post("/api/task/repair")
async def api_task_repair(request: Request, user: str = Depends(require_admin)):
    result = await _to_thread(ensure_windows_startup_task, True)
    if not result.get("ok"):
        audit_event(user,"TASK_REPAIR",detail=str(result.get("message") or "gagal"),success=False,ip=_client_ip(request)); raise HTTPException(status_code=500, detail=str(result.get("message") or "Task Scheduler gagal."))
    audit_event(user,"TASK_REPAIR",detail=str(result.get("message") or "ok"),ip=_client_ip(request)); return result


@app.post("/api/backup")
async def api_backup(request: Request, user: str = Depends(require_admin)):
    try:
        out = await _to_thread(create_backup_archive)
        audit_event(user,"BACKUP",target=out.name,detail="%s bytes"%out.stat().st_size,ip=_client_ip(request))
        return {"success": True, "name": out.name, "size": out.stat().st_size}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/maintenance/cleanup")
async def api_cleanup(request: Request, user: str = Depends(require_admin)):
    previews = await _to_thread(cleanup_preview_files)
    backups = await _to_thread(cleanup_old_backups)
    audit_event(user,"CLEANUP",detail="previews=%s backups=%s"%(previews,backups),ip=_client_ip(request))
    return {"success": True, "previews_removed": previews, "backups_removed": backups}


@app.get("/api/stats")
async def api_stats(user: str = Depends(require_login)):
    return job_stats(top_n=10)


@app.get("/api/browse")
async def api_browse(root: str, path: str = "", page: int = 0, sort: str = "name",
                      user: str = Depends(require_login)):
    try:
        return await _to_thread(fm_list_dir, root, path, sort, page)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/search")
async def api_search(q: str, root: Optional[str] = None, user: str = Depends(require_login)):
    try:
        results = await _to_thread(fm_search, root, q)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"query": q, "results": results, "count": len(results)}


@app.get("/api/file-info")
async def api_file_info(root: str, path: str, user: str = Depends(require_login)):
    try:
        p = fm_resolve(root, path)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not p.is_file():
        raise HTTPException(status_code=404, detail="File tidak ditemukan.")
    st = p.stat()
    ext = p.suffix.lower()
    info = {
        "root": root, "rel_path": path, "name": p.name,
        "size": st.st_size, "modified": st.st_mtime, "ext": ext,
        "printable": ext in PRINTABLE_EXTS,
        "is_spreadsheet": ext in {".xls", ".xlsx", ".xlsm"},
    }
    if info["is_spreadsheet"]:
        info["sheets"] = fm_get_excel_sheet_names(p)
    return info


@app.get("/api/download")
async def api_download(root: str, path: str, user: str = Depends(require_login)):
    try:
        abs_path = fm_resolve(root, path)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not abs_path.is_file():
        raise HTTPException(status_code=404, detail="File tidak ditemukan.")
    size_mb = abs_path.stat().st_size / (1024 * 1024)
    if size_mb > settings.max_download_mb:
        raise HTTPException(status_code=413, detail=f"File terlalu besar ({size_mb:.1f} MB).")
    recent_touch(user,root,path,"download")
    return FileResponse(abs_path, filename=abs_path.name)


@app.get("/api/preview")
async def api_preview(root: str, path: str, user: str = Depends(require_login)):
    try:
        abs_path = fm_resolve(root, path)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not abs_path.is_file():
        raise HTTPException(status_code=404, detail="File tidak ditemukan.")
    jpg = await _to_thread(generate_preview, abs_path)
    if not jpg or not jpg.is_file():
        raise HTTPException(status_code=422, detail="Preview tidak tersedia (perlu LibreOffice / poppler-utils).")
    return FileResponse(
        jpg, media_type="image/jpeg",
        headers={"Cache-Control": "no-store, max-age=0", "Pragma": "no-cache"},
        background=BackgroundTask(lambda: jpg.unlink(missing_ok=True)),
    )


@app.post("/api/device-print/upload")
async def api_device_print_upload(request: Request, file: UploadFile = File(...), user: str = Depends(require_login)):
    """Stage file dari storage perangkat/browser tanpa memasukkannya ke File Manager."""
    original = Path(file.filename or "").name.strip()
    ext = Path(original).suffix.lower()
    if not original or ext not in PRINTABLE_EXTS or is_upload_blocked(original):
        raise HTTPException(status_code=422, detail="Format file tidak didukung untuk print.")
    token = uuid.uuid4().hex; root = _device_print_dir(); dest = root / (token + ext); meta_path = root / (token + ".json")
    max_bytes = settings.max_upload_mb * 1024 * 1024; size = 0
    try:
        with dest.open("wb") as fh:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk: break
                size += len(chunk)
                if size > max_bytes:
                    raise HTTPException(status_code=413, detail="File terlalu besar. Maks %s MB." % settings.max_upload_mb)
                fh.write(chunk)
        meta = {"token":token,"name":original,"ext":ext,"size":size,"created_at":time.time(),"username":user}
        tmp = meta_path.with_suffix(".json.tmp"); tmp.write_text(json.dumps(meta,ensure_ascii=False,indent=2),encoding="utf-8"); os.replace(str(tmp),str(meta_path))
        sheets = await _to_thread(fm_get_excel_sheet_names, dest) if ext in {".xls", ".xlsx", ".xlsm"} else []
        audit_event(user,"DEVICE_PRINT_STAGE",target=original,detail="%s bytes"%size,ip=_client_ip(request))
        return {"success":True,"token":token,"name":original,"ext":ext,"size":size,"sheets":sheets,"expires_in":DEVICE_PRINT_UPLOAD_TTL_SEC}
    except HTTPException:
        try: dest.unlink()
        except OSError: pass
        raise
    except Exception as exc:
        try: dest.unlink()
        except OSError: pass
        record_error("DEVICE_UPLOAD",str(exc),original); raise HTTPException(status_code=500,detail="Gagal menyiapkan file perangkat: %s"%exc)


@app.post("/api/device-print/queue")
async def api_device_print_queue(payload: DevicePrintRequest, request: Request, user: str = Depends(require_login)):
    if is_maintenance_mode():
        raise HTTPException(status_code=503, detail="Maintenance Mode aktif. Print baru sementara dinonaktifkan.")
    is_admin = request.session.get("role") == ROLE_ADMIN
    try: enforce_print_quota(user, payload.copies, is_admin)
    except FileManagerError as exc: raise HTTPException(status_code=429, detail=str(exc))
    if payload.scheduled_at is not None:
        if payload.scheduled_at < time.time() - 30: raise HTTPException(status_code=400, detail="Jadwal print sudah lewat.")
        if payload.scheduled_at > time.time() + 366 * 86400: raise HTTPException(status_code=400, detail="Jadwal print terlalu jauh.")
    try: staged, meta = device_print_resolve(payload.token)
    except FileManagerError as exc: raise HTTPException(status_code=410,detail=str(exc))
    if str(meta.get("username") or "") not in ("", user) and not is_admin:
        raise HTTPException(status_code=403,detail="File perangkat ini milik sesi/user lain.")
    if not local_print_available(): raise HTTPException(status_code=503,detail="Engine printer lokal tidak tersedia. Pastikan pywin32/driver printer terpasang.")
    printers=await _to_thread(local_list_printers); printer=(payload.printer or "").strip() or local_default_printer()
    if not printer: raise HTTPException(status_code=400,detail="Tidak ada default printer. Pilih printer lokal terlebih dahulu.")
    if printers and printer not in printers: raise HTTPException(status_code=404,detail="Printer yang dipilih tidak ditemukan.")
    status=await _to_thread(printer_status_details,printer)
    if status.get("problem"): raise HTTPException(status_code=409,detail="Printer belum siap: %s"%status.get("message","Unknown"))
    paper=(payload.paper or "AUTO").upper(); orientation=(payload.orientation or "AUTO").upper(); scale_mode=(payload.scale_mode or "FIT").upper(); duplex=(payload.duplex or "DEFAULT").upper(); color_mode=(payload.color_mode or "DEFAULT").upper()
    if paper not in {"AUTO","A4","F4","LEGAL","LETTER"}: raise HTTPException(status_code=400,detail="Ukuran kertas tidak valid.")
    if orientation not in {"AUTO","PORTRAIT","LANDSCAPE"}: raise HTTPException(status_code=400,detail="Orientasi tidak valid.")
    if scale_mode not in {"FIT","ACTUAL"}: raise HTTPException(status_code=400,detail="Skala print tidak valid.")
    if duplex not in {"DEFAULT","OFF","LONG","SHORT"}: raise HTTPException(status_code=400,detail="Mode duplex tidak valid.")
    if color_mode not in {"DEFAULT","COLOR","MONO"}: raise HTTPException(status_code=400,detail="Mode warna tidak valid.")
    try: _parse_page_numbers(payload.pages)
    except ValueError as exc: raise HTTPException(status_code=400,detail=str(exc))
    try: archive_path=await _to_thread(archive_print_source,staged)
    except OSError as exc: raise HTTPException(status_code=500,detail="Gagal membuat snapshot print: %s"%exc)
    original_name=str(meta.get("name") or staged.name)
    try:
        job=job_create(source="webapp-device",sender=user,sender_ref="web:%s"%user,file_path=str(staged),file_name=original_name,label=original_name,
            copies=payload.copies,pages=payload.pages,sheet=payload.sheet,method=PrintMethod.DIRECT,printer=printer,archive_path=str(archive_path),
            paper=paper,orientation=orientation,scale_mode=scale_mode,duplex=duplex,color_mode=color_mode,priority=payload.priority if is_admin else 0,scheduled_at=payload.scheduled_at)
    except Exception:
        try: archive_path.unlink()
        except OSError: pass
        raise
    await _to_thread(device_print_remove,payload.token)
    audit_event(user,"DEVICE_PRINT_QUEUE",target=job["id"],detail="file=%s copies=%s printer=%s"%(original_name,payload.copies,printer),ip=_client_ip(request))
    await event_publish({"type":"job","action":"queued","job":job}); return job


@app.get("/api/device-print/preview/meta")
async def api_device_print_preview_meta(token: str, user: str = Depends(require_login)):
    try: p, meta=device_print_resolve(token)
    except FileManagerError as exc: raise HTTPException(status_code=410,detail=str(exc))
    return await _to_thread(preview_document_meta,p)


@app.get("/api/device-print/preview/page")
async def api_device_print_preview_page(token: str, page: int = 1, user: str = Depends(require_login)):
    try: p, meta=device_print_resolve(token)
    except FileManagerError as exc: raise HTTPException(status_code=410,detail=str(exc))
    img=await _to_thread(generate_preview_page,p,max(1,page))
    if not img or not img.is_file(): raise HTTPException(status_code=422,detail="Halaman preview tidak tersedia.")
    return FileResponse(img,media_type="image/jpeg",headers={"Cache-Control":"no-store"},background=BackgroundTask(lambda: img.unlink(missing_ok=True)))


@app.post("/api/print")
async def api_print(payload: PrintRequest, request: Request, user: str = Depends(require_login)):
    if is_maintenance_mode():
        raise HTTPException(status_code=503, detail="Maintenance Mode aktif. Print baru sementara dinonaktifkan.")
    is_admin = request.session.get("role") == ROLE_ADMIN
    try:
        enforce_print_quota(user, payload.copies, is_admin)
    except FileManagerError as exc:
        raise HTTPException(status_code=429, detail=str(exc))
    if payload.scheduled_at is not None:
        if payload.scheduled_at < time.time() - 30:
            raise HTTPException(status_code=400, detail="Jadwal print sudah lewat.")
        if payload.scheduled_at > time.time() + 366 * 86400:
            raise HTTPException(status_code=400, detail="Jadwal print terlalu jauh.")
    try:
        abs_path = fm_resolve(payload.root, payload.path)
    except FileManagerError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not abs_path.is_file():
        raise HTTPException(status_code=404, detail="File tidak ditemukan.")
    if abs_path.suffix.lower() not in PRINTABLE_EXTS:
        raise HTTPException(status_code=422, detail="Format file tidak didukung untuk print.")
    if not local_print_available():
        raise HTTPException(status_code=503, detail="Engine printer lokal tidak tersedia. Pastikan pywin32/driver printer terpasang.")

    printers = await _to_thread(local_list_printers)
    printer = (payload.printer or "").strip() or local_default_printer()
    if not printer:
        raise HTTPException(status_code=400, detail="Tidak ada default printer. Pilih printer lokal terlebih dahulu.")
    if printers and printer not in printers:
        raise HTTPException(status_code=404, detail="Printer yang dipilih tidak ditemukan.")
    status = await _to_thread(printer_status_details, printer)
    if status.get("problem"):
        raise HTTPException(status_code=409, detail="Printer belum siap: %s" % status.get("message", "Unknown"))

    paper = (payload.paper or "AUTO").upper()
    orientation = (payload.orientation or "AUTO").upper()
    scale_mode = (payload.scale_mode or "FIT").upper()
    duplex = (payload.duplex or "DEFAULT").upper()
    color_mode = (payload.color_mode or "DEFAULT").upper()
    if paper not in {"AUTO", "A4", "F4", "LEGAL", "LETTER"}:
        raise HTTPException(status_code=400, detail="Ukuran kertas tidak valid.")
    if orientation not in {"AUTO", "PORTRAIT", "LANDSCAPE"}:
        raise HTTPException(status_code=400, detail="Orientasi tidak valid.")
    if scale_mode not in {"FIT", "ACTUAL"}:
        raise HTTPException(status_code=400, detail="Skala print tidak valid.")
    if duplex not in {"DEFAULT", "OFF", "LONG", "SHORT"}:
        raise HTTPException(status_code=400, detail="Mode duplex tidak valid.")
    if color_mode not in {"DEFAULT", "COLOR", "MONO"}:
        raise HTTPException(status_code=400, detail="Mode warna tidak valid.")
    try:
        _parse_page_numbers(payload.pages)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    # Snapshot khusus riwayat/reprint. File sumber asli tidak dipindah/diubah.
    try:
        archive_path = await _to_thread(archive_print_source, abs_path)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Gagal membuat snapshot print: {exc}")

    job = job_create(
        source="webapp", sender=user, sender_ref=f"web:{user}",
        file_path=str(abs_path), file_name=abs_path.name, label=abs_path.name,
        copies=payload.copies, pages=payload.pages, sheet=payload.sheet,
        method=PrintMethod.DIRECT, printer=printer, archive_path=str(archive_path),
        paper=paper, orientation=orientation, scale_mode=scale_mode, duplex=duplex,
        color_mode=color_mode, priority=payload.priority if is_admin else 0, scheduled_at=payload.scheduled_at,
    )
    log.info("PRINT QUEUED (web) job=%s file=%s copies=%s printer=%s user=%s",
             job["id"], abs_path.name, payload.copies, printer, user)
    recent_touch(user,payload.root,payload.path,"print")
    audit_event(user,"PRINT_QUEUE",target=job["id"],detail="file=%s copies=%s printer=%s scheduled=%s"%(abs_path.name,payload.copies,printer,payload.scheduled_at or "now"),ip=_client_ip(request))
    await event_publish({"type":"job","action":"queued","job":job})
    return job


@app.get("/api/jobs")
async def api_jobs(request: Request, status: Optional[str] = None, limit: int = 50, user: str = Depends(require_login)):
    rows=job_list(status=status,limit=limit) if request.session.get("role")==ROLE_ADMIN else job_list_for_sender(user,status=status,limit=limit)
    return {"jobs":rows}


@app.get("/api/jobs/{job_id}")
async def api_job_get(job_id: str, request: Request, user: str = Depends(require_login)):
    job = job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job tidak ditemukan.")
    if request.session.get("role")!=ROLE_ADMIN and job.get("sender")!=user:
        raise HTTPException(status_code=403,detail="Anda tidak boleh melihat job pengguna lain.")
    return job


@app.post("/api/jobs/{job_id}/cancel")
async def api_job_cancel(job_id: str, request: Request, user: str = Depends(require_login)):
    job = job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job tidak ditemukan.")
    if request.session.get("role") != ROLE_ADMIN and job.get("sender") != user:
        raise HTTPException(status_code=403, detail="Anda tidak boleh membatalkan job pengguna lain.")
    if job.get("status") not in (JobStatus.QUEUED, JobStatus.PROCESSING, JobStatus.PRINTING):
        raise HTTPException(status_code=400, detail="Job sudah selesai/gagal dan tidak dapat dibatalkan.")

    if job.get("status") == JobStatus.QUEUED:
        if not job_cancel(job_id):
            raise HTTPException(status_code=409, detail="Job keburu diproses. Muat ulang status lalu coba lagi.")
        log.info("PRINT CANCEL QUEUED job=%s user=%s", job_id, user)
        audit_event(user,"PRINT_CANCEL",target=job_id,detail="queued",ip=_client_ip(request)); await event_publish({"type":"job","action":"cancelled","job_id":job_id})
        return {"status": "success", "job_id": job_id, "message": "Job dibatalkan sebelum masuk printer."}

    # Tandai event worker agar jika spool ID baru muncul setelah request cancel,
    # callback akan langsung membatalkan job Windows tersebut.
    ev = _ACTIVE_CANCEL_EVENTS.get(job_id)
    if ev is not None:
        ev.set()
    spool_id = int(job.get("spool_job_id") or 0)
    spool_msg = "Pembatalan diminta sebelum Windows Job ID tersedia."
    if spool_id:
        ok, spool_msg = await _to_thread(cancel_windows_spool_job, job.get("printer") or local_default_printer(), spool_id)
        if not ok and "tidak ditemukan" not in spool_msg.lower():
            raise HTTPException(status_code=500, detail=spool_msg)
    with _db_lock, _db_connect() as conn:
        conn.execute("UPDATE jobs SET status=?, error=?, finished_at=?, cancelled_at=?, spool_cancelled=1, spool_status=? WHERE id=?",
                     (JobStatus.FAILED, "Dibatalkan pengguna.", time.time(), time.time(), "Cancelled", job_id)); conn.commit()
    log.info("PRINT CANCEL ACTIVE job=%s spool=%s user=%s", job_id, spool_id or "-", user)
    audit_event(user,"PRINT_CANCEL",target=job_id,detail="spool=%s"%(spool_id or 0),ip=_client_ip(request)); await event_publish({"type":"job","action":"cancelled","job_id":job_id,"spool_job_id":spool_id})
    return {"status": "success", "job_id": job_id, "message": spool_msg}


@app.post("/api/jobs/{job_id}/reprint")
async def api_job_reprint(job_id: str, request: Request, user: str = Depends(require_login)):
    if is_maintenance_mode(): raise HTTPException(status_code=503,detail="Maintenance Mode aktif.")
    original = job_get(job_id)
    if not original:
        raise HTTPException(status_code=404, detail="Riwayat print tidak ditemukan.")
    is_admin=request.session.get("role") == ROLE_ADMIN
    if not is_admin and original.get("sender") != user:
        raise HTTPException(status_code=403, detail="Anda tidak boleh print ulang file pengguna lain.")
    try: enforce_print_quota(user,int(original.get("copies") or 1),is_admin)
    except FileManagerError as exc: raise HTTPException(status_code=429,detail=str(exc))
    archive_path = resolve_print_archive(original.get("archive_path") or "")
    original_path = Path(original.get("file_path") or "")
    source = archive_path or (original_path if original_path.is_file() else None)
    if not source:
        raise HTTPException(status_code=410, detail="File sumber/snapshot print sudah tidak tersedia.")
    printer = original.get("printer") or local_default_printer()
    if not printer:
        raise HTTPException(status_code=400, detail="Printer untuk job lama tidak tersedia.")
    # Buat snapshot baru supaya lifecycle reprint tidak tergantung snapshot job lama.
    try:
        new_archive = await _to_thread(archive_print_source, source)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Gagal membuat snapshot reprint: {exc}")
    job = job_create(
        source="webapp-reprint", sender=user, sender_ref=f"web:{user}",
        file_path=original.get("file_path") or str(source), file_name=original.get("file_name") or source.name,
        label=original.get("label") or original.get("file_name") or source.name,
        copies=original.get("copies") or 1, pages=original.get("pages") or "", sheet=original.get("sheet") or "",
        method=PrintMethod.DIRECT, printer=printer, reprint_of=original.get("reprint_of") or original["id"],
        archive_path=str(new_archive), paper=original.get("paper") or "AUTO", orientation=original.get("orientation") or "AUTO",
        scale_mode=original.get("scale_mode") or "FIT", duplex=original.get("duplex") or "DEFAULT",
        color_mode=original.get("color_mode") or "DEFAULT", priority=(original.get("priority") or 0) if is_admin else 0,
    )
    log.info("REPRINT LOCAL original=%s new=%s file=%s user=%s", job_id, job["id"], original.get("file_name"), user)
    audit_event(user,"REPRINT",target=job["id"],detail="original=%s"%job_id,ip=_client_ip(request)); await event_publish({"type":"job","action":"queued","job":job})
    return job


@app.get("/api/status")
async def api_status(user: str = Depends(require_login)):
    available = local_print_available()
    default = local_default_printer()
    default_status = await _to_thread(printer_status_details, default) if default else None
    return {
        "counts": job_counts(),
        "preview_tools": preview_available_tools(),
        "roots": get_roots(),
        "local_print_available": available,
        "printer_system": {
            "available": available, "default_printer": default, "default_status": default_status,
            "queue_paused": _PRINT_QUEUE_PAUSED, "active_job": _PRINT_WORKER_ACTIVE_JOB,
        },
    }


# ============================================================================
# WINDOWS AUTOSTART / SINGLE INSTANCE
# ============================================================================
_INSTANCE_MUTEX_HANDLE = None


def acquire_single_instance() -> bool:
    global _INSTANCE_MUTEX_HANDLE
    if _INSTANCE_MUTEX_HANDLE is not None:
        return True
    if not sys.platform.startswith("win"):
        return True
    try:
        import ctypes
        name = "Local\\PrintBotWeb_" + uuid.uuid5(uuid.NAMESPACE_URL, str(BASE_DIR).lower()).hex
        handle = ctypes.windll.kernel32.CreateMutexW(None, False, name)
        if not handle: return True
        exists = int(ctypes.windll.kernel32.GetLastError()) == 183
        _INSTANCE_MUTEX_HANDLE = handle
        return not exists
    except Exception as exc:
        log.warning("Single-instance mutex gagal: %s", exc); return True


def _pythonw_path() -> Path:
    exe = Path(sys.executable).resolve()
    candidate = exe.with_name("pythonw.exe")
    return candidate if sys.platform.startswith("win") and candidate.is_file() else exe


def _startup_task_action() -> str:
    return '"%s" "%s"' % (str(_pythonw_path()), str(Path(__file__).resolve()))


def _task_creation_flags() -> int:
    if not sys.platform.startswith("win"):
        return 0
    return int(getattr(subprocess, "CREATE_NO_WINDOW", 0) or 0)


def _current_user_sid() -> str:
    """Ambil SID user interaktif saat ini tanpa meminta password."""
    if not sys.platform.startswith("win"):
        return ""
    try:
        import win32api, win32con, win32security
        token = win32security.OpenProcessToken(win32api.GetCurrentProcess(), win32con.TOKEN_QUERY)
        sid = win32security.GetTokenInformation(token, win32security.TokenUser)[0]
        return str(win32security.ConvertSidToStringSid(sid) or "")
    except Exception:
        pass
    try:
        r = subprocess.run(["whoami", "/user", "/fo", "csv", "/nh"],
                           stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                           universal_newlines=True, timeout=10,
                           creationflags=_task_creation_flags())
        if r.returncode == 0 and r.stdout.strip():
            import csv as _csv
            row = next(_csv.reader([r.stdout.strip()]))
            if len(row) >= 2:
                return str(row[1]).strip()
    except Exception:
        pass
    return ""


def startup_task_status() -> dict:
    """Baca Task Scheduler via schtasks.exe saja.

    COM sengaja tidak dipakai di jalur startup/status karena pada beberapa Windows
    account lokal/Microsoft Account COM RegisterTaskDefinition dapat menghasilkan
    0x8007052E dan noise `releasing IUnknown` walaupun schtasks bekerja normal.
    """
    if not sys.platform.startswith("win"):
        return {"supported": False, "exists": False, "ok": False, "message": "Non-Windows"}
    try:
        r = subprocess.run(["schtasks", "/Query", "/TN", settings.task_name, "/XML"],
                           stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                           universal_newlines=True, timeout=20,
                           creationflags=_task_creation_flags())
        xml = r.stdout or ""
        exists = r.returncode == 0
        script = str(Path(__file__).resolve()).lower()
        pythonw = str(_pythonw_path()).lower()
        xml_l = xml.lower()
        action_ok = script in xml_l and pythonw in xml_l
        enabled = "<enabled>false</enabled>" not in xml_l if exists else False
        return {
            "supported": True,
            "exists": exists,
            "ok": bool(exists and enabled and action_ok),
            "enabled": enabled,
            "action_ok": action_ok,
            "message": "Task Scheduler native (schtasks/XML)" if exists else ((r.stderr or "Task belum ada").strip()[-800:]),
            "action": _startup_task_action(),
        }
    except Exception as exc:
        return {"supported": True, "exists": False, "ok": False, "message": str(exc), "action": _startup_task_action()}


def _create_startup_task_xml() -> dict:
    """Buat task ONLOGON dengan InteractiveToken melalui XML + schtasks.

    Tidak meminta/menyimpan password. Task berjalan pada sesi user yang sedang
    login sehingga printer per-user dan Microsoft Office COM tetap tersedia.
    """
    if not sys.platform.startswith("win"):
        return {"ok": True, "created": False, "message": "Non-Windows"}
    sid = _current_user_sid()
    if not sid:
        raise RuntimeError("SID user Windows tidak dapat ditentukan.")
    from xml.sax.saxutils import escape as _xml_escape
    command = _xml_escape(str(_pythonw_path()))
    script = _xml_escape('"%s"' % str(Path(__file__).resolve()))
    workdir = _xml_escape(str(BASE_DIR))
    sid_xml = _xml_escape(sid)
    xml = f'''<?xml version="1.0" encoding="UTF-16"?>
<Task version="1.2" xmlns="http://schemas.microsoft.com/windows/2004/02/mit/task">
  <RegistrationInfo>
    <Description>PrintBot WebApp - local printer web console</Description>
  </RegistrationInfo>
  <Triggers>
    <LogonTrigger>
      <Enabled>true</Enabled>
      <UserId>{sid_xml}</UserId>
    </LogonTrigger>
  </Triggers>
  <Principals>
    <Principal id="Author">
      <UserId>{sid_xml}</UserId>
      <LogonType>InteractiveToken</LogonType>
      <RunLevel>HighestAvailable</RunLevel>
    </Principal>
  </Principals>
  <Settings>
    <MultipleInstancesPolicy>IgnoreNew</MultipleInstancesPolicy>
    <DisallowStartIfOnBatteries>false</DisallowStartIfOnBatteries>
    <StopIfGoingOnBatteries>false</StopIfGoingOnBatteries>
    <AllowHardTerminate>true</AllowHardTerminate>
    <StartWhenAvailable>true</StartWhenAvailable>
    <RunOnlyIfNetworkAvailable>false</RunOnlyIfNetworkAvailable>
    <IdleSettings>
      <StopOnIdleEnd>false</StopOnIdleEnd>
      <RestartOnIdle>false</RestartOnIdle>
    </IdleSettings>
    <AllowStartOnDemand>true</AllowStartOnDemand>
    <Enabled>true</Enabled>
    <Hidden>false</Hidden>
    <RunOnlyIfIdle>false</RunOnlyIfIdle>
    <WakeToRun>false</WakeToRun>
    <ExecutionTimeLimit>PT0S</ExecutionTimeLimit>
    <Priority>7</Priority>
    <RestartOnFailure>
      <Interval>PT1M</Interval>
      <Count>999</Count>
    </RestartOnFailure>
  </Settings>
  <Actions Context="Author">
    <Exec>
      <Command>{command}</Command>
      <Arguments>{script}</Arguments>
      <WorkingDirectory>{workdir}</WorkingDirectory>
    </Exec>
  </Actions>
</Task>
'''
    tmp_path = None
    try:
        fd, raw_path = tempfile.mkstemp(prefix="printbot_task_", suffix=".xml")
        os.close(fd)
        tmp_path = Path(raw_path)
        tmp_path.write_text(xml, encoding="utf-16")
        r = subprocess.run(["schtasks", "/Create", "/TN", settings.task_name,
                            "/XML", str(tmp_path), "/F"],
                           stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                           universal_newlines=True, timeout=30,
                           creationflags=_task_creation_flags())
        msg = ((r.stdout or "") + " " + (r.stderr or "")).strip()
        if r.returncode != 0:
            raise RuntimeError(msg or "schtasks /XML gagal")
        st = startup_task_status()
        if not st.get("ok"):
            raise RuntimeError("Task dibuat tetapi verifikasi action gagal: %s" % st.get("message", "unknown"))
        return {"ok": True, "created": True, "message": msg or "Task dibuat via schtasks/XML", "action": _startup_task_action()}
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _create_startup_task_basic() -> dict:
    """Fallback paling kompatibel jika XML ditolak oleh instalasi Windows tertentu."""
    action = _startup_task_action()
    base = ["schtasks", "/Create", "/TN", settings.task_name, "/TR", action,
            "/SC", "ONLOGON", "/F"]
    last = ""
    for cmd in (base + ["/RL", "HIGHEST"], base):
        try:
            r = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                               universal_newlines=True, timeout=30,
                               creationflags=_task_creation_flags())
            last = ((r.stdout or "") + " " + (r.stderr or "")).strip()
            if r.returncode == 0:
                return {"ok": True, "created": True,
                        "message": last or "Task dibuat via schtasks", "action": action}
        except Exception as exc:
            last = str(exc)
    return {"ok": False, "created": False, "message": last, "action": action}


def ensure_windows_startup_task(force: bool = False) -> dict:
    if not sys.platform.startswith("win"):
        return {"ok": True, "created": False, "message": "Non-Windows"}
    if not settings.auto_start_task and not force:
        return {"ok": True, "created": False, "message": "AUTO_START_TASK nonaktif"}
    current = startup_task_status()
    if current.get("ok") and not force:
        return {"ok": True, "created": False, "message": "Task sudah aktif", "action": _startup_task_action()}
    try:
        result = _create_startup_task_xml()
        log.info("Task Scheduler '%s' siap via native schtasks/XML.", settings.task_name)
        return result
    except Exception as exc:
        log.warning("Task Scheduler XML gagal, fallback schtasks basic: %s", exc)
        return _create_startup_task_basic()

def schedule_windows_task_restart(delay_sec: int = 3) -> dict:
    if not sys.platform.startswith("win"):
        return {"ok":False,"message":"Restart otomatis via Task Scheduler hanya tersedia di Windows."}
    ready=ensure_windows_startup_task(False)
    if not ready.get("ok"): return ready
    wait=max(2,min(int(delay_sec),30)); cmd='ping 127.0.0.1 -n %s >nul & schtasks /Run /TN "%s"' % (wait+1,settings.task_name.replace('"',''))
    flags=getattr(subprocess,"CREATE_NO_WINDOW",0)|getattr(subprocess,"DETACHED_PROCESS",0)
    subprocess.Popen(["cmd.exe","/d","/c",cmd],creationflags=flags,close_fds=True)
    return {"ok":True,"message":"Restart dijadwalkan melalui Task Scheduler.","delay_sec":wait}


# ============================================================================
# ENTRYPOINT
# ============================================================================
def main() -> None:
    global _TASK_ENSURED_AT_STARTUP
    import uvicorn
    errs = settings.validate()
    if errs:
        for e in errs: print(f"[CONFIG ERROR] {e}")
        sys.exit(1)
    if not acquire_single_instance():
        print("[INFO] PrintBot WebApp sudah berjalan. Instance kedua dihentikan.")
        return
    if sys.platform.startswith("win") and settings.auto_start_task:
        result = ensure_windows_startup_task(False)
        _TASK_ENSURED_AT_STARTUP = True
        if not result.get("ok"):
            log.warning("Task Scheduler belum siap: %s", result.get("message"))
    print("=" * 62)
    print(" PrintBot WebApp - LOCAL PRINTER v%s" % APP_VERSION)
    print(" URL     : http://127.0.0.1:%s" % settings.webapp_port)
    print(" Printer : %s" % (local_default_printer() or "belum ada/default belum diset"))
    print(" Queue   : persistent SQLite / sequential worker")
    print(" Bridge  : DISABLED")
    print("=" * 62)
    uvicorn.run(app, host=settings.webapp_host, port=settings.webapp_port, log_level="info", access_log=False, use_colors=False)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        pass
    except Exception as exc:
        try: record_error("STARTUP",str(exc))
        except Exception: pass
        logging.getLogger("webapp").exception("Fatal startup error")
        raise
